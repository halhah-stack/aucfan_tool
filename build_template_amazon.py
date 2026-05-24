"""
build_template_amazon.py — AucFan + Amazon 統合リサーチシート テンプレート生成（2シート構成）

【シート構成】
  シート1「①リサーチ」:
    A〜L列: AucFanデータ（元テンプレートと同じ）
    M〜Q列: Amazonデータ（アプリが自動入力: ASIN・現在価格・評価・レビュー件数・A+フラグ）
    R〜Y列: 手入力・利益計算（ライバル件数・仕入値・販売価格・FBA手数料・利益・利益率・メモ）

  シート2「②Amazon詳細」:
    縦並びでAmazon全テキストデータを表示（タイトル・説明文・仕様文など）
    各項目が独立したセルなのでクリック1つでコピー可能

実行方法:
  cd ~/Downloads/aucfan_tool
  python3 build_template_amazon.py

出力ファイル: リサーチ_テンプレート_Amazon.xlsx
"""

from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_PATH = Path(__file__).parent / "リサーチ_テンプレート_Amazon.xlsx"

# ── フォント・スタイル定数 ──────────────────────────────────────────────
FONT_NAME = "BIZ UDGothic"
FONT_SIZE = 11

# シート1 色定数
HDR_BG  = "1F4E79"  # タイトルバー（紺）
HDR_FG  = "FFFFFF"
SUB_BG  = "2E75B6"  # AucFan列ヘッダー（青）
SUB_FG  = "FFFFFF"
AMZ_BG  = "375623"  # Amazon列ヘッダー（緑）
AMZ_FG  = "FFFFFF"
LBL_BG  = "D6E4F0"  # ラベル（薄青）
INP_BG  = "FFF9C4"  # AucFan手入力（黄）
FML_BG  = "E8F5E9"  # 自動計算（薄緑）
IMG_BG  = "F0F0F0"  # 画像エリア（薄灰）
AUTO_BG = "DBEAFE"  # Amazon自動入力（薄青）
MINP_BG = "FFF3E0"  # Amazon手入力（薄橙）
MFML_BG = "E8F5E9"  # Amazon自動計算（薄緑）

# シート2 色定数
S2_HDR_BG  = "1F4E79"  # シート2タイトル
S2_LBL_BG  = "DBEAFE"  # 項目ラベル（薄青）
S2_VAL_BG  = "F9FAFB"  # 値セル（薄灰）
S2_SECT_BG = "375623"  # セクションヘッダー（緑）

DATA_ROW           = 6   # シート1データ行
EXCHANGE_RATE_DEFAULT = 22  # 初期為替レート（円/元）


# ── ヘルパー ──────────────────────────────────────────────────────────────
def fnt(bold=False, size=FONT_SIZE, color="000000"):
    return Font(name=FONT_NAME, size=size, bold=bold, color=color)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def aln(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def bdr(color="CCCCCC"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def cell(ws, row, col, value=None, font=None, bg=None,
         align=None, border=None, fmt=None):
    c = ws.cell(row=row, column=col)
    if value  is not None: c.value  = value
    if font:               c.font   = font
    if bg:                 c.fill   = fill(bg)
    if align:              c.alignment = align
    if border:             c.border = border
    if fmt:                c.number_format = fmt
    return c


# ─────────────────────────────────────────────────────────────────────────
# ワークブック生成
# ─────────────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ═══════════════════════════════════════════════════════════════════════════
# シート1「①リサーチ」
# ═══════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "①リサーチ"

# ── 列定義: (列記号, 幅, ヘッダーテキスト, 背景色, 書式, グループ) ──
# グループ: "aucfan" / "amz_auto"（自動入力）/ "amz_manual"（手入力・計算）
COLUMNS = [
    # AucFan列（元テンプレートと同じ）
    ("A",  4,  "No",                     None,    None,          "aucfan"),
    ("B",  16, "画像",                   IMG_BG,  None,          "aucfan"),
    ("C",  11, "日付",                   INP_BG,  "YYYY/MM/DD",  "aucfan"),
    ("D",  42, "代表キーワード",         INP_BG,  None,          "aucfan"),
    ("E",  12, "件数\n（直近30日）",    FML_BG,  "#,##0",       "aucfan"),
    ("F",  9,  "4件\n以上",             FML_BG,  None,          "aucfan"),
    ("G",  16, "最安値\n（価格＋送料）", FML_BG, "#,##0",       "aucfan"),
    ("H",  18, "セラーID①",             INP_BG,  None,          "aucfan"),
    ("I",  18, "セラーID②",             INP_BG,  None,          "aucfan"),
    ("J",  18, "セラーID③",             INP_BG,  None,          "aucfan"),
    ("K",  38, "検索URL\n（Aucfan）",   INP_BG,  None,          "aucfan"),
    ("L",  22, "備考",                   INP_BG,  None,          "aucfan"),
    # Amazon自動入力列（アプリが自動記入）
    ("M",  14, "ASIN\n（自動）",        AUTO_BG, None,          "amz_auto"),
    ("N",  14, "Amazon\n現在価格（自動）", AUTO_BG, "#,##0",    "amz_auto"),
    ("O",  10, "評価\n（自動）",        AUTO_BG, "0.0",         "amz_auto"),
    ("P",  14, "レビュー件数\n（自動）", AUTO_BG, None,         "amz_auto"),
    ("Q",  10, "A+\n（自動）",          AUTO_BG, None,          "amz_auto"),
    # 手入力・計算列
    ("R",  12, "ライバル\n件数",         MINP_BG, "#,##0",      "amz_manual"),
    ("S",  14, "1688\n仕入値（元）",    MINP_BG, "#,##0.00",   "amz_manual"),
    ("T",  14, "仕入値（円）\n自動計算", MFML_BG, "#,##0",     "amz_manual"),
    ("U",  14, "販売\n予定価格",         MINP_BG, "#,##0",      "amz_manual"),
    ("V",  14, "FBA\n手数料",            MINP_BG, "#,##0",      "amz_manual"),
    ("W",  14, "利益\n自動計算",         MFML_BG, "#,##0",      "amz_manual"),
    ("X",  10, "利益率\n自動計算",       MFML_BG, "0.0%",       "amz_manual"),
    ("Y",  24, "メモ",                   MINP_BG, None,          "amz_manual"),
]

LAST_COL = COLUMNS[-1][0]  # "Y"

# 列幅設定
for col_letter, width, *_ in COLUMNS:
    ws1.column_dimensions[col_letter].width = width

# ── Row 1: タイトルバー ────────────────────────────────────────────────
ws1.row_dimensions[1].height = 26
ws1.merge_cells(f"A1:{LAST_COL}1")
c = ws1["A1"]
c.value     = "Aucfan + Amazon 統合リサーチシート  ―  シート①リサーチ"
c.font      = Font(name=FONT_NAME, size=13, bold=True, color=HDR_FG)
c.fill      = fill(HDR_BG)
c.alignment = aln("center", "center")

# ── Row 2: サマリー・設定行 ────────────────────────────────────────────
ws1.row_dimensions[2].height = 22

cell(ws1, 2, 1, "調査商品数",
     font=fnt(bold=True), bg=LBL_BG, align=aln("center","center"), border=bdr())
cell(ws1, 2, 2, f"=COUNTA(D{DATA_ROW}:D{DATA_ROW})",
     font=fnt(bold=True), bg=FML_BG, align=aln("center","center"), border=bdr())
cell(ws1, 2, 3, "4件以上",
     font=fnt(bold=True), bg=LBL_BG, align=aln("center","center"), border=bdr())
cell(ws1, 2, 4, f'=COUNTIF(E{DATA_ROW}:E{DATA_ROW},">=4")',
     font=fnt(bold=True), bg=FML_BG, align=aln("center","center"), border=bdr())
cell(ws1, 2, 5, "セッション",
     font=fnt(bold=True), bg=LBL_BG, align=aln("center","center"), border=bdr())
ws1.merge_cells("F2:L2")
cell(ws1, 2, 6, "", font=Font(name=FONT_NAME, size=10, color="444444"),
     bg=INP_BG, align=aln("left","center"), border=bdr())

# 為替レート（Amazon仕入計算用）
cell(ws1, 2, 19, "為替レート\n（円/元）",
     font=fnt(bold=True, size=9), bg="FFF3E0",
     align=aln("center","center", wrap=True), border=bdr())
c_rate = ws1.cell(row=2, column=20)
c_rate.value         = EXCHANGE_RATE_DEFAULT
c_rate.font          = fnt(bold=True)
c_rate.fill          = fill(MFML_BG)
c_rate.alignment     = aln("center","center")
c_rate.border        = bdr()
c_rate.number_format = "#,##0.00"
ws1.merge_cells("U2:Y2")
cell(ws1, 2, 21, "← 為替レートをT2セルに入力（仕入値（円）列が自動計算されます）",
     font=Font(name=FONT_NAME, size=9, color="888888"),
     bg="FFFBF5", align=aln("left","center"), border=bdr())

# ── Row 3: 凡例 ───────────────────────────────────────────────────────
ws1.row_dimensions[3].height = 18
ws1.merge_cells(f"A3:{LAST_COL}3")
c = ws1["A3"]
c.value = (
    "  ■ 黄色 = 手入力（AucFan）　"
    "■ 薄青 = Amazon自動入力（アプリが記入）　"
    "■ 薄橙 = 手入力（Amazon）　"
    "■ 薄緑 = 自動計算　"
    "→ 詳細テキストはシート②を参照"
)
c.font      = Font(name=FONT_NAME, size=9, color="555555")
c.fill      = fill("FAFAFA")
c.alignment = aln("left", "center")

# ── Row 4: スペーサー ──────────────────────────────────────────────────
ws1.row_dimensions[4].height = 4

# ── Row 5: 列ヘッダー ─────────────────────────────────────────────────
ws1.row_dimensions[5].height = 40
for col_idx, (_, _, header, _, _, group) in enumerate(COLUMNS, 1):
    c = ws1.cell(row=5, column=col_idx)
    c.value = header
    if group == "aucfan":
        c.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True, color=SUB_FG)
        c.fill = fill(SUB_BG)
    else:  # amz_auto / amz_manual どちらも緑ヘッダー
        c.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True, color=AMZ_FG)
        c.fill = fill(AMZ_BG)
    c.alignment = aln("center", "center", wrap=True)
    c.border    = bdr()

ws1.freeze_panes = "A6"

# ── Row 6: データ行テンプレート ──────────────────────────────────────
ws1.row_dimensions[DATA_ROW].height = 72
for col_idx, (col_letter, _, _, bg_color, fmt_str, _) in enumerate(COLUMNS, 1):
    c = ws1.cell(row=DATA_ROW, column=col_idx)
    c.font      = fnt()
    c.alignment = aln("left", "center", wrap=(col_letter in ["D","L","Y"]))
    c.border    = bdr()
    if bg_color:
        c.fill = fill(bg_color)
    if fmt_str:
        c.number_format = fmt_str

# AucFan側
ws1.cell(DATA_ROW, 1).alignment = aln("center","center")
ws1.cell(DATA_ROW, 3).alignment = aln("center","center")
ws1.cell(DATA_ROW, 5).alignment = aln("center","center")
ws1.cell(DATA_ROW, 6).value     = f'=IF(E{DATA_ROW}="","",IF(E{DATA_ROW}>=4,"✓","✗"))'
ws1.cell(DATA_ROW, 6).font      = fnt(bold=True)
ws1.cell(DATA_ROW, 6).alignment = aln("center","center")
ws1.cell(DATA_ROW, 7).alignment = aln("right","center")
img_cell = ws1.cell(DATA_ROW, 2)
img_cell.value     = "画像\nここに貼付"
img_cell.font      = Font(name=FONT_NAME, size=9, color="AAAAAA")
img_cell.alignment = aln("center","center", wrap=True)

# Amazon自動入力側（M=13〜Q=17）
ws1.cell(DATA_ROW, 13).alignment = aln("center","center")   # ASIN
ws1.cell(DATA_ROW, 14).alignment = aln("right","center")    # 現在価格
ws1.cell(DATA_ROW, 15).alignment = aln("center","center")   # 評価
ws1.cell(DATA_ROW, 16).alignment = aln("center","center")   # レビュー件数
ws1.cell(DATA_ROW, 17).alignment = aln("center","center")   # A+

# 手入力・計算側（R=18〜Y=25）
ws1.cell(DATA_ROW, 18).alignment = aln("center","center")   # ライバル件数
ws1.cell(DATA_ROW, 19).alignment = aln("right","center")    # 1688仕入値（元）
# 仕入値（円）= 1688価格 × T2の為替レート（T=20）
ws1.cell(DATA_ROW, 20).value     = f"=IF(S{DATA_ROW}=\"\",\"\",ROUND(S{DATA_ROW}*$T$2,0))"
ws1.cell(DATA_ROW, 20).alignment = aln("right","center")
ws1.cell(DATA_ROW, 21).alignment = aln("right","center")    # 販売予定価格
ws1.cell(DATA_ROW, 22).alignment = aln("right","center")    # FBA手数料
# 利益 = 販売価格 - 仕入値（円）- FBA手数料（W=23）
ws1.cell(DATA_ROW, 23).value     = (
    f"=IF(OR(U{DATA_ROW}=\"\",T{DATA_ROW}=\"\"),\"\","
    f"U{DATA_ROW}-T{DATA_ROW}-IF(V{DATA_ROW}=\"\",0,V{DATA_ROW}))"
)
ws1.cell(DATA_ROW, 23).font      = fnt(bold=True)
ws1.cell(DATA_ROW, 23).alignment = aln("right","center")
# 利益率 = 利益 / 販売価格（X=24）
ws1.cell(DATA_ROW, 24).value     = (
    f"=IF(OR(W{DATA_ROW}=\"\",U{DATA_ROW}=0),\"\","
    f"W{DATA_ROW}/U{DATA_ROW})"
)
ws1.cell(DATA_ROW, 24).font      = fnt(bold=True)
ws1.cell(DATA_ROW, 24).alignment = aln("center","center")
ws1.cell(DATA_ROW, 25).alignment = aln("left","center", wrap=True)  # メモ


# ═══════════════════════════════════════════════════════════════════════════
# シート2「②Amazon詳細」
# ═══════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("②Amazon詳細")

# 列幅
ws2.column_dimensions["A"].width = 20   # 項目名
ws2.column_dimensions["B"].width = 80   # 値

# ── Row 1: タイトルバー ────────────────────────────────────────────────
ws2.row_dimensions[1].height = 26
ws2.merge_cells("A1:B1")
c = ws2["A1"]
c.value     = "Amazon 詳細データ  ―  シート②（商品カードからExcel出力すると自動入力されます）"
c.font      = Font(name=FONT_NAME, size=13, bold=True, color=HDR_FG)
c.fill      = fill(S2_HDR_BG)
c.alignment = aln("center", "center")

# ── 項目定義: (ラベル, 行の高さ, 折り返し, セルID用) ─────────────────────
DETAIL_ROWS = [
    ("ASIN",         22,  False, "asin"),
    ("タイトル",     44,  True,  "title"),
    ("現在価格",     22,  False, "price"),
    ("評価（星）",   22,  False, "rating"),
    ("レビュー件数", 22,  False, "review_count"),
    ("A+コンテンツ", 22,  False, "has_aplus"),
    ("Amazon URL",   22,  False, "url"),
    ("───",          8,   False, "sep1"),   # セパレーター
    ("商品の特徴",   22,  False, "bullets_hdr"),  # セクションヘッダー
    # ← 箇条書きは exporter が動的に追加
    ("───",          8,   False, "sep2"),
    ("商品説明",     22,  False, "desc_hdr"),     # セクションヘッダー
    # ← 説明文は exporter が動的に追加
    ("───",          8,   False, "sep3"),
    ("仕様・詳細",   22,  False, "specs_hdr"),    # セクションヘッダー
    # ← 仕様は exporter が動的に追加
]

for row_idx, (label, height, wrap_val, _) in enumerate(DETAIL_ROWS, 2):
    ws2.row_dimensions[row_idx].height = height

    is_sep     = label.startswith("───")
    is_sect_hd = label.endswith("_hdr") or label in ("商品の特徴","商品説明","仕様・詳細")
    # ラベル列
    ca = ws2.cell(row=row_idx, column=1)
    ca.font   = fnt(bold=True, size=10)
    ca.border = bdr()

    if is_sep:
        ws2.merge_cells(f"A{row_idx}:B{row_idx}")
        ca.value     = ""
        ca.fill      = fill("E5E7EB")
        ca.alignment = aln("center","center")
    elif label in ("商品の特徴","商品説明","仕様・詳細"):
        ws2.merge_cells(f"A{row_idx}:B{row_idx}")
        ca.value     = f"【{label}】"
        ca.font      = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
        ca.fill      = fill(S2_SECT_BG)
        ca.alignment = aln("left","center")
        ws2.cell(row=row_idx, column=2).border = bdr()
    else:
        ca.value     = label
        ca.fill      = fill(S2_LBL_BG)
        ca.alignment = aln("right","center")
        # 値列
        cb = ws2.cell(row=row_idx, column=2)
        cb.font      = fnt(size=10)
        cb.fill      = fill(S2_VAL_BG)
        cb.alignment = aln("left","center", wrap=wrap_val)
        cb.border    = bdr()

ws2.freeze_panes = "A2"


# ─────────────────────────────────────────────────────────────────────────
# 保存
# ─────────────────────────────────────────────────────────────────────────
wb.save(OUT_PATH)
print(f"✅ Amazon統合テンプレート生成完了（2シート構成）: {OUT_PATH}")
print()
print("【シート1「①リサーチ」の列構成】")
print("  A〜L列: AucFan リサーチ（元テンプレートと同じ）")
print("  M列: ASIN（アプリが自動入力・薄青）")
print("  N列: Amazon現在価格（自動入力・薄青）")
print("  O列: 評価（自動入力・薄青）")
print("  P列: レビュー件数（自動入力・薄青）")
print("  Q列: A+あり（自動入力・薄青）")
print("  R列: ライバル件数（手入力・薄橙）")
print("  S列: 1688仕入値（元・手入力・薄橙）")
print("  T列: 仕入値（円）= S × T2の為替レート（自動計算・薄緑）")
print("  U列: 販売予定価格（手入力・薄橙）")
print("  V列: FBA手数料（手入力・薄橙）")
print("  W列: 利益 = U - T - V（自動計算・薄緑）")
print("  X列: 利益率 = W ÷ U（自動計算・薄緑）")
print("  Y列: メモ（手入力・薄橙）")
print()
print("【シート2「②Amazon詳細」】")
print("  縦並びで Amazon テキスト全データを表示")
print("  ASIN / タイトル / 現在価格 / 評価 / レビュー件数 / A+")
print("  / URL / 商品の特徴（箇条書き）/ 商品説明 / 仕様・詳細")
print("  各項目がセル単位なのでクリック→コピー可能")
print()
print("【為替レート】")
print(f"  T2セルに {EXCHANGE_RATE_DEFAULT}（円/元）を初期値として設定。")
print("  実際のレートに書き換えると仕入値（円）列が自動再計算されます。")
print()
print("【使い方】")
print("  1. このファイルを リサーチ_テンプレート.xlsx にコピーして使用:")
print("       cp リサーチ_テンプレート_Amazon.xlsx リサーチ_テンプレート.xlsx")
print("  2. 商品カードの「🔍 Amazon調査」→ Amazon取得 → 📗 Excel で")
print("     シート1のAmazon列とシート2の詳細が自動入力されます")
print("  3. 元のテンプレートに戻す: python3 build_template.py")
