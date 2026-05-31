"""
routes/research.py — リサーチ追記ツール関連APIルート

/research および /api/research/* エンドポイントを Flask Blueprint として定義。
app.py から切り出し済み（2026-05-29）。
"""
import os
import re
import json
import threading
import logging
import datetime as dt
from pathlib import Path

import config
from flask import Blueprint, jsonify, render_template, request
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

research_bp = Blueprint("research", __name__)

# Amazon取得中ステータス（フロントエンドがポーリングして進捗表示に使う）
_research_fetch_status = {
    "running": False,
    "step":    "",      # 現在の処理ステップ説明
    "elapsed": 0,       # 開始からの経過秒（フロント側で計算）
}
# 2重実行防止ロック（同時リクエストが来ても1件しか処理しない）
_research_fetch_lock = threading.Lock()

# 1688取得中ステータス（フロントエンドがポーリングして進捗表示に使う）
_research_1688_fetch_status = {
    "running": False,
    "step":    "",
}
# 1688 2重実行防止ロック
_research_1688_fetch_lock = threading.Lock()



@research_bp.route("/research")
def research_page():
    """Excel 追記ツール（AucFan アプリ内ページ）"""
    return render_template("research.html")


@research_bp.route("/api/research/excel/list", methods=["POST"])
def api_research_excel_list():
    """指定フォルダ（省略時は EXCEL_BASE_DIR、なければ OUTPUT_BASE_DIR）の xlsx ファイル一覧を返す。"""
    import time
    body   = request.get_json(silent=True) or {}
    explicit_folder = (body.get("folder") or "").strip()

    if explicit_folder:
        folder = explicit_folder
    else:
        # 新しい保存先 (リサーチシート) を優先、なければ旧保存先 (リサーチ結果) にフォールバック
        excel_base = Path(config.EXCEL_BASE_DIR)
        output_base = Path(config.OUTPUT_BASE_DIR)
        if excel_base.exists():
            folder = str(excel_base)
        elif output_base.exists():
            folder = str(output_base)
        else:
            folder = str(excel_base)  # 存在しなくても新しいパスを返す

    folder_path = Path(folder)

    if not folder_path.exists():
        return jsonify({"success": False, "error": f"フォルダが見つかりません: {folder}", "files": []})

    try:
        import datetime as dt
        files = []
        # フラット（旧形式）+ サブフォルダ（新形式）の両方を拾う
        xlsx_paths = list(folder_path.glob("*.xlsx")) + list(folder_path.glob("*/*.xlsx"))
        xlsx_paths = sorted(xlsx_paths, key=lambda x: x.stat().st_mtime, reverse=True)

        # quick=true の場合はAmazon件数カウントをスキップして高速化
        quick_mode = body.get("quick", False)

        for p in xlsx_paths:
            stat = p.stat()
            # Amazon取得済み件数を簡易チェック（quickモード時はスキップ）
            amazon_count = 0
            if not quick_mode:
                try:
                    wb = load_workbook(str(p), read_only=True, data_only=True)
                    if "②Amazonライバル" in wb.sheetnames:
                        ws = wb["②Amazonライバル"]
                        for r in range(4, (ws.max_row or 3) + 1):
                            if ws.cell(r, 1).value:
                                amazon_count += 1
                    wb.close()
                except Exception as e:
                    logger.debug(f"Excel Amazon件数カウントスキップ: {e}")

            modified = dt.datetime.fromtimestamp(stat.st_mtime).strftime("%Y/%m/%d %H:%M")
            size_kb  = stat.st_size // 1024
            size_str = f"{size_kb}KB" if size_kb < 1024 else f"{size_kb//1024}MB"

            # サブフォルダ内のファイルは「フォルダ名/ファイル名」で表示
            rel = p.relative_to(folder_path)
            display_name = str(rel) if len(rel.parts) > 1 else p.name

            files.append({
                "name":          display_name,
                "path":          str(p),
                "modified":      modified,
                "size":          size_str,
                "amazon_count":  amazon_count,
            })

        return jsonify({"success": True, "folder": str(folder_path), "files": files})

    except Exception as e:
        logger.error(f"ファイル一覧取得エラー: {e}")
        return jsonify({"success": False, "error": str(e), "files": []})


@research_bp.route("/api/research/excel/download", methods=["GET"])
def api_research_excel_download():
    """Excelファイルをブラウザにダウンロードさせる。"""
    from urllib.parse import unquote
    from flask import send_file
    path = unquote(request.args.get("path", "").strip())
    if not path:
        return "パスが指定されていません", 400
    p = Path(path)
    if not p.exists() or not p.is_file():
        return "ファイルが見つかりません", 404
    # パストラバーサル防止: OUTPUT_BASE_DIR または EXCEL_BASE_DIR 配下のみ許可
    resolved = p.resolve()
    allowed = False
    for base_dir in [config.OUTPUT_BASE_DIR, config.EXCEL_BASE_DIR]:
        try:
            resolved.relative_to(Path(base_dir).resolve())
            allowed = True
            break
        except ValueError:
            pass
    if not allowed:
        return "アクセスできないパスです", 403
    return send_file(
        str(p),
        as_attachment=True,
        download_name=p.name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@research_bp.route("/api/research/excel/delete", methods=["POST"])
def api_research_excel_delete():
    """Excelファイルを削除する。パストラバーサル防止あり。"""
    body = request.get_json(silent=True) or {}
    path = (body.get("path") or "").strip()
    if not path:
        return jsonify({"success": False, "error": "パスが指定されていません"}), 400

    p = Path(path)
    if not p.exists() or not p.is_file():
        return jsonify({"success": False, "error": "ファイルが見つかりません"}), 404

    # パストラバーサル防止: OUTPUT_BASE_DIR または EXCEL_BASE_DIR 配下のみ許可
    resolved = p.resolve()
    allowed = False
    for base_dir in [config.OUTPUT_BASE_DIR, config.EXCEL_BASE_DIR]:
        try:
            resolved.relative_to(Path(base_dir).resolve())
            allowed = True
            break
        except ValueError:
            pass
    if not allowed:
        return jsonify({"success": False, "error": "アクセスできないパスです"}), 403

    try:
        p.unlink()
        logger.info(f"Excel削除: {p}")
        return jsonify({"success": True, "deleted": str(p)})
    except Exception as e:
        logger.error(f"Excel削除エラー: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@research_bp.route("/api/research/excel/load", methods=["POST"])
def api_research_excel_load():
    """Excel ファイルの情報を返す。"""
    from excel_append import get_excel_info
    body = request.get_json(silent=True) or {}
    path = (body.get("excel_path") or body.get("path") or "").strip()
    if not path:
        return jsonify({"success": False, "error": "パスが指定されていません"})
    return jsonify(get_excel_info(path))


@research_bp.route("/api/research/amazon/append", methods=["POST"])
def api_research_amazon_append():
    """Chrome の現在開いている Amazon ページを取得して Excel に追記する。"""
    from amazon_scraper import fetch_amazon_product
    from excel_append import append_amazon

    body = request.get_json(silent=True) or {}
    excel_path = body.get("path", "").strip()
    if not excel_path:
        return jsonify({"success": False, "error": "Excelパスが指定されていません"})

    amazon_data = fetch_amazon_product()
    if not amazon_data.get("success"):
        return jsonify(amazon_data), 400

    result = append_amazon(excel_path, amazon_data)
    status = 200 if result.get("success") else 500
    return jsonify(result), status


@research_bp.route("/api/research/amazon/status")
def api_research_amazon_status():
    """Amazon取得中ステータスをポーリングで返す"""
    return jsonify(_research_fetch_status)


@research_bp.route("/api/research/amazon/fetch-url-append", methods=["POST"])
def api_research_amazon_fetch_url_append():
    """URLを指定してAmazonデータを取得し、Excelに追記する。短縮URL対応。"""
    global _research_fetch_status
    from amazon_scraper import fetch_amazon_from_url
    from excel_append import append_amazon

    body = request.get_json(silent=True) or {}
    excel_path = body.get("path", "").strip()
    url        = body.get("url", "").strip()

    if not excel_path:
        return jsonify({"success": False, "error": "Excelパスが指定されていません"})
    if not url:
        return jsonify({"success": False, "error": "URLが指定されていません"})

    # 2重実行防止：前の取得がまだ実行中なら即座に拒否
    if not _research_fetch_lock.acquire(blocking=False):
        return jsonify({
            "success": False,
            "error": "取得処理が実行中です。完了をお待ちください。"
        }), 429

    _research_fetch_status = {"running": True, "step": "① URLを解析中..."}
    try:
        _research_fetch_status["step"] = "② ChromeでAmazonページを開いています..."
        amazon_data = fetch_amazon_from_url(url)
        if not amazon_data.get("success"):
            return jsonify(amazon_data), 400

        _research_fetch_status["step"] = "③ 商品情報を解析中..."
        # 画像・テキスト取得済み確認
        img_count = len(amazon_data.get("image_urls") or [])
        logger.debug(f"Amazon取得完了: ASIN={amazon_data.get('asin')} 画像={img_count}枚")

        _research_fetch_status["step"] = "④ 画像をダウンロード中..."
        _research_fetch_status["step"] = "⑤ Excelに書き込み中..."
        result = append_amazon(excel_path, amazon_data)
        status = 200 if result.get("success") else 500
        return jsonify(result), status
    finally:
        _research_fetch_status = {"running": False, "step": ""}
        _research_fetch_lock.release()


@research_bp.route("/api/research/amazon/open-calculator", methods=["POST"])
def api_open_fba_calculator():
    """
    FBA料金シミュレータをChromeで開き、ASINを自動入力する。
    1688仕入れ値は手動入力のため、自動入力するのは ASIN のみ。

    リクエスト body:
      asin       (str, 任意) : 入力するASIN。省略時はURLを開くだけ。
      excel_path (str, 任意) : asin省略時にSheet2の最初のASINを自動取得する。
    """
    import config as _cfg
    body       = request.get_json(silent=True) or {}
    asin       = body.get("asin", "").strip()
    excel_path = body.get("excel_path", "").strip()

    # asin 未指定 & excel_path あり → Sheet2 から最初の ASIN を取得
    if not asin and excel_path:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(excel_path, read_only=True, data_only=True)
            if "②Amazonライバル" in wb.sheetnames:
                ws2 = wb["②Amazonライバル"]
                for r in range(4, (ws2.max_row or 3) + 1):
                    val = ws2.cell(r, 1).value  # A列 = ASIN
                    if val:
                        asin = str(val).strip()
                        break
            wb.close()
        except Exception as e:
            logger.warning(f"Excel ASIN取得失敗: {e}")

    try:
        from amazon_scraper import _connect_chrome
        import time
        from selenium.webdriver.common.by import By

        CALC_URL = _cfg.REVCAL_URL

        driver = _connect_chrome()
        if not driver:
            return jsonify({"success": False, "error": "Chromeに接続できません（port 9222）"})

        # 既存のシミュレータタブを探す（revcal URLのみ再利用）
        calc_handle = None
        for handle in driver.window_handles:
            try:
                driver.switch_to.window(handle)
                if "revcal" in driver.current_url:
                    calc_handle = handle
                    break
            except Exception as e:
                logger.debug(f"revcalタブ探索スキップ: {e}")

        if calc_handle:
            # 既存タブをリロード
            driver.switch_to.window(calc_handle)
            driver.get(CALC_URL)
        else:
            # Selenium で新タブを開いてからURLに移動（window.open より確実）
            driver.switch_to.new_window("tab")
            driver.get(CALC_URL)

        time.sleep(3)

        # ASIN なし → URL を開くだけ
        if not asin:
            logger.info("FBA料金シミュレータ: URLを開くだけ（ASIN未指定）")
            return jsonify({
                "success": True, "asin": "", "url": CALC_URL,
                "message": "シミュレータを開きました。ASINと仕入れ値を手動で入力してください。"
            })

        # ── ASIN 入力欄を探す（通常セレクター → Shadow DOM 再帰） ────────
        input_el = None
        for sel in ["input[id*='asin']", "input[name*='asin']",
                    "input[placeholder*='ASIN']", "input[placeholder*='asin']",
                    "input[data-testid*='asin']", "#asin-search-input"]:
            try:
                for el in driver.find_elements(By.CSS_SELECTOR, sel):
                    if el.is_displayed() and el.is_enabled():
                        input_el = el
                        break
                if input_el:
                    break
            except Exception as e:
                logger.debug(f"Shadow DOM input探索スキップ: {e}")

        if not input_el:
            try:
                input_el = driver.execute_script("""
                    var NAV = '#navbar,#navbar-main,nav,header,[id*="nav-search"],[id*="topnav"],' +
                              '[class*="navbar"],[class*="nav-bar"],[class*="nav-search"]';
                    function isInNav(el) {
                        try { return !!el.closest(NAV); } catch(e) { return false; }
                    }
                    function findInput(root) {
                        var pri = root.querySelectorAll(
                            'input[id*="asin"],input[name*="asin"],' +
                            'input[placeholder*="ASIN"],input[data-testid*="asin"]');
                        for (var i=0;i<pri.length;i++)
                            if (pri[i].offsetParent!==null && !isInNav(pri[i])) return pri[i];
                        var all2 = root.querySelectorAll('input[type="text"],input:not([type])');
                        for (var j=0;j<all2.length;j++)
                            if (all2[j].offsetParent!==null && !isInNav(all2[j])) return all2[j];
                        var all = root.querySelectorAll('*');
                        for (var k=0;k<all.length;k++)
                            if (all[k].shadowRoot) {
                                var f = findInput(all[k].shadowRoot);
                                if (f) return f;
                            }
                        return null;
                    }
                    return findInput(document);
                """)
            except Exception as e:
                logger.debug(f"JS input探索スキップ: {e}")

        if not input_el:
            logger.warning("FBA料金シミュレータ: ASIN入力欄が見つかりませんでした")
            return jsonify({
                "success": True, "asin": asin, "url": CALC_URL,
                "warning": "シミュレータを開きましたが、ASIN入力欄が見つかりませんでした。手動で入力してください。"
            })

        # ASIN 入力（1文字ずつKeyboardEventを発火してKATフレームワークに認識させる）
        driver.execute_script("""
            var el = arguments[0], val = arguments[1];
            el.scrollIntoView({block:'center'});
            el.click();
            el.focus();
            el.value = '';
            for (var i = 0; i < val.length; i++) {
                var ch = val[i];
                el.dispatchEvent(new KeyboardEvent('keydown',  {key:ch, bubbles:true}));
                el.dispatchEvent(new KeyboardEvent('keypress', {key:ch, bubbles:true}));
                el.value += ch;
                el.dispatchEvent(new Event('input', {bubbles:true}));
                el.dispatchEvent(new KeyboardEvent('keyup',    {key:ch, bubbles:true}));
            }
            el.dispatchEvent(new Event('change', {bubbles:true}));
        """, input_el, asin)
        time.sleep(0.5)

        # 送信ボタンをクリック（Shadow DOM 対応）
        submitted = False
        try:
            submitted = driver.execute_script("""
                function findBtn(root) {
                    var btns = root.querySelectorAll(
                        'button[type="submit"],input[type="submit"],' +
                        'button[id*="search"],button[id*="submit"],' +
                        '.kat-button--primary,button.a-button-primary,kat-button[variant="primary"]');
                    for (var i=0;i<btns.length;i++)
                        if (btns[i].offsetParent!==null) { btns[i].click(); return true; }
                    var all = root.querySelectorAll('*');
                    for (var k=0;k<all.length;k++)
                        if (all[k].shadowRoot && findBtn(all[k].shadowRoot)) return true;
                    return false;
                }
                return findBtn(document);
            """)
        except Exception as e:
            logger.debug(f"Submitボタン検索スキップ: {e}")

        if not submitted:
            # フォールバック: ActionChains でクリックしてから Enter
            try:
                from selenium.webdriver.common.action_chains import ActionChains
                from selenium.webdriver.common.keys import Keys
                ActionChains(driver).move_to_element(input_el).click().send_keys(Keys.RETURN).perform()
            except Exception as e:
                logger.debug(f"ActionChains Enterキースキップ: {e}")

        logger.info(f"FBA料金シミュレータ: ASIN={asin} 入力完了（仕入れ値は手動入力）")
        return jsonify({"success": True, "asin": asin, "url": CALC_URL})

    except Exception as e:
        logger.error(f"FBA料金シミュレータ起動エラー: {e}", exc_info=True)
        return jsonify({"success": False, "error": str(e)})


@research_bp.route("/api/research/amazon/read-calc", methods=["POST"])
def api_read_fba_calc():
    """
    現在開いている revcal タブから計算結果をスクレイピングし、
    Excelの Sheet1 に転記する。

    リクエスト body:
      excel_path (str, 必須) : 書き込み先 Excel のフルパス

    書き込み先:
      Sheet1 B12 = 販売価格（空の場合のみ上書き）
      Sheet1 B13 = FBA手数料合計（常に上書き）
    """
    import config as _cfg
    body       = request.get_json(silent=True) or {}
    excel_path = body.get("excel_path", "").strip()

    if not excel_path:
        return jsonify({"success": False, "error": "Excelパスが指定されていません"})

    try:
        from amazon_scraper import _connect_chrome
        from pathlib import Path
        import time

        driver = _connect_chrome()
        if not driver:
            return jsonify({"success": False, "error": "Chromeに接続できません（port 9222）"})

        # revcal タブに切り替え
        calc_handle = None
        for handle in driver.window_handles:
            try:
                driver.switch_to.window(handle)
                if "revcal" in driver.current_url or "revcal" in driver.current_url.lower():
                    calc_handle = handle
                    break
            except Exception as e:
                logger.debug(f"revcalタブ探索(read-calc)スキップ: {e}")

        if not calc_handle:
            return jsonify({
                "success": False,
                "error": "FBA料金シミュレータのタブが見つかりません。先に「シミュレータを開く」ボタンを押してください。"
            })

        driver.switch_to.window(calc_handle)
        time.sleep(1)

        # ── ページ全体のテキストを取得（Shadow DOM も含む） ────────────
        page_text = driver.execute_script("""
            function collectText(root) {
                var texts = [];
                var walker = document.createTreeWalker(root, NodeFilter.SHOW_TEXT, null, false);
                var node;
                while (node = walker.nextNode()) {
                    var t = node.textContent.trim();
                    if (t) texts.push(t);
                }
                root.querySelectorAll('*').forEach(function(el) {
                    if (el.shadowRoot) texts = texts.concat(collectText(el.shadowRoot));
                });
                return texts;
            }
            return collectText(document).join('\\n');
        """) or ""

        # ── 数値抽出ヘルパー ──────────────────────────────────────────
        import re

        def extract_yen(pattern, text):
            """¥X,XXX 形式の数値を抽出して int を返す。"""
            m = re.search(pattern, text, re.DOTALL)
            if m:
                raw = re.sub(r'[¥,\s円]', '', m.group(1))
                try:
                    return int(float(raw))
                except (ValueError, TypeError):
                    pass
            return None

        def extract_pct(pattern, text):
            """XX.X% 形式の数値を抽出して float を返す。"""
            m = re.search(pattern, text, re.DOTALL)
            if m:
                try:
                    return float(m.group(1))
                except (ValueError, TypeError):
                    pass
            return None

        # ── 各値を抽出 ────────────────────────────────────────────────
        # ページ上の数値は「ラベル」の直後（または近傍）にある
        selling_price = extract_yen(
            r'(?:商品売上|販売価格|売上)[^\d¥\-]*[\-]?¥?\s*([\d,]+)', page_text
        )
        referral_fee = extract_yen(
            r'参照手数料[^\d¥\-]*[\-]?¥?\s*([\d,]+)', page_text
        )
        fba_fee = extract_yen(
            r'FBA配送代行手数料[^\d¥\-]*[\-]?¥?\s*([\d,]+)', page_text
        )
        # 合計手数料は「合計」または「手数料合計」など
        total_fee = extract_yen(
            r'(?:合計手数料|手数料合計|合計)[^\d¥\-\n]*[\-]?¥?\s*([\d,]+)', page_text
        )
        # 合計が取れなければ referral + fba で計算
        if total_fee is None and referral_fee is not None and fba_fee is not None:
            total_fee = referral_fee + fba_fee

        profit = extract_yen(
            r'利益[^\d¥\-\n%]*[\-]?¥?\s*([\d,]+)', page_text
        )
        profit_rate = extract_pct(
            r'利益率[^\d%\n]*(\d+\.?\d*)\s*%', page_text
        )

        if total_fee is None:
            return jsonify({
                "success": False,
                "error": (
                    "FBA手数料が取得できませんでした。\n"
                    "シミュレータでASINを検索して結果が表示された状態で再試行してください。"
                ),
                "page_sample": page_text[:300]   # デバッグ用
            })

        # ── Excel に書き込む ──────────────────────────────────────────
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill

        path = Path(excel_path)
        if not path.exists():
            return jsonify({"success": False, "error": f"ファイルが見つかりません: {excel_path}"})

        wb = load_workbook(str(path))
        SHEET1 = "①概要"
        if SHEET1 not in wb.sheetnames:
            return jsonify({"success": False, "error": "シート「①概要」が見つかりません"})

        ws1 = wb[SHEET1]

        # B12: 販売価格（空の場合のみ書き込む）
        b12_written = False
        if selling_price and not ws1["B12"].value:
            ws1["B12"].value = selling_price
            b12_written = True

        # B13: FBA手数料合計（常に上書き）
        ws1["B13"].value = total_fee
        # セルを黄色でハイライト（転記済みマーク）
        ws1["B13"].fill = PatternFill("solid", fgColor="FFF2CC")

        wb.save(str(path))

        # ── 仕入れ判断 ────────────────────────────────────────────────
        judgment = None
        if profit is not None and profit_rate is not None:
            if profit_rate >= _cfg.PROFIT_RATE_THRESHOLD or profit >= _cfg.PROFIT_YEN_THRESHOLD:
                judgment = "◎ GO"
            else:
                judgment = "× 再検討"

        logger.info(
            f"revcal転記完了: FBA={total_fee}円 / 利益={profit}円 / 利益率={profit_rate}%"
        )
        return jsonify({
            "success":       True,
            "selling_price": selling_price,
            "referral_fee":  referral_fee,
            "fba_fee":       fba_fee,
            "total_fee":     total_fee,
            "profit":        profit,
            "profit_rate":   profit_rate,
            "judgment":      judgment,
            "b12_written":   b12_written,
            "profit_rate_threshold": _cfg.PROFIT_RATE_THRESHOLD,
            "profit_yen_threshold":  _cfg.PROFIT_YEN_THRESHOLD,
        })

    except Exception as e:
        logger.error(f"revcal転記エラー: {e}", exc_info=True)
        return jsonify({"success": False, "error": str(e)})


@research_bp.route("/api/research/1688/status")
def api_research_1688_status():
    """1688取得中ステータスをポーリングで返す"""
    return jsonify(_research_1688_fetch_status)


@research_bp.route("/api/research/1688/fetch-url", methods=["POST"])
def api_research_1688_fetch_url():
    """
    URLを指定して1688データを取得する（Excelへの追記はしない）。
    取得したバリアント一覧をUIに返して、ユーザーが選択後に /append を呼ぶ。
    """
    global _research_1688_fetch_status

    body = request.get_json(silent=True) or {}
    url  = body.get("url", "").strip()

    if not url:
        return jsonify({"success": False, "error": "URLが指定されていません"})

    if not _research_1688_fetch_lock.acquire(blocking=False):
        return jsonify({
            "success": False,
            "error": "取得処理が実行中です。完了をお待ちください。"
        }), 429

    excel_path = body.get("path", "").strip()
    if excel_path:
        excel_path = _validate_excel_path(excel_path) or ""

    _research_1688_fetch_status = {"running": True, "step": "① URLを解析中..."}
    try:
        from scraper_1688 import fetch_1688_from_url
        _research_1688_fetch_status["step"] = "② Chromeで1688ページを開いています..."
        data_1688 = fetch_1688_from_url(url)
        if not data_1688.get("success"):
            return jsonify(data_1688), 400

        # ── 重複チェック ──────────────────────────────────────
        duplicate = False
        if excel_path:
            try:
                wb = load_workbook(excel_path, read_only=True, data_only=True)
                if "④1688仕入れ" in wb.sheetnames:
                    ws = wb["④1688仕入れ"]
                    new_shop  = (data_1688.get("shop_name") or "").strip()
                    new_title = (data_1688.get("title") or "").strip()
                    for row in ws.iter_rows(min_row=4, values_only=True):
                        ex_shop  = str(row[1] or "").strip()   # B列: ショップ名
                        ex_title = str(row[5] or "").strip()   # F列: 商品名（中）
                        if ex_shop == new_shop and ex_title == new_title:
                            duplicate = True
                            break
                wb.close()
            except Exception:
                pass  # チェック失敗時はスルーして通常処理

        return jsonify({
            "success":   True,
            "data":      data_1688,
            "duplicate": duplicate,
        }), 200
    finally:
        _research_1688_fetch_status = {"running": False, "step": ""}
        _research_1688_fetch_lock.release()


@research_bp.route("/api/research/1688/append", methods=["POST"])
def api_research_1688_append():
    """
    取得済みの1688データをExcelに追記する。
    selected_flags: {バリアントname: 1 or 0} を受け取りA列に反映する。
    """
    from excel_append import append_1688

    body        = request.get_json(silent=True) or {}
    excel_path  = body.get("path", "").strip()
    data_1688   = body.get("data", {})
    flags       = body.get("selected_flags", {})  # {variant_name: 1 or 0}

    if not excel_path:
        return jsonify({"success": False, "error": "Excelパスが指定されていません"})
    if not data_1688:
        return jsonify({"success": False, "error": "1688データがありません"})

    excel_path = _validate_excel_path(excel_path)
    if not excel_path:
        return jsonify({"success": False, "error": "不正なパスです"}), 400

    # flagsをdata_1688のvariantsに反映
    for v in data_1688.get("variants", []):
        v["selected"] = flags.get(v.get("name", ""), 0)

    result = append_1688(excel_path, data_1688)
    if not result.get("success"):
        return jsonify(result), 500

    profit_result = _calc_profit_from_excel(excel_path)
    result.update(profit_result)
    return jsonify(result), 200


@research_bp.route("/api/research/1688/variants", methods=["POST"])
def api_research_1688_variants():
    """
    Excelの④1688仕入れシートから現在のバリアント一覧とA列フラグを読み込んで返す。
    """
    body       = request.get_json(silent=True) or {}
    excel_path = body.get("path", "").strip()
    if not excel_path:
        return jsonify({"success": False, "error": "Excelパスが指定されていません"})
    excel_path = _validate_excel_path(excel_path)
    if not excel_path:
        return jsonify({"success": False, "error": "不正なパスです"}), 400

    try:
        wb = load_workbook(excel_path, data_only=True)
        if "④1688仕入れ" not in wb.sheetnames:
            return jsonify({"success": False, "error": "④1688仕入れシートがありません"})

        ws = wb["④1688仕入れ"]
        variants = []
        for row in ws.iter_rows(min_row=4, values_only=True):
            flag      = row[0]   # A列
            shop      = row[1]   # B列
            variant   = row[8] or row[7] or ""  # I列 or H列
            price     = row[10]  # K列
            if price is None:
                continue
            variants.append({
                "flag":    1 if flag == 1 else 0,
                "shop":    str(shop or ""),
                "variant": str(variant),
                "price":   price,
            })

        return jsonify({"success": True, "variants": variants})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@research_bp.route("/api/research/1688/update-flags", methods=["POST"])
def api_research_1688_update_flags():
    """
    Excelの④1688仕入れシートのA列フラグをUIからの選択内容で更新する。
    rows: [{row_index: N, flag: 1 or 0}, ...] (4行目=0ベースのインデックス)
    """
    body       = request.get_json(silent=True) or {}
    excel_path = body.get("path", "").strip()
    rows       = body.get("rows", [])  # [{row_index: N, flag: 1 or 0}]

    if not excel_path:
        return jsonify({"success": False, "error": "Excelパスが指定されていません"})
    excel_path = _validate_excel_path(excel_path)
    if not excel_path:
        return jsonify({"success": False, "error": "不正なパスです"}), 400

    try:
        wb = load_workbook(excel_path)
        if "④1688仕入れ" not in wb.sheetnames:
            return jsonify({"success": False, "error": "④1688仕入れシートがありません"})

        ws = wb["④1688仕入れ"]
        for r in rows:
            excel_row = int(r["row_index"]) + 4  # 0ベース → 4行目スタート
            ws.cell(excel_row, 1).value = int(r["flag"])

        wb.save(excel_path)
        profit_result = _calc_profit_from_excel(excel_path)
        return jsonify({"success": True, **profit_result})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@research_bp.route("/api/research/1688/fetch-url-append", methods=["POST"])
def api_research_1688_fetch_url_append():
    """URLを指定して1688データを取得し、Excelに追記する（旧API・後方互換用）。"""
    global _research_1688_fetch_status
    from excel_append import append_1688

    body = request.get_json(silent=True) or {}
    excel_path = body.get("path", "").strip()
    url        = body.get("url", "").strip()

    if not excel_path:
        return jsonify({"success": False, "error": "Excelパスが指定されていません"})
    if not url:
        return jsonify({"success": False, "error": "URLが指定されていません"})

    if not _research_1688_fetch_lock.acquire(blocking=False):
        return jsonify({
            "success": False,
            "error": "取得処理が実行中です。完了をお待ちください。"
        }), 429

    _research_1688_fetch_status = {"running": True, "step": "① URLを解析中..."}
    try:
        from scraper_1688 import fetch_1688_from_url
        _research_1688_fetch_status["step"] = "② Chromeで1688ページを開いています..."
        data_1688 = fetch_1688_from_url(url)
        if not data_1688.get("success"):
            return jsonify(data_1688), 400

        _research_1688_fetch_status["step"] = "③ Excelに書き込み中..."
        result = append_1688(excel_path, data_1688)
        if not result.get("success"):
            return jsonify(result), 500

        profit_result = _calc_profit_from_excel(excel_path)
        result.update(profit_result)
        return jsonify(result), 200
    finally:
        _research_1688_fetch_status = {"running": False, "step": ""}
        _research_1688_fetch_lock.release()


# ─────────────────────────────────────────────────────────────────────────────
# SP-API: ASIN → 商品情報 + FBA手数料取得 → Excel 転記
# ─────────────────────────────────────────────────────────────────────────────

@research_bp.route("/api/research/amazon/sp-fetch", methods=["POST"])
def api_sp_fetch():
    """
    SP-API (Catalog Items + Products Fees) でAmazon商品情報を取得し、
    Excelの Sheet1 B12（販売価格）・B13（FBA手数料）に転記する。

    Request JSON:
      {
        "excel_path": "/path/to/xxx_リサーチ.xlsx",
        "asin":       "B0XXXXXXXX",
        "price":      3000          # 省略可。省略時はAmazon参考価格を使用。
      }

    Response JSON:
      {
        "success": true/false,
        "asin":         str,
        "title":        str,
        "rank":         int | null,
        "category":     str,
        "list_price":   int | null,
        "price":        int | null,
        "fba_fee":      int | null,
        "referral_fee": int | null,
        "total_fee":    int | null,
        "profit_ok":    bool,        # ◎ GO 条件を満たすかどうか
        "message":      str,
        "error":        str | null
      }
    """
    body = request.get_json(force=True, silent=True) or {}
    excel_path = body.get("excel_path", "")
    asin       = (body.get("asin") or "").strip().upper()
    price      = body.get("price")   # int | None

    if not asin:
        return jsonify({"success": False, "error": "asin が指定されていません"}), 400

    # price を int に変換
    if price is not None:
        try:
            price = int(price)
        except (ValueError, TypeError):
            price = None

    try:
        from sp_api_client import get_client
        client = get_client()
        info = client.fetch_product_info(asin, price)
    except Exception as e:
        logger.error(f"[SP-API] 取得エラー: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

    if info.get("error"):
        return jsonify({"success": False, **info}), 500

    # ── Excel 転記（パスが指定されている場合のみ） ──────────────────────
    if excel_path:
        excel_path = _validate_excel_path(excel_path)
        if not excel_path:
            return jsonify({"success": False, "error": "不正なExcelパスです"}), 400
        try:
            from openpyxl.styles import PatternFill
            sp_fill = PatternFill("solid", fgColor=config.EXCEL_COLOR_SP_API)

            wb = load_workbook(excel_path)
            ws = wb["①概要"]
            effective_price = info.get("price")
            fba_fee         = info.get("fba_fee")

            if effective_price and ws["B12"].value is None:
                ws["B12"] = effective_price
                ws["B12"].fill = sp_fill
            if fba_fee:
                ws["B13"] = fba_fee
                ws["B13"].fill = sp_fill

            wb.save(excel_path)
            logger.info(f"[SP-API] Excel転記完了: {excel_path} B12={effective_price} B13={fba_fee}")
        except Exception as e:
            logger.error(f"[SP-API] Excel転記エラー: {e}")
            return jsonify({"success": False, "error": f"Excel転記失敗: {e}"}), 500

    # ── 1688原価込み利益計算（共通関数で処理）────────────────────────────
    profit_result   = _calc_profit_from_excel(excel_path) if excel_path else {}
    variants_profit = profit_result.get("variants_profit", [])
    profit_ok       = profit_result.get("profit_ok", False)

    return jsonify({
        "success":          True,
        "asin":             info["asin"],
        "title":            info["title"],
        "brand":            info.get("brand", ""),
        "rank":             info["rank"],
        "category":         info["category"],
        "list_price":       info["list_price"],
        "price":            info["price"],
        "fba_fee":          info["fba_fee"],
        "referral_fee":     info.get("referral_fee"),
        "total_fee":        info.get("total_fee"),
        "profit_ok":        profit_ok,
        "variants_profit":  variants_profit,   # 全バリアントの利益計算結果
        "fee_error":        info.get("fee_error"),
        "message":          "SP-API 取得・転記完了",
        "error":            None,
    })


def _calc_profit_from_excel(excel_path: str) -> dict:
    """
    Excelから販売価格(B12)・FBA手数料(B13)・Sheet4の1688単価を読み取り、
    全バリアントの利益計算を行う共通関数。

    Returns:
        {
          "profit_ok":       bool,
          "variants_profit": list,   # バリアント別 {variant, cny_price, cost, profit, profit_rate, ok}
          "selling_price":   int,
          "fba_fee":         int,
        }
    """
    result = {"profit_ok": False, "variants_profit": [], "selling_price": None, "fba_fee": None}
    try:
        wb = load_workbook(excel_path, data_only=True)

        # B12/B13 を読む
        ws1 = wb["①概要"]
        selling_price = ws1["B12"].value
        fba_fee_val   = ws1["B13"].value or 0

        if not isinstance(selling_price, (int, float)) or selling_price <= 0:
            return result   # 販売価格未設定 → 計算不可

        result["selling_price"] = int(selling_price)
        result["fba_fee"]       = int(fba_fee_val)

        # Sheet4 の K列・L列を読む
        if "④1688仕入れ" not in wb.sheetnames:
            return result

        ws4 = wb["④1688仕入れ"]
        variants_profit = []
        for row in ws4.iter_rows(min_row=4, values_only=True):
            flag       = row[0]    # A列: 1=仕入れ対象 / 0or空=除外
            cny_price  = row[10]   # K列
            rate_val   = row[11]   # L列
            variant_ja = row[8] or row[7] or ""   # I列 or H列

            # A列が 1 の行だけ利益計算対象にする（0・空白は無視）
            if flag != 1:
                continue
            if not isinstance(cny_price, (int, float)) or cny_price <= 0:
                continue

            rate    = rate_val if isinstance(rate_val, (int, float)) else config.CNY_TO_JPY_RATE
            cost    = round(cny_price * rate)
            profit  = round(selling_price - fba_fee_val - cost)
            p_rate  = round(profit / selling_price * 100, 1) if selling_price > 0 else 0
            ok      = (p_rate >= config.PROFIT_RATE_THRESHOLD or
                       profit >= config.PROFIT_YEN_THRESHOLD)

            variants_profit.append({
                "variant":     str(variant_ja),
                "cny_price":   cny_price,
                "cost":        cost,
                "profit":      profit,
                "profit_rate": p_rate,
                "ok":          ok,
            })

        result["variants_profit"] = variants_profit
        result["profit_ok"]       = any(v["ok"] for v in variants_profit)
        logger.info(f"[利益計算] {len(variants_profit)}バリアント profit_ok={result['profit_ok']}")
    except Exception as e:
        logger.warning(f"[利益計算] エラー（スキップ）: {e}")
    return result


def _validate_excel_path(path: str) -> str | None:
    """パスが許可ディレクトリ内にあるかチェックして正規化して返す。不正なら None。"""
    try:
        p = Path(path).resolve()
        allowed = [
            Path(config.EXCEL_BASE_DIR).resolve(),
            Path(config.OUTPUT_BASE_DIR).resolve(),
        ]
        for a in allowed:
            try:
                p.relative_to(a)
                return str(p)
            except ValueError:
                continue
        return None
    except Exception:
        return None


