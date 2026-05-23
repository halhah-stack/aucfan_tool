"""
build_template.py — Aucfan リサーチシート テンプレート生成

実行するとリサーチ_テンプレート.xlsx を aucfan_tool/ フォルダに作成する。
書式を変えたいときはこのスクリプトを編集して再実行するか、
生成されたExcelを直接 Excel アプリで編集してください。

実行方法:
  cd ~/Downloads/aucfan_tool
  python3 build_template.py
"""

from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 出力先 ────────────────────────────────────────────────────────────
OUT_PATH = Path(__file__).parent / "リサーチ_テンプレート.xlsx"

# ── スタイル定数（ここを変更すれば全体に反映） ───────────────────────
FONT_NAME = "BIZ UDGothic"
FONT_SIZE = 11

HDR_BG = "1F4E79"   # タイトルバー（紺）
HDR_FG = "FFFFFF"
SUB_BG = "2E75B6"   # 列ヘッダー（青）
SUB_FG = "FFFFFF"
LBL_BG = "D6E4F0"   # ラベル（薄青）
INP_BG = "FFF9C4"   # 手入力セル（黄）
FML_BG = "E8F5E9"   # 自動計算セル（緑）
IMG_BG = "F0F0F0"   # 画像エリア（薄灰）

DATA_ROW = 6        # データ行（テンプレートは1行分のみ用意）

# ── 列定義 ────────────────────────────────────────────────────────────
# (列名, 幅, ヘッダーテキスト, 背景色, 書式)
COLUMNS = [
    ("A", 4,  "No",                    None,   None),
    ("B", 16, "画像",                  IMG_BG, None),
    ("C", 11, "日付",                  INP_BG, "YYYY/MM/DD"),
    ("D", 42, "代表キーワード",        INP_BG, None),
    ("E", 12, "件数\n（直近30日）",   FML_BG, "#,##0"),
    ("F", 9,  "4件\n以上",            FML_BG, None),
    ("G", 16, "最安値\n（価格＋送料）", FML_BG, "#,##0"),
    ("H", 18, "セラーID①",            INP_BG, None),
    ("I", 18, "セラーID②",            INP_BG, None),
    ("J", 18, "セラーID③",            INP_BG, None),
    ("K", 38, "検索URL\n（Aucfan）",  INP_BG, None),
    ("L", 22, "備考",                  INP_BG, None),
]

# ── ヘルパー ──────────────────────────────────────────────────────────
def fnt(bold=False, size=FONT_SIZE, color="000000"):
    return Font(name=FONT_NAME, size=size, bold=bold, color=color)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def aln(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def bdr(color="CCCCCC"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def cell(ws, row, col, value=None, font=None, bg=None,
         align=None, border=None, fmt=None):
    c = ws.cell(row=row, column=col)
    if value  is not None: c.value  = value
    if font:               c.font   = font
    if bg:                 c.fill   = fill(bg)
    if align:              c.alignment = align
    if border:             c.border = border
    if fmt:                c.number_format = fmt
    return c

# ── ワークブック生成 ──────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "①Aucfan"

# 列幅
for col_letter, width, *_ in COLUMNS:
    ws.column_dimensions[col_letter].width = width

# ── Row 1: タイトルバー ──────────────────────────────────────────────
ws.row_dimensions[1].height = 26
ws.merge_cells("A1:L1")
c = ws["A1"]
c.value     = "Aucfan リサーチシート"
c.font      = Font(name=FONT_NAME, size=13, bold=True, color=HDR_FG)
c.fill      = fill(HDR_BG)
c.alignment = aln("center", "center")

# ── Row 2: サマリー行 ────────────────────────────────────────────────
ws.row_dimensions[2].height = 22

# 調査商品数
cell(ws, 2, 1, "調査商品数",
     font=fnt(bold=True), bg=LBL_BG,
     align=aln("center","center"), border=bdr())
cell(ws, 2, 2, f"=COUNTA(D{DATA_ROW}:D{DATA_ROW})",
     font=fnt(bold=True), bg=FML_BG,
     align=aln("center","center"), border=bdr())

# 4件以上
cell(ws, 2, 3, "4件以上の商品",
     font=fnt(bold=True), bg=LBL_BG,
     align=aln("center","center"), border=bdr())
cell(ws, 2, 4, f'=COUNTIF(E{DATA_ROW}:E{DATA_ROW},">=4")',
     font=fnt(bold=True), bg=FML_BG,
     align=aln("center","center"), border=bdr())

# セッション名（コードが書き込む）
cell(ws, 2, 5, "セッション",
     font=fnt(bold=True), bg=LBL_BG,
     align=aln("center","center"), border=bdr())
ws.merge_cells("F2:L2")
cell(ws, 2, 6, "",
     font=Font(name=FONT_NAME, size=10, color="444444"),
     bg=INP_BG, align=aln("left","center"), border=bdr())

# ── Row 3: 凡例 ──────────────────────────────────────────────────────
ws.row_dimensions[3].height = 18
ws.merge_cells("A3:L3")
c = ws["A3"]
c.value     = "  ■ 黄色 = 手入力セル　　■ 緑色 = 自動計算セル"
c.font      = Font(name=FONT_NAME, size=10, color="555555")
c.fill      = fill("FAFAFA")
c.alignment = aln("left", "center")

# ── Row 4: スペーサー ─────────────────────────────────────────────────
ws.row_dimensions[4].height = 4

# ── Row 5: 列ヘッダー ────────────────────────────────────────────────
ws.row_dimensions[5].height = 34
for col_idx, (_, _, header, _, _) in enumerate(COLUMNS, 1):
    c = ws.cell(row=5, column=col_idx)
    c.value     = header
    c.font      = Font(name=FONT_NAME, size=FONT_SIZE, bold=True, color=SUB_FG)
    c.fill      = fill(SUB_BG)
    c.alignment = aln("center", "center", wrap=True)
    c.border    = bdr()

ws.freeze_panes = "A6"

# ── Row 6: データ行テンプレート（書式のみ、値は空） ──────────────────
ws.row_dimensions[DATA_ROW].height = 72

for col_idx, (_, _, _, bg_color, fmt_str) in enumerate(COLUMNS, 1):
    c = ws.cell(row=DATA_ROW, column=col_idx)
    c.font      = fnt()
    c.alignment = aln("left", "center", wrap=(col_idx in [4, 12]))
    c.border    = bdr()
    if bg_color:
        c.fill = fill(bg_color)
    if fmt_str:
        c.number_format = fmt_str

# No列だけ中央揃え
ws.cell(DATA_ROW, 1).alignment = aln("center", "center")
# 日付列
ws.cell(DATA_ROW, 3).alignment = aln("center", "center")
# 件数列
ws.cell(DATA_ROW, 5).alignment = aln("center", "center")
# 4件以上列（数式プレースホルダー）
ws.cell(DATA_ROW, 6).value     = f'=IF(E{DATA_ROW}="","",IF(E{DATA_ROW}>=4,"✓","✗"))'
ws.cell(DATA_ROW, 6).font      = fnt(bold=True)
ws.cell(DATA_ROW, 6).alignment = aln("center", "center")
# 最安値列
ws.cell(DATA_ROW, 7).alignment = aln("right", "center")
# 画像エリア
img_cell = ws.cell(DATA_ROW, 2)
img_cell.value     = "画像\nここに貼付"
img_cell.font      = Font(name=FONT_NAME, size=9, color="AAAAAA")
img_cell.alignment = aln("center", "center", wrap=True)

# ── 保存 ─────────────────────────────────────────────────────────────
wb.save(OUT_PATH)
print(f"✅ テンプレート生成完了: {OUT_PATH}")
print("   このファイルを Excel で開いて書式を自由に調整してください。")
print("   調整後は上書き保存するだけで次回エクスポートに反映されます。")
