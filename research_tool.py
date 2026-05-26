"""
research_tool.py — スタンドアロン リサーチ追記ツール

aucfantool アプリとは完全に独立して動作する。
既存の Excel リサーチファイルを読み込み、
Chrome で開いている Amazon ページのデータを Sheet2/3 に追記する。

起動方法:
    cd ~/Downloads/aucfan_tool
    source .venv/bin/activate
    python3 research_tool.py

ブラウザ: http://localhost:5002
"""

import io
import json
import logging
import os
import re
import shutil
import tempfile
import urllib.request
from datetime import datetime
from pathlib import Path
from typing import Optional

from flask import Flask, jsonify, render_template_string, request, send_file

app = Flask(__name__)
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)

# ── 設定 ───────────────────────────────────────────────────────────────
CHROME_HOST = "localhost"
CHROME_PORT = 9222
PORT        = 5002

# シート名（excel_exporter.py と揃える）
SHEET_AMAZON_LIST = "②Amazonライバル"
SHEET_AMAZON_TEXT = "③Amazonテキスト"
SHEET_1688_LIST   = "④1688仕入れ"
SHEET_1688_TEXT   = "⑤1688テキスト"

# Amazon画像保存フォルダ（Excelと同じディレクトリに作成）
IMAGE_FOLDER_NAME = "_images"

# ── Chrome 接続 ────────────────────────────────────────────────────────
def _connect_chrome():
    """既存 Chrome に接続して driver を返す。失敗時は None。"""
    try:
        from selenium import webdriver
        from selenium.webdriver.chrome.options import Options
        options = Options()
        options.add_experimental_option("debuggerAddress", f"{CHROME_HOST}:{CHROME_PORT}")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        driver = webdriver.Chrome(options=options)
        driver.set_page_load_timeout(15)
        return driver
    except Exception as e:
        logger.error(f"Chrome接続失敗: {e}")
        return None


# ── Amazon ページ解析 ──────────────────────────────────────────────────
def _fetch_amazon_from_chrome() -> dict:
    """Chrome の Amazon 商品ページからデータを取得する。"""
    driver = _connect_chrome()
    if not driver:
        return {"success": False, "error": "Chromeに接続できません（port 9222）"}

    try:
        from bs4 import BeautifulSoup

        amazon_handle = None
        current_url = ""
        for handle in driver.window_handles:
            driver.switch_to.window(handle)
            url = driver.current_url
            if "amazon.co.jp" in url and ("/dp/" in url or "/gp/product/" in url):
                amazon_handle = handle
                current_url = url
                break

        if not amazon_handle:
            return {"success": False, "error": "amazon.co.jp の商品ページ（/dp/...）をChromeで開いてください"}

        # ASIN
        m = re.search(r"/dp/([A-Z0-9]{10})", current_url)
        asin = m.group(1) if m else ""

        html = driver.page_source
        soup = BeautifulSoup(html, "html.parser")

        # タイトル
        title = ""
        el = soup.select_one("#productTitle")
        if el:
            title = el.get_text(strip=True)

        # 価格
        price = ""
        for sel in [".a-price .a-offscreen", "#priceblock_ourprice",
                    ".apexPriceToPay .a-offscreen", ".a-price-whole"]:
            el = soup.select_one(sel)
            if el:
                price = el.get_text(strip=True)
                break

        # 画像URL
        image_url = ""
        img_el = soup.select_one("#landingImage, #imgTagWrapperId img, #main-image")
        if img_el:
            image_url = (
                img_el.get("data-old-hires")
                or img_el.get("src", "")
            )
            if not image_url and img_el.get("data-a-dynamic-image"):
                parts = img_el.get("data-a-dynamic-image", "{}").split('"')
                if len(parts) > 1:
                    image_url = parts[1]

        # 評価・レビュー
        rating = ""
        rating_el = soup.select_one("span[data-hook='rating-out-of-text'], #acrPopover .a-icon-alt")
        if rating_el:
            m2 = re.search(r"(\d+\.?\d*)", rating_el.get_text(strip=True))
            if m2:
                rating = m2.group(1)

        review_count = ""
        rev_el = soup.select_one("#acrCustomerReviewText, span[data-hook='total-review-count']")
        if rev_el:
            review_count = rev_el.get_text(strip=True)

        # A+
        has_aplus = bool(soup.select_one("#aplus, #aplus3PModule, .aplus-v2"))

        # 箇条書き
        bullets = []
        for el in soup.select("#feature-bullets ul li span.a-list-item"):
            text = el.get_text(strip=True)
            if text and len(text) > 3:
                bullets.append(text)

        # 商品説明
        description = ""
        desc_el = soup.select_one("#productDescription p, #productDescription_feature_div p")
        if desc_el:
            description = desc_el.get_text(separator="\n", strip=True)

        # 仕様表
        specs = {}
        for row in soup.select(
            "#productDetails_techSpec_section_1 tr, "
            "#productDetails_techSpec_section_2 tr, "
            "#prodDetails .prodDetTable tr"
        ):
            th = row.select_one("th")
            td = row.select_one("td")
            if th and td:
                key = th.get_text(strip=True)
                val = td.get_text(strip=True)
                if key and val:
                    specs[key] = val

        if not specs:
            for li in soup.select("#detailBullets_feature_div ul li"):
                parts = [p.strip() for p in li.get_text(separator="\n", strip=True).split("\n")
                         if p.strip() and p.strip() != ":"]
                if len(parts) >= 2:
                    specs[parts[0]] = parts[1]

        return {
            "success":      True,
            "url":          current_url,
            "asin":         asin,
            "title":        title,
            "price":        price,
            "image_url":    image_url,
            "rating":       rating,
            "review_count": review_count,
            "has_aplus":    has_aplus,
            "bullets":      bullets,
            "description":  description,
            "specs":        specs,
        }

    except Exception as e:
        logger.error(f"Amazon取得エラー: {e}", exc_info=True)
        return {"success": False, "error": str(e)}
    finally:
        try:
            driver.quit()
        except Exception:
            pass


# ── 画像ダウンロード ────────────────────────────────────────────────────
def _download_image(url: str, save_dir: Path, asin: str) -> Optional[Path]:
    """画像をローカルに保存して Path を返す。失敗時は None。"""
    if not url:
        return None
    try:
        save_dir.mkdir(parents=True, exist_ok=True)
        ext = ".jpg"
        if ".png" in url.lower():
            ext = ".png"
        dest = save_dir / f"{asin}{ext}"
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=10) as resp:
            dest.write_bytes(resp.read())
        logger.info(f"画像保存: {dest}")
        return dest
    except Exception as e:
        logger.warning(f"画像ダウンロード失敗 ({url}): {e}")
        return None


# ── Excel 追記 ─────────────────────────────────────────────────────────
def _append_amazon_to_excel(excel_path: str, data: dict) -> dict:
    """
    指定した Excel ファイルの Sheet2/3 に Amazon データを追記する。
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        path = Path(excel_path)
        if not path.exists():
            return {"success": False, "error": f"ファイルが見つかりません: {excel_path}"}

        wb = load_workbook(str(path))

        # シート存在確認
        if SHEET_AMAZON_LIST not in wb.sheetnames:
            return {"success": False, "error": f"シート「{SHEET_AMAZON_LIST}」が見つかりません。正しいリサーチシートか確認してください。"}

        # ── Sheet2: ②Amazonライバル に追記 ────────────────────────
        ws2 = wb[SHEET_AMAZON_LIST]
        # ヘッダー行 = 2、データ開始 = 3（行3はメモなので4以降）
        # 最終行を探す（4行目以降で最初の空行）
        next_row = 4
        for r in range(4, ws2.max_row + 2):
            if ws2.cell(r, 1).value is None:
                next_row = r
                break

        thin = Side(style="thin", color="BFBFBF")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        row_data = [
            data.get("asin", ""),
            data.get("title", ""),
            data.get("price", ""),
            data.get("rating", ""),
            data.get("review_count", ""),
            "あり" if data.get("has_aplus") else "なし",
            data.get("url", ""),
            "",  # 画像（後で埋め込み）
        ]
        ws2.row_dimensions[next_row].height = 90  # 画像用に高さを確保

        for col, val in enumerate(row_data, 1):
            c = ws2.cell(next_row, col)
            c.value = val
            c.font = Font(name="BIZ UDGothic")
            c.border = border
            c.alignment = Alignment(vertical="center", wrap_text=(col == 2))

        # ── 画像ダウンロード & 埋め込み ──────────────────────────
        image_col_letter = get_column_letter(8)
        image_path = None
        if data.get("image_url"):
            img_dir = path.parent / IMAGE_FOLDER_NAME
            image_path = _download_image(data["image_url"], img_dir, data.get("asin", "img"))

        tmp_dir = tempfile.mkdtemp()
        try:
            if image_path and image_path.exists():
                # リサイズ
                try:
                    from PIL import Image as PILImage
                    img = PILImage.open(image_path).convert("RGBA")
                    img.thumbnail((100, 85), PILImage.LANCZOS)
                    bg = PILImage.new("RGB", img.size, (255, 255, 255))
                    bg.paste(img, mask=img.split()[3])
                    thumb_path = Path(tmp_dir) / image_path.name
                    bg.save(str(thumb_path), "JPEG", quality=85)

                    from openpyxl.drawing.image import Image as XLImage
                    xl_img = XLImage(str(thumb_path))
                    xl_img.anchor = f"{image_col_letter}{next_row}"
                    ws2.add_image(xl_img)
                    ws2.cell(next_row, 8).value = None
                except Exception as e:
                    logger.warning(f"画像埋め込みエラー: {e}")
                    ws2.cell(next_row, 8).value = data.get("image_url", "")

            # ── Sheet3: ③Amazonテキスト に追記 ──────────────────
            if SHEET_AMAZON_TEXT in wb.sheetnames:
                ws3 = wb[SHEET_AMAZON_TEXT]
                start_row = ws3.max_row + 2
                if start_row <= 3:
                    start_row = 4

                def write_text_row(r, label, value):
                    ws3.row_dimensions[r].height = max(18, min(120, len(str(value)) // 3 + 18))
                    ws3.cell(r, 1).value = label
                    ws3.cell(r, 1).font = Font(name="BIZ UDGothic", bold=True)
                    ws3.cell(r, 1).fill = PatternFill("solid", fgColor="E2EFDA")
                    ws3.cell(r, 2).value = value
                    ws3.cell(r, 2).font = Font(name="BIZ UDGothic")
                    ws3.cell(r, 2).alignment = Alignment(vertical="top", wrap_text=True)

                # セパレーター
                ws3.row_dimensions[start_row - 1].height = 8
                ws3.merge_cells(f"A{start_row}:B{start_row}")
                sep = ws3.cell(start_row, 1)
                sep.value = f"── {data.get('asin','')} {data.get('title','')[:40]} ──"
                sep.font = Font(name="BIZ UDGothic", bold=True, color="FFFFFF")
                sep.fill = PatternFill("solid", fgColor="375623")
                start_row += 1

                write_text_row(start_row,     "タイトル",     data.get("title", ""))
                write_text_row(start_row + 1, "価格",         data.get("price", ""))
                write_text_row(start_row + 2, "商品の特徴",   "\n".join(data.get("bullets", [])))
                write_text_row(start_row + 3, "商品説明",     data.get("description", ""))
                specs_text = "\n".join(f"{k}: {v}" for k, v in data.get("specs", {}).items())
                write_text_row(start_row + 4, "仕様・詳細",   specs_text)

            # 保存
            wb.save(str(path))
        finally:
            shutil.rmtree(tmp_dir, ignore_errors=True)

        return {
            "success":  True,
            "row":      next_row,
            "asin":     data.get("asin", ""),
            "title":    data.get("title", ""),
            "has_image": image_path is not None,
        }

    except Exception as e:
        logger.error(f"Excel追記エラー: {e}", exc_info=True)
        return {"success": False, "error": str(e)}


# ── Excel 情報取得 ──────────────────────────────────────────────────────
def _get_excel_info(excel_path: str) -> dict:
    """Excel ファイルの現在の状態を返す。"""
    try:
        from openpyxl import load_workbook
        path = Path(excel_path)
        if not path.exists():
            return {"success": False, "error": "ファイルが見つかりません"}

        wb = load_workbook(str(path), read_only=True, data_only=True)
        sheets = wb.sheetnames

        # Sheet1からタイトル取得
        title = ""
        if "①概要" in sheets:
            ws = wb["①概要"]
            title = ws["A1"].value or ""
            title = str(title).replace("リサーチシート　", "").strip()

        # Sheet2の追記件数
        amazon_count = 0
        if SHEET_AMAZON_LIST in sheets:
            ws2 = wb[SHEET_AMAZON_LIST]
            for r in range(4, ws2.max_row + 1):
                if ws2.cell(r, 1).value:
                    amazon_count += 1

        wb.close()
        return {
            "success":      True,
            "filename":     path.name,
            "title":        title,
            "sheets":       sheets,
            "amazon_count": amazon_count,
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


# ── HTML テンプレート ───────────────────────────────────────────────────
HTML = """<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>リサーチ追記ツール</title>
<style>
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body { font-family: "BIZ UDGothic", "Hiragino Sans", sans-serif;
         background: #f0f2f5; color: #333; min-height: 100vh; }

  .header { background: #1F4E79; color: white; padding: 16px 24px;
            display: flex; align-items: center; gap: 12px; }
  .header h1 { font-size: 18px; font-weight: bold; }
  .header .sub { font-size: 12px; opacity: 0.75; margin-top: 2px; }

  .main { max-width: 800px; margin: 24px auto; padding: 0 16px; }

  .card { background: white; border-radius: 10px; padding: 20px;
          box-shadow: 0 2px 8px rgba(0,0,0,0.08); margin-bottom: 16px; }
  .card h2 { font-size: 14px; font-weight: bold; color: #555;
             border-left: 4px solid #1F4E79; padding-left: 10px; margin-bottom: 14px; }

  .file-row { display: flex; gap: 8px; align-items: center; }
  .file-input { flex: 1; padding: 10px 12px; border: 2px solid #ddd;
                border-radius: 6px; font-size: 14px; transition: border-color 0.2s; }
  .file-input:focus { outline: none; border-color: #1F4E79; }

  .btn { padding: 10px 20px; border: none; border-radius: 6px; font-size: 14px;
         font-weight: bold; cursor: pointer; transition: all 0.2s; white-space: nowrap; }
  .btn:hover { filter: brightness(0.9); }
  .btn:active { transform: scale(0.97); }
  .btn-blue   { background: #1F4E79; color: white; }
  .btn-orange { background: #C55A11; color: white; }
  .btn-green  { background: #375623; color: white; }
  .btn-gray   { background: #ccc; color: #555; cursor: not-allowed; }

  .excel-info { background: #EBF3FB; border-radius: 8px; padding: 14px 16px; margin-top: 12px; }
  .excel-info .name { font-size: 15px; font-weight: bold; color: #1F4E79; }
  .excel-info .meta { font-size: 12px; color: #666; margin-top: 4px; }
  .excel-info .sheets { display: flex; gap: 6px; flex-wrap: wrap; margin-top: 8px; }
  .sheet-tag { background: #1F4E79; color: white; font-size: 11px;
               padding: 3px 8px; border-radius: 12px; }

  .result { margin-top: 12px; padding: 12px 14px; border-radius: 8px; font-size: 13px; }
  .result.success { background: #E9F7EF; border: 1px solid #27AE60; color: #1a6b3a; }
  .result.error   { background: #FDE8E8; border: 1px solid #E74C3C; color: #a11a1a; }
  .result.info    { background: #EBF5FB; border: 1px solid #2980B9; color: #1a4a6b; }

  .spinner { display: inline-block; width: 16px; height: 16px;
             border: 2px solid rgba(255,255,255,0.4);
             border-top-color: white; border-radius: 50%;
             animation: spin 0.7s linear infinite; margin-right: 6px; vertical-align: middle; }
  @keyframes spin { to { transform: rotate(360deg); } }

  .hidden { display: none; }
  .row-count { display: inline-block; background: #C55A11; color: white;
               font-size: 11px; padding: 2px 8px; border-radius: 10px; margin-left: 8px; }
</style>
</head>
<body>
<div class="header">
  <div>
    <div class="h1" style="font-size:18px; font-weight:bold;">📊 リサーチ追記ツール</div>
    <div class="sub">aucfantool から独立したスタンドアロンツール</div>
  </div>
</div>

<div class="main">

  <!-- Excel読み込み -->
  <div class="card">
    <h2>① Excel ファイルを開く</h2>
    <div class="file-row">
      <input id="filePath" class="file-input" type="text"
             placeholder="/Users/shino/... リサーチシート名 ...xlsx">
      <button class="btn btn-blue" onclick="loadExcel()">読み込む</button>
    </div>
    <div id="excelInfo" class="hidden"></div>
  </div>

  <!-- Amazon取得 -->
  <div class="card" id="amazonCard">
    <h2>② Amazon ライバルデータ取得
      <span id="amazonCount" class="row-count hidden">0件</span>
    </h2>
    <p style="font-size:13px; color:#666; margin-bottom:12px;">
      Chrome でライバルの Amazon 商品ページ（/dp/...）を開いてから「取得」を押してください。
    </p>
    <button id="btnAmazon" class="btn btn-orange" onclick="fetchAmazon()">
      🔍 Amazonデータ取得 → Excel Sheet2/3 に追記
    </button>
    <div id="amazonResult" class="hidden"></div>
  </div>

  <!-- 1688取得（将来実装） -->
  <div class="card" style="opacity:0.5;">
    <h2>③ 1688 仕入れデータ取得 <span style="font-size:11px; color:#999;">（準備中）</span></h2>
    <p style="font-size:13px; color:#666; margin-bottom:12px;">
      Chrome で 1688 商品ページを開いてから「取得」を押してください。
    </p>
    <button class="btn btn-gray" disabled>🔍 1688データ取得 → Excel Sheet4/5 に追記</button>
  </div>

</div>

<script>
let loadedPath = "";

async function loadExcel() {
  const p = document.getElementById("filePath").value.trim();
  if (!p) { alert("ファイルパスを入力してください"); return; }

  const res = await fetch("/api/excel/load", {
    method: "POST",
    headers: {"Content-Type": "application/json"},
    body: JSON.stringify({path: p})
  });
  const data = await res.json();
  const infoDiv = document.getElementById("excelInfo");

  if (data.success) {
    loadedPath = p;
    const countBadge = data.amazon_count > 0
      ? `<span style="color:#C55A11; font-weight:bold;">Amazon ${data.amazon_count}件取得済み</span>`
      : "Amazonデータ未取得";

    infoDiv.innerHTML = `
      <div class="excel-info">
        <div class="name">📄 ${data.filename}</div>
        <div class="meta">${data.title || ""} &nbsp;|&nbsp; ${countBadge}</div>
        <div class="sheets">
          ${(data.sheets || []).map(s => `<span class="sheet-tag">${s}</span>`).join("")}
        </div>
      </div>`;
    infoDiv.classList.remove("hidden");

    // Amazon件数バッジ更新
    const badge = document.getElementById("amazonCount");
    if (data.amazon_count > 0) {
      badge.textContent = data.amazon_count + "件";
      badge.classList.remove("hidden");
    }
  } else {
    infoDiv.innerHTML = `<div class="result error">❌ ${data.error}</div>`;
    infoDiv.classList.remove("hidden");
  }
}

async function fetchAmazon() {
  if (!loadedPath) { alert("先にExcelファイルを読み込んでください"); return; }

  const btn = document.getElementById("btnAmazon");
  const resultDiv = document.getElementById("amazonResult");
  btn.innerHTML = '<span class="spinner"></span>取得中...';
  btn.disabled = true;
  resultDiv.className = "result info";
  resultDiv.textContent = "Chromeのタブを確認しています...";
  resultDiv.classList.remove("hidden");

  try {
    const res = await fetch("/api/amazon/fetch-and-append", {
      method: "POST",
      headers: {"Content-Type": "application/json"},
      body: JSON.stringify({path: loadedPath})
    });
    const data = await res.json();

    if (data.success) {
      const imgMsg = data.has_image ? "（画像あり）" : "（画像なし）";
      resultDiv.className = "result success";
      resultDiv.innerHTML = `
        ✅ 追記完了 ${imgMsg}<br>
        <strong>ASIN:</strong> ${data.asin}<br>
        <strong>タイトル:</strong> ${data.title}<br>
        <span style="font-size:12px; color:#666;">Sheet2 の ${data.row} 行目に追記しました</span>`;

      // 件数バッジ更新
      const badge = document.getElementById("amazonCount");
      const cur = parseInt(badge.textContent) || 0;
      badge.textContent = (cur + 1) + "件";
      badge.classList.remove("hidden");
    } else {
      resultDiv.className = "result error";
      resultDiv.textContent = "❌ " + data.error;
    }
  } catch(e) {
    resultDiv.className = "result error";
    resultDiv.textContent = "❌ 通信エラー: " + e.message;
  } finally {
    btn.innerHTML = "🔍 Amazonデータ取得 → Excel Sheet2/3 に追記";
    btn.disabled = false;
  }
}

// Enterキーで読み込み
document.getElementById("filePath").addEventListener("keydown", e => {
  if (e.key === "Enter") loadExcel();
});
</script>
</body>
</html>
"""


# ── Flask ルート ────────────────────────────────────────────────────────
@app.route("/")
def index():
    return render_template_string(HTML)


@app.route("/api/excel/load", methods=["POST"])
def api_load_excel():
    data = request.get_json() or {}
    path = data.get("path", "").strip()
    if not path:
        return jsonify({"success": False, "error": "パスが指定されていません"})
    info = _get_excel_info(path)
    return jsonify(info)


@app.route("/api/amazon/fetch-and-append", methods=["POST"])
def api_amazon_fetch_and_append():
    data = request.get_json() or {}
    excel_path = data.get("path", "").strip()
    if not excel_path:
        return jsonify({"success": False, "error": "Excelパスが指定されていません"})

    # Chrome から Amazon データ取得
    amazon_data = _fetch_amazon_from_chrome()
    if not amazon_data.get("success"):
        return jsonify(amazon_data)

    # Excel に追記
    result = _append_amazon_to_excel(excel_path, amazon_data)
    return jsonify(result)


@app.route("/api/status", methods=["GET"])
def api_status():
    return jsonify({"status": "ok", "port": PORT, "tool": "research_tool"})


# ── 起動 ────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("=" * 50)
    print("📊 リサーチ追記ツール")
    print(f"   http://localhost:{PORT}")
    print("=" * 50)
    app.run(host="0.0.0.0", port=PORT, debug=False)
