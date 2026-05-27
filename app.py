"""
app.py — AucFan リサーチツール メインエントリーポイント

【役割】
  - Flask Webアプリ起動（ポート 5001、0.0.0.0 バインドで iPhone/iPad からも接続可能）
  - AucFanScraper / SellerAnalyzer をバックグラウンドスレッドで実行
  - スクレイピング進捗を SSE（Server-Sent Events）でリアルタイム配信
  - STEP 1（キーワードリサーチ）/ STEP 2（セラーリサーチ）/ STEP 3（マスターリサーチ）
    の3ステップ分の API エンドポイントを提供

【グローバル状態】
  _data_manager      : STEP 1 用 DataManager（スクレイパースレッドと共有）
  _scraper_thread    : STEP 1 スクレイピングスレッド
  _stop_event        : 停止シグナル（set() で全スレッドに停止を通知）
  _seller_state      : STEP 2（セラーリサーチ）の実行状態
  _master_state      : STEP 3（マスターセラーリサーチ）の実行状態
  _sellers_master    : SellersMaster シングルトン（data/sellers_master.json を管理）

【werkzeug ログ抑制】
  werkzeug のアクセスログ（ポーリング系 GET /api/... 200 の大量ログ）は
  Warning レベルに抑制済み。スクレイピング完了後は logger.info で
  「待機中」メッセージをターミナルに表示する。
"""
import json
import logging
import os
import sys
import threading
import time
import webbrowser
from datetime import datetime
from pathlib import Path

import io
import shutil
import pandas as pd
from flask import (
    Flask, Response, jsonify, render_template, request,
    send_from_directory, abort
)

import config
from data_manager import DataManager, make_output_dir, make_session_id
from image_processor import ImageProcessor
from gemini_client import GeminiClient, get_rate_limit_status, reset_rate_limit_flag
from scraper import AucFanScraper
from seller_analyzer import SellerAnalyzer
from sellers_master import SellersMaster

# ─────────────────────────────────────────────
# ロギング設定
# ─────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="[%(asctime)s] %(levelname)s %(name)s: %(message)s",
    datefmt="%H:%M:%S",
    handlers=[
        logging.StreamHandler(sys.stdout),
    ]
)
logger = logging.getLogger(__name__)

# werkzeug のアクセスログを抑制（ポーリング系の大量ログを非表示にする）
logging.getLogger("werkzeug").setLevel(logging.WARNING)

# ─────────────────────────────────────────────
# Flask アプリ
# ─────────────────────────────────────────────
app = Flask(__name__, template_folder="templates", static_folder="static")
app.config["SECRET_KEY"] = os.urandom(24)

# ─── STEP 1: キーワードリサーチ グローバル状態 ───
# スクレイパースレッドと Flask ルートハンドラが同一オブジェクトを共有する。
# 操作は _lock でスレッドセーフに保護する。
_scraper_thread: threading.Thread = None
_stop_event = threading.Event()          # set() でスクレイパーに停止を通知
_data_manager: DataManager = None        # 現在アクティブな DataManager
_image_processor: ImageProcessor = None  # 画像ダウンロード・pHash 計算
_gemini_client: GeminiClient = None      # Gemini API クライアント
_session_output_dir: Path = None         # 現セッションの出力ディレクトリ
_lock = threading.Lock()
# ログイン即時チェックトリガー（UIの「今すぐ確認」ボタンから set() する）
# STEP1/2/3 すべてのスクレイパーで共有する
_login_check_event = threading.Event()

# ─── STEP 2: セラーリサーチ グローバル状態 ───
# SellerAnalyzer スレッドと Flask ルートハンドラが共有する辞書。
# 操作は _seller_lock でスレッドセーフに保護する。
_seller_state = {
    "sellers": [],        # [{"seller_id": str, "seller_url": str, "status": "pending"|"running"|"done"|"error"}]
    "current_index": -1,  # 現在処理中のセラーインデックス（-1 = 未開始）
    "running": False,     # SellerAnalyzer スレッドが実行中かどうか
    "phase": "idle",      # "idle"|"scraping_list"|"grouping"|"vision_check"|"done"|"stopped"|"error"
    "stop_event": None,   # threading.Event（実行中のみ有効、停止リクエスト用）
    "thread": None,       # SellerAnalyzer スレッド
    "dm": None,           # 実行中の DataManager（/api/seller/status でのポーリング用）
    "session_id": None,   # 完了セッション ID（履歴表示用）
    "output_dir": None,   # 完了セッション出力ディレクトリ（Path）
    "source_keyword": "", # 元になった STEP1 キーワード名（セッション履歴の表示に使用）
}
_seller_lock = threading.Lock()

# ─── STEP 3: マスターセラーリサーチ グローバル状態 ───
# _seller_state と同じ構造。SellersMaster からセラーを自動取得して
# SellerAnalyzer を順次実行する。
_master_state = {
    "running": False,
    "phase": "idle",       # "idle"|"scraping_list"|"grouping"|"vision_check"|"done"|"stopped"|"error"
    "stop_event": None,    # threading.Event（実行中のみ有効）
    "thread": None,        # SellerAnalyzer スレッド
    "dm": None,            # 実行中の DataManager
    "session_id": None,
    "output_dir": None,
    "total": 0,
    "done": 0,
    "current_seller": "",
}
_master_lock = threading.Lock()
_sellers_master = SellersMaster()   # シングルトン


def get_dm() -> DataManager:
    """
    グローバルな DataManager インスタンスを返すヘルパー。

    スクレイピングスレッドと Flask ルートハンドラが同一の DataManager を
    共有するためのシングルトンアクセスポイント。
    読み取り専用アクセスのためロック不要で呼び出せる。
    スクレイピング未開始・データ未ロード時は None を返す。
    """
    global _data_manager
    return _data_manager


# ─────────────────────────────────────────────
# ルート
# ─────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/images/<path:filename>")
def serve_image(filename):
    """
    保存済み画像を配信。
    構成: LOCAL_IMAGE_CACHE_DIR / セッション名 / images / ファイル名
    """
    if _session_output_dir is not None:
        session_images = Path(config.LOCAL_IMAGE_CACHE_DIR) / _session_output_dir.name / "images"
        if (session_images / filename).exists():
            return send_from_directory(str(session_images), filename)
    abort(404)


# ─────────────────────────────────────────────
# スクレイピング制御
# ─────────────────────────────────────────────

@app.route("/api/start", methods=["POST"])
def api_start():
    """スクレイピング開始"""
    global _scraper_thread, _stop_event, _data_manager, _image_processor
    global _gemini_client, _session_output_dir

    with _lock:
        if _scraper_thread and _scraper_thread.is_alive():
            return jsonify({"success": False, "message": "スクレイピングは既に実行中です"}), 400

        data = request.get_json(silent=True) or {}
        keyword   = data.get("keyword", "unknown")
        resume    = data.get("resume", False)
        start_url = data.get("start_url", "").strip()  # iPhoneから貼り付けたURL（任意）

        # セッション初期化 (S1_YYYYMMDD_NN_keyword)
        out_dir, session_id = make_output_dir(keyword, step=1)
        _session_output_dir = out_dir
        _data_manager = DataManager(session_id, out_dir)
        _image_processor = ImageProcessor(Path(config.LOCAL_IMAGE_CACHE_DIR) / out_dir.name / "images")
        _gemini_client = GeminiClient()

        # 再開の場合は前回データをロード
        if resume:
            _data_manager.load_previous_session()

        _stop_event.clear()

        def run_scraper():
            scraper = AucFanScraper(
                data_manager=_data_manager,
                image_processor=_image_processor,
                gemini_client=_gemini_client,
                stop_event=_stop_event,
                login_check_event=_login_check_event,
            )
            scraper.run(resume=resume, start_url=start_url or None)
            # STEP1完了後にCSV・HTML・PDFを自動保存
            final_status = _data_manager.get_progress().get("status", "done")
            if final_status in ("done", "stopped"):
                _save_export_files(_data_manager, out_dir)

        _scraper_thread = threading.Thread(target=run_scraper, daemon=True, name="scraper")
        _scraper_thread.start()

        logger.info(f"スクレイピング開始: keyword={keyword}, session={session_id}, start_url={start_url or '（現在タブ）'}")
        return jsonify({
            "success": True,
            "session_id": session_id,
            "output_dir": str(out_dir),
            "resume": resume,
            "start_url": start_url or None,
        })


@app.route("/api/stop", methods=["POST"])
def api_stop():
    """スクレイピングを停止"""
    _stop_event.set()
    logger.info("停止リクエストを受け取りました")
    return jsonify({"success": True, "message": "停止中..."})


@app.route("/api/resume", methods=["POST"])
def api_resume():
    """前回セッションを選択して再開準備"""
    sessions = _list_sessions()
    return jsonify({"sessions": sessions})


# ─────────────────────────────────────────────
# SSE（リアルタイム進捗）
# ─────────────────────────────────────────────

@app.route("/api/stream")
def api_stream():
    """Server-Sent Events で進捗をリアルタイム配信"""
    def generate():
        # SSEペイロード構造（クライアントの updateProgressUI が期待するフィールド）:
        #   progress          : DataManager.get_progress() の戻り値（dict）
        #     .status         : 'idle'|'scraping_list'|'scraping_detail'|'grouping'|
        #                       'vision_check'|'done'|'stopped'|'error'
        #     .keyword        : スクレイピング対象キーワード
        #     .pages_done     : 取得済み一覧ページ数
        #     .total_pages    : 推定総ページ数
        #     .total_items    : 取得済み商品数（累計）
        #     .detail_pages_done  : 詳細取得済み件数
        #     .detail_pages_total : 詳細取得対象の総件数
        #     .candidates_found   : 仕入れ候補として検出された件数
        #     .processed_items    : 処理（判定）済みアイテム数
        #   stats             : DataManager.get_stats() の戻り値（dict）
        #     .total          : 全アイテム数
        #     .by_status      : {candidate, next_candidate, ok, ng, review, waiting} 各件数
        #   is_running        : スクレイパースレッドが生存中かどうか (bool)
        #   gemini_enabled    : Gemini API が有効かどうか (bool)
        #   is_seller_analysis: セラー分析モード中かどうか（keyword == 'seller_analysis'）(bool)
        last_total = -1
        while True:
            try:
                dm = get_dm()
                if dm:
                    progress = dm.get_progress()
                    stats = dm.get_stats()
                    is_running = _scraper_thread is not None and _scraper_thread.is_alive()

                    payload = {
                        "progress": progress,
                        "stats": stats,
                        "is_running": is_running,
                        "gemini_enabled": _gemini_client.available if _gemini_client else False,
                        "is_seller_analysis": progress.get("keyword", "") == "seller_analysis",
                    }
                    yield f"data: {json.dumps(payload, ensure_ascii=False)}\n\n"
                else:
                    yield f"data: {json.dumps({'progress': {'status': 'idle'}, 'is_running': False})}\n\n"
            except GeneratorExit:
                break
            except Exception as e:
                logger.debug(f"SSEエラー: {e}")
            time.sleep(1.5)

    return Response(
        generate(),
        mimetype="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
        }
    )


# ─────────────────────────────────────────────
# データ取得 API
# ─────────────────────────────────────────────

@app.route("/api/progress")
def api_progress():
    """現在の進捗を返す"""
    dm = get_dm()
    if not dm:
        return jsonify({"status": "idle", "total_items": 0})
    return jsonify({
        "progress": dm.get_progress(),
        "stats": dm.get_stats(),
        "is_running": _scraper_thread is not None and _scraper_thread.is_alive(),
    })


@app.route("/api/items")
def api_items():
    """商品一覧をグループ形式で返す"""
    dm = get_dm()
    if not dm:
        return jsonify({"groups": []})

    # フィルターパラメータ
    keyword = request.args.get("keyword", "")
    status = request.args.get("status", "")
    min_price = int(request.args.get("min_price", 0))
    max_price = int(request.args.get("max_price", 99999))
    min_group = int(request.args.get("min_group", 0))
    page = int(request.args.get("page", 1))
    per_page = int(request.args.get("per_page", 30))

    filtered = dm.get_items_filtered(
        keyword=keyword,
        status=status,
        min_price=min_price,
        max_price=max_price,
        min_group=min_group,
    )

    # グループ化
    group_map = {}
    for item in filtered:
        gid = item.get("group_id") or item["item_id"]
        group_map.setdefault(gid, []).append(item)

    # グループサイズで降順ソート
    groups = sorted(group_map.values(), key=lambda g: len(g), reverse=True)

    # ページネーション
    total_groups = len(groups)
    start = (page - 1) * per_page
    end = start + per_page
    page_groups = groups[start:end]

    # シリアライズ用に整形
    result_groups = []
    for group in page_groups:
        result_groups.append({
            "group_id": group[0].get("group_id") or group[0]["item_id"],
            "count": len(group),
            "items": group,
            "status": group[0].get("status", ""),
            "title": (group[0].get("title_full") or group[0].get("title_short", ""))[:100],
            "min_price": min((int(i.get("price") or 0) for i in group if int(i.get("price") or 0) > 0), default=0),
            "max_price": max((int(i.get("price") or 0) for i in group if int(i.get("price") or 0) > 0), default=0),
            "seller_ids": list(set(i.get("seller_id", "") for i in group if i.get("seller_id"))),
        })

    progress = dm.get_progress()
    return jsonify({
        "groups": result_groups,
        "total_groups": total_groups,
        "page": page,
        "per_page": per_page,
        "total_pages": (total_groups + per_page - 1) // per_page,
        "is_seller_analysis": progress.get("keyword", "") == "seller_analysis",
        "seller_detail_min_group": config.SELLER_DETAIL_MIN_GROUP,
    })


@app.route("/api/item/<item_id>")
def api_item_detail(item_id):
    """個別商品の詳細"""
    dm = get_dm()
    if not dm:
        abort(404)
    item = dm.get_item(item_id)
    if not item:
        abort(404)
    return jsonify(item)


# ─────────────────────────────────────────────
# ステータス更新 API
# ─────────────────────────────────────────────

@app.route("/api/item/<item_id>/status", methods=["POST"])
def api_update_status(item_id):
    """商品ステータスを更新（OK / NG / etc.）"""
    dm = get_dm()
    if not dm:
        return jsonify({"success": False}), 400

    data = request.get_json(silent=True) or {}
    new_status = data.get("status")

    valid_statuses = [config.STATUS_OK, config.STATUS_NG, config.STATUS_CANDIDATE,
                      config.STATUS_WAITING, config.STATUS_REVIEW]
    if new_status not in valid_statuses:
        return jsonify({"success": False, "message": f"無効なステータス: {new_status}"}), 400

    dm.update_status(item_id, new_status)

    # グループ全体に反映する場合
    apply_group = data.get("apply_group", False)
    if apply_group:
        item = dm.get_item(item_id)
        if item and item.get("group_id"):
            all_items = dm.get_all_items()
            for i in all_items:
                if i.get("group_id") == item["group_id"]:
                    dm.update_status(i["item_id"], new_status)

    dm.save_csv()
    return jsonify({"success": True, "item_id": item_id, "status": new_status})


@app.route("/api/group/<group_id>/status", methods=["POST"])
def api_update_group_status(group_id):
    """グループ全体のステータスを更新"""
    dm = get_dm()
    if not dm:
        return jsonify({"success": False}), 400

    data = request.get_json(silent=True) or {}
    new_status = data.get("status")
    ng_reason = data.get("ng_reason", "").strip()  # 手動NG理由（任意）

    all_items = dm.get_all_items()
    updated = 0
    for item in all_items:
        if item.get("group_id") == group_id or item["item_id"] == group_id:
            updates = {"status": new_status}
            # NG理由が入力された場合は exclude_reason に保存
            if new_status == "ng" and ng_reason:
                updates["exclude_reason"] = ng_reason
                updates["gemini_source"] = "manual"
            dm.update_item(item["item_id"], updates)
            updated += 1

    dm.save_csv()
    return jsonify({"success": True, "updated": updated, "ng_reason": ng_reason})


@app.route("/api/ng/analyze", methods=["POST"])
def api_ng_analyze():
    """
    手動NG理由テキストをGeminiで分析し、除外ルール整理情報を返す。
    Request:  { "reason": "フライパンなのでNG" }
    Response: { "category": "衛生リスク商品", "explanation": "...", "keywords": [...] }
    """
    data = request.get_json(silent=True) or {}
    reason = data.get("reason", "").strip()
    if not reason:
        return jsonify({"error": "理由テキストが空です"}), 400

    gc = _gemini_client
    if gc is None or not gc.available:
        return jsonify({"error": "Gemini API が無効です"}), 503

    result = gc.analyze_ng_reason(reason)
    if result is None:
        return jsonify({"error": "Gemini分析に失敗しました"}), 500

    return jsonify(result)


# ─────────────────────────────────────────────
# CSV エクスポート
# ─────────────────────────────────────────────

def _generate_export_html(dm, images_dir: Path) -> str:
    """
    DataManager と images_dir から HTML エクスポート文字列を生成する共通ヘルパー。
    api_export_html（ブラウザ経由）と _save_export_files（自動保存）の両方から呼ばれる。
    """
    import base64
    from collections import defaultdict

    items = dm.get_all_items()
    if not items:
        return ""

    # グループ化
    group_map = defaultdict(list)
    for item in items:
        gid = item.get("group_id") or item["item_id"]
        group_map[gid].append(item)
    groups = sorted(group_map.values(), key=lambda g: len(g), reverse=True)

    # 画像を base64 に変換
    def encode_image(local_path):
        if not local_path or not images_dir:
            return ""
        try:
            img_path = images_dir / Path(local_path).name
            if img_path.exists():
                with open(img_path, "rb") as f:
                    raw = f.read()
                ext = img_path.suffix.lower().lstrip(".")
                mime = {
                    "jpg": "image/jpeg", "jpeg": "image/jpeg",
                    "png": "image/png", "gif": "image/gif",
                    "webp": "image/webp",
                }.get(ext, "image/jpeg")
                return f"data:{mime};base64,{base64.b64encode(raw).decode()}"
        except Exception as e:
            logger.debug(f"画像base64変換エラー: {e}")
        return ""

    # グループデータを整形
    groups_data = []
    for group in groups:
        first = group[0]
        thumbs = [encode_image(item.get("thumbnail_local", "")) for item in group[:5]]
        thumbs = [t for t in thumbs if t]

        price    = first.get("price", 0)
        shipping = first.get("shipping", 0)
        total    = first.get("total", price + shipping)
        seller_ids = list({i.get("seller_id", "") for i in group if i.get("seller_id")})

        groups_data.append({
            "count":      len(group),
            "status":     first.get("status", "waiting"),
            "title":      (first.get("title_full") or first.get("title_short", ""))[:120],
            "price":      price,
            "shipping":   shipping,
            "total":      total,
            "seller_ids": seller_ids[:5],
            "thumbs":     thumbs,
            "url":        first.get("url", ""),
        })

    progress    = dm.get_progress()
    keyword     = progress.get("keyword", "")
    exported_at = datetime.now().strftime("%Y年%m月%d日 %H:%M")
    return _build_export_html(groups_data, keyword, exported_at, len(items), len(groups_data))


def _save_export_files(dm, output_dir: Path):
    """
    CSV / Mac用HTML / PDF をセッションフォルダと Google Drive に自動保存する。
    セラー分析完了時・キーワードスクレイピング完了時に呼ばれる。

    Google Drive 保存先:
      ~/Library/CloudStorage/GoogleDrive-shinozakistore@gmail.com/マイドライブ/AucFanToolData/
      セッション名_Mac表示用.html  ← ブラウザで開く用
      セッション名_仕入れ候補.pdf  ← iPhone の「ファイル」アプリで開く用
    """
    import shutil

    # Google Drive 保存先フォルダ（config._GDRIVE_ROOT と同じパス）
    _GDRIVE_DIR = Path(os.path.expanduser(
        "~/Library/CloudStorage/"
        "GoogleDrive-shinozakistore@gmail.com/"
        "マイドライブ/AucFanToolData"
    ))

    session_name = output_dir.name   # 例: S1_20260508_01_LEDライト
    images_dir   = Path(config.LOCAL_IMAGE_CACHE_DIR) / session_name / "images"

    # ── CSV ──
    try:
        dm.save_csv()
        logger.info(f"自動CSV保存: {output_dir / 'results.csv'}")
    except Exception as e:
        logger.error(f"CSV自動保存エラー: {e}")

    # ── Mac用HTML（セッションフォルダ + Google Drive） ──
    try:
        html_mac = _generate_export_html(dm, images_dir)
        if html_mac:
            # セッションフォルダに保存
            (output_dir / "result.html").write_text(html_mac, encoding="utf-8")
            logger.info(f"自動HTML保存(Mac用): {output_dir / 'result.html'}")
            # Google Drive にコピー
            _gdrive_copy_html(
                html_mac,
                _GDRIVE_DIR / f"{session_name}_Mac表示用.html",
                label="Mac表示用"
            )
    except Exception as e:
        logger.error(f"HTML自動保存エラー(Mac用): {e}")

    # iPhone用HTMLは廃止。iPhone向け出力はPDFに一本化。

    # ── PDF（仕入れ候補・次期候補 / 件数降順→価格降順） ──
    try:
        from pdf_exporter import generate_pdf
        pdf_bytes = generate_pdf(dm, output_dir)
        if pdf_bytes:
            # セッションフォルダに保存
            pdf_path = output_dir / "result.pdf"
            pdf_path.write_bytes(pdf_bytes)
            logger.info(f"自動PDF保存: {pdf_path}")
            # Google Drive にPDFを保存
            # scraper Mac（十王）: GDrive API で直接アップロード
            # reader  Mac（守谷）: ミラーリング経由でファイル書き込み
            if config.SITE_ROLE == "scraper" and config.GDRIVE_UPLOAD_ENABLED:
                try:
                    import gdrive_uploader
                    import tempfile
                    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tf:
                        tf.write(pdf_bytes)
                        tmp_path = Path(tf.name)
                    folder_id = gdrive_uploader.get_or_create_folder_path(
                        ["AucFanToolData", "リサーチ結果", session_name]
                    )
                    if folder_id:
                        gdrive_uploader.upload_file(
                            tmp_path, folder_id, f"{session_name}_仕入れ候補.pdf"
                        )
                        logger.info(f"Google Drive PDF保存(API): {session_name}_仕入れ候補.pdf")
                    tmp_path.unlink(missing_ok=True)
                except Exception as e2:
                    logger.warning(f"Google Drive PDF保存(API)スキップ: {e2}")
            else:
                gdrive_pdf = _GDRIVE_DIR / f"{session_name}_仕入れ候補.pdf"
                try:
                    gdrive_pdf.parent.mkdir(parents=True, exist_ok=True)
                    gdrive_pdf.write_bytes(pdf_bytes)
                    logger.info(f"Google Drive PDF保存: {gdrive_pdf}")
                except Exception as e2:
                    logger.warning(f"Google Drive PDF保存スキップ: {e2}")
        else:
            logger.info("PDF出力スキップ（候補グループなし）")
    except Exception as e:
        logger.error(f"PDF自動保存エラー: {e}")


def _gdrive_copy_html(html_content: str, dest_path: Path, label: str = ""):
    """
    HTML 文字列を Google Drive の指定パスに書き込む。
    Google Drive フォルダが存在しない場合は自動作成する。
    エラーが発生しても例外を握り潰してメインフローを止めない。
    """
    try:
        dest_path.parent.mkdir(parents=True, exist_ok=True)
        dest_path.write_text(html_content, encoding="utf-8")
        logger.info(f"Google Drive 保存({label}): {dest_path}")
    except Exception as e:
        logger.warning(f"Google Drive 保存スキップ({label}): {e}")


@app.route("/api/export/html")
def api_export_html():
    """スタンドアロンHTMLエクスポート（Mac用・ブラウザで開く用）"""
    from urllib.parse import quote

    dm = get_dm()
    if not dm:
        return jsonify({"error": "データがありません"}), 400

    images_dir = (
        Path(config.LOCAL_IMAGE_CACHE_DIR) / _session_output_dir.name / "images"
        if _session_output_dir else Path(config.LOCAL_IMAGE_CACHE_DIR)
    )
    html = _generate_export_html(dm, images_dir)
    if not html:
        return jsonify({"error": "商品データがありません"}), 400

    # セッションIDをファイル名に使う（例: S1_20260507_01_バフ_Mac用.html）
    if _session_output_dir:
        session_id = _session_output_dir.name
    else:
        keyword = dm.get_progress().get("keyword", "")
        safe_kw = "".join(c for c in keyword if c.isalnum() or c in ("_", "-"))[:20] or "result"
        session_id = f"aucfan_{safe_kw}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    filename = f"{session_id}_Mac用.html"

    response = Response(html, mimetype="text/html; charset=utf-8")
    response.headers["Content-Disposition"] = (
        f"attachment; filename*=UTF-8''{quote(filename)}"
    )
    return response


@app.route("/api/export/pdf")
def api_export_pdf():
    """
    PDF エクスポート（手動ダウンロード用）。
    仕入れ候補・次期候補を件数降順→価格降順で A4 縦 2 カラム PDF に書き出す。
    """
    from urllib.parse import quote
    from pdf_exporter import generate_pdf

    dm = get_dm()
    if not dm:
        return jsonify({"error": "データがありません"}), 400

    # 画像は現在のセッションフォルダ内 images/ を参照
    session_dir = _session_output_dir
    if session_dir is None:
        return jsonify({"error": "セッションが読み込まれていません"}), 400

    pdf_bytes = generate_pdf(dm, session_dir)
    if not pdf_bytes:
        return jsonify({"error": "出力対象グループがありません（候補・次期候補 0件）"}), 400

    session_name = session_dir.name
    filename = f"{session_name}_仕入れ候補.pdf"

    response = Response(pdf_bytes, mimetype="application/pdf")
    response.headers["Content-Disposition"] = (
        f"attachment; filename*=UTF-8''{quote(filename)}"
    )
    return response


@app.route("/api/export/excel/<group_id>", methods=["POST"])
def api_export_excel_single(group_id):
    """
    指定した group_id 1件分を Excel リサーチシートとして生成し、
    リサーチ結果フォルダ直下に保存する。
    ファイル名 = 商品タイトル_リサーチ.xlsx
    """
    from excel_exporter import generate_excel_single_with_session

    dm = get_dm()
    if not dm:
        return jsonify({"error": "データがありません"}), 400

    if _session_output_dir is None:
        return jsonify({"error": "セッションが読み込まれていません"}), 400

    session_name = _session_output_dir.name
    result = generate_excel_single_with_session(dm, group_id, session_name, embed_images=True)
    if not result:
        return jsonify({"error": "グループが見つかりません"}), 404

    excel_bytes, filename = result

    # ── 商品名フォルダを作成してその中にExcelを保存 ──────────────────
    # 例: リサーチ結果/商品名/商品名_リサーチ.xlsx
    # フォルダ名 = xlsx ファイル名から "_リサーチ.xlsx" を除いた部分
    from excel_exporter import sanitize_filename
    folder_name = filename.replace("_リサーチ.xlsx", "")
    product_dir = Path(config.OUTPUT_BASE_DIR) / folder_name
    save_path = product_dir / filename
    try:
        product_dir.mkdir(parents=True, exist_ok=True)
        # amazon/ と 1688/ サブフォルダも先に作っておく
        (product_dir / "amazon").mkdir(exist_ok=True)
        (product_dir / "1688").mkdir(exist_ok=True)
        save_path.write_bytes(excel_bytes)
        logger.info(f"Excel(単品)保存: {save_path}")
    except Exception as e:
        logger.error(f"Excel(単品)保存エラー: {e}")
        return jsonify({"error": f"保存に失敗しました: {e}"}), 500

    return jsonify({"success": True, "filename": filename, "folder": str(product_dir)})


def _build_export_html(groups_data, keyword, exported_at, total_items, total_groups):
    """iPhone/iPad 向けスタンドアロン HTML を生成する"""
    import json
    groups_json = json.dumps(groups_data, ensure_ascii=False)

    return f"""<!DOCTYPE html>
<html lang="ja">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>AucFan リサーチ結果 - {keyword}</title>
  <style>
    *, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}
    :root {{
      --primary: #2563eb; --primary-dark: #1d4ed8;
      --success: #16a34a; --warning: #d97706;
      --danger: #dc2626; --review: #7c3aed;
      --gray-50: #f9fafb; --gray-100: #f3f4f6;
      --gray-200: #e5e7eb; --gray-300: #d1d5db;
      --gray-600: #4b5563; --gray-700: #374151;
      --gray-900: #111827;
    }}
    body {{
      font-family: -apple-system, BlinkMacSystemFont, 'Hiragino Sans', sans-serif;
      font-size: 14px; background: var(--gray-50);
      color: var(--gray-900); line-height: 1.5;
    }}
    /* ─ ヘッダー ─ */
    .header {{
      background: var(--primary); color: #fff;
      position: sticky; top: 0; z-index: 100;
      padding: 10px 16px; box-shadow: 0 2px 6px rgba(0,0,0,.2);
    }}
    .header h1 {{ font-size: 16px; font-weight: 700; }}
    .header-meta {{ font-size: 11px; opacity: .8; margin-top: 2px; }}
    /* ─ フィルタータブ ─ */
    .filter-bar {{
      background: #fff; border-bottom: 1px solid var(--gray-200);
      padding: 8px 12px; overflow-x: auto; white-space: nowrap;
    }}
    .filter-bar button {{
      display: inline-block; padding: 5px 14px; margin-right: 6px;
      border: 1px solid var(--gray-300); border-radius: 20px;
      background: #fff; font-size: 13px; cursor: pointer; white-space: nowrap;
    }}
    .filter-bar button.active {{
      background: var(--primary); color: #fff; border-color: var(--primary);
    }}
    /* ─ 統計バー ─ */
    .stats {{
      background: #fff; padding: 8px 16px;
      display: flex; gap: 10px; flex-wrap: wrap;
      border-bottom: 1px solid var(--gray-200); font-size: 12px;
    }}
    .stat {{ text-align: center; }}
    .stat-num {{ font-size: 18px; font-weight: 700; }}
    .stat-label {{ color: var(--gray-600); }}
    /* ─ 検索バー ─ */
    .search-bar {{
      padding: 8px 12px; background: var(--gray-50);
      border-bottom: 1px solid var(--gray-200);
    }}
    .search-bar input {{
      width: 100%; padding: 8px 12px;
      border: 1px solid var(--gray-300); border-radius: 20px;
      font-size: 14px; outline: none;
    }}
    /* ─ グリッド ─ */
    .grid {{
      display: grid;
      grid-template-columns: repeat(auto-fill, minmax(320px, 1fr));
      gap: 12px; padding: 12px;
    }}
    @media (max-width: 480px) {{ .grid {{ grid-template-columns: 1fr; }} }}
    /* ─ カード ─ */
    .card {{
      background: #fff; border-radius: 10px;
      box-shadow: 0 1px 3px rgba(0,0,0,.1);
      overflow: hidden; border: 2px solid transparent;
    }}
    .card.candidate {{ border-color: var(--primary); }}
    .card.ok         {{ border-color: var(--success); }}
    .card.review     {{ border-color: var(--review); }}
    .card.ng         {{ opacity: .5; }}
    /* ステータスバー */
    .card-status {{
      padding: 5px 12px; font-size: 11px; font-weight: 700;
      color: #fff; display: flex; justify-content: space-between;
    }}
    .card-status.candidate {{ background: var(--primary); }}
    .card-status.waiting   {{ background: var(--gray-600); }}
    .card-status.review    {{ background: var(--review); }}
    .card-status.ok        {{ background: var(--success); }}
    .card-status.ng        {{ background: var(--danger); }}
    /* サムネール */
    .card-images {{
      display: flex; gap: 4px; padding: 8px;
      overflow-x: auto; background: var(--gray-50); min-height: 80px;
    }}
    .card-images img {{
      width: 72px; height: 72px; object-fit: cover;
      border-radius: 6px; flex-shrink: 0;
    }}
    .no-image {{
      width: 72px; height: 72px; border-radius: 6px;
      background: var(--gray-200); display: flex;
      align-items: center; justify-content: center;
      font-size: 24px; flex-shrink: 0;
    }}
    /* カード本文 */
    .card-body {{ padding: 10px 12px; }}
    .card-title {{
      font-size: 13px; font-weight: 600; line-height: 1.4;
      margin-bottom: 8px;
      display: -webkit-box; -webkit-line-clamp: 2;
      -webkit-box-orient: vertical; overflow: hidden;
    }}
    .card-price-row {{
      display: flex; gap: 12px; margin-bottom: 6px; font-size: 13px;
    }}
    .card-price {{ font-size: 18px; font-weight: 700; color: var(--danger); }}
    .card-price-label {{ font-size: 10px; color: var(--gray-600); }}
    .card-price-sub {{ font-weight: 600; }}
    .card-sellers {{
      font-size: 11px; color: var(--gray-600); margin-bottom: 8px;
    }}
    .seller-badge {{
      display: inline-block; background: var(--gray-100);
      border-radius: 3px; padding: 1px 5px; margin-right: 2px;
      font-family: monospace;
    }}
    .group-badge {{
      display: inline-block; background: #dbeafe;
      color: var(--primary); border-radius: 4px;
      padding: 2px 8px; font-weight: 700; font-size: 12px;
    }}
    /* アクション */
    .card-actions {{
      display: flex; gap: 6px; flex-wrap: wrap;
      padding: 8px 12px; border-top: 1px solid var(--gray-100);
      background: var(--gray-50);
    }}
    .btn {{
      display: inline-flex; align-items: center;
      padding: 7px 14px; border: none; border-radius: 6px;
      font-size: 13px; font-weight: 600; cursor: pointer;
      text-decoration: none; color: inherit;
    }}
    .btn-gray  {{ background: var(--gray-200); color: var(--gray-700); }}
    /* カウント表示 */
    .count-info {{
      text-align: center; padding: 12px;
      color: var(--gray-600); font-size: 13px;
    }}
    /* empty */
    .empty {{
      text-align: center; padding: 60px 20px; color: var(--gray-600);
    }}
    .empty-icon {{ font-size: 48px; margin-bottom: 12px; }}
  </style>
</head>
<body>

<div class="header">
  <h1>🔍 AucFan リサーチ結果</h1>
  <div class="header-meta">
    キーワード: {keyword} ／ {total_items}件 / {total_groups}グループ ／ 書き出し: {exported_at}
  </div>
</div>

<div class="filter-bar" id="filterBar">
  <button class="active" onclick="setFilter('')">すべて</button>
  <button onclick="setFilter('candidate')">🔵 仕入れ候補</button>
  <button onclick="setFilter('ok')">✅ OK</button>
  <button onclick="setFilter('waiting')">⏳ 確認待ち</button>
  <button onclick="setFilter('review')">⚠️ 要確認</button>
  <button onclick="setFilter('ng')">❌ NG</button>
</div>

<div class="search-bar">
  <input type="search" id="searchInput" placeholder="タイトルで絞り込み..."
         oninput="renderCards()" autocomplete="off" autocorrect="off">
</div>

<div id="countInfo" class="count-info"></div>
<div class="grid" id="grid"></div>

<script>
const GROUPS = {groups_json};

let currentFilter = '';

function setFilter(status) {{
  currentFilter = status;
  document.querySelectorAll('#filterBar button').forEach(btn => btn.classList.remove('active'));
  event.target.classList.add('active');
  document.getElementById('searchInput').value = '';
  renderCards();
}}

function esc(str) {{
  if (!str) return '';
  return String(str)
    .replace(/&/g,'&amp;').replace(/</g,'&lt;')
    .replace(/>/g,'&gt;').replace(/"/g,'&quot;');
}}

const STATUS_LABEL = {{
  candidate: '仕入れ候補', waiting: '確認待ち',
  review: '要確認', ok: '✅ OK', ng: '❌ NG'
}};

function renderCards() {{
  const grid = document.getElementById('grid');
  const kw   = document.getElementById('searchInput').value.trim().toLowerCase();

  const filtered = GROUPS.filter(g => {{
    if (currentFilter && g.status !== currentFilter) return false;
    if (kw && !g.title.toLowerCase().includes(kw)) return false;
    return true;
  }});

  document.getElementById('countInfo').textContent =
    `表示中: ${{filtered.length}} グループ`;

  if (filtered.length === 0) {{
    grid.innerHTML = '<div class="empty"><div class="empty-icon">📦</div><p>該当する商品がありません</p></div>';
    return;
  }}

  grid.innerHTML = filtered.map(g => {{
    const statusLabel = STATUS_LABEL[g.status] || g.status;
    const countLabel  = g.count > 1
      ? `同一商品 <span class="group-badge">${{g.count}}件</span>` : '単品';

    const thumbsHtml = g.thumbs.length > 0
      ? g.thumbs.map(src => `<img src="${{src}}" loading="lazy" alt="">`).join('')
      : '<div class="no-image">📦</div>';

    const shipping = g.shipping === 0 ? '無料' : '¥' + g.shipping.toLocaleString();
    const sellers  = (g.seller_ids || []).map(s => `<span class="seller-badge">${{esc(s)}}</span>`).join('');

    const searchQ  = encodeURIComponent((g.title || '').substring(0, 50));
    const aliUrl   = `https://aliprice.com/search?q=${{searchQ}}`;
    const amaUrl   = `https://www.amazon.co.jp/s?k=${{searchQ}}`;

    return `<div class="card ${{g.status}}">
      <div class="card-status ${{g.status}}">
        <span>${{esc(statusLabel)}}</span>
        <span>${{countLabel}}</span>
      </div>
      <div class="card-images">${{thumbsHtml}}</div>
      <div class="card-body">
        <div class="card-title">${{esc(g.title || '（タイトルなし）')}}</div>
        <div class="card-price-row">
          <div>
            <div class="card-price-label">合計</div>
            <div class="card-price">¥${{(g.total||0).toLocaleString()}}</div>
          </div>
          <div>
            <div class="card-price-label">落札価格</div>
            <div class="card-price-sub">¥${{(g.price||0).toLocaleString()}}</div>
          </div>
          <div>
            <div class="card-price-label">送料</div>
            <div class="card-price-sub">${{shipping}}</div>
          </div>
        </div>
        <div class="card-sellers">出品者: ${{sellers || '—'}}</div>
      </div>
      <div class="card-actions">
        <a class="btn btn-gray" href="${{aliUrl}}" target="_blank" rel="noopener">🛒 AliPrice</a>
        <a class="btn btn-gray" href="${{amaUrl}}" target="_blank" rel="noopener">📦 Amazon</a>
        ${{g.url ? `<a class="btn btn-gray" href="${{esc(g.url)}}" target="_blank" rel="noopener">🔗 元ページ</a>` : ''}}
      </div>
    </div>`;
  }}).join('');
}}

renderCards();
</script>
</body>
</html>"""


@app.route("/api/export/csv")
def api_export_csv():
    """CSVファイルをダウンロード"""
    dm = get_dm()
    if not dm:
        return jsonify({"error": "データがありません"}), 400

    csv_path = dm.export_csv()
    if not csv_path.exists():
        return jsonify({"error": "CSVファイルが見つかりません"}), 404

    return send_from_directory(
        str(csv_path.parent),
        csv_path.name,
        as_attachment=True,
        download_name=f"aucfan_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
        mimetype="text/csv; charset=utf-8-sig"
    )


@app.route("/api/load_csv", methods=["POST"])
def api_load_csv():
    """
    results.csv をアップロードしてグリッドに表示する。
    CSV の各行をアイテムとして新規 DataManager セッションにロードし、
    _data_manager を差し替える（画像は別途 /images/ から参照）。
    """
    global _data_manager, _image_processor, _gemini_client, _session_output_dir

    file = request.files.get("file")
    if not file:
        return jsonify({"error": "ファイルがありません"}), 400

    try:
        raw = file.read()
        df = None
        for enc in ("utf-8-sig", "utf-8", "shift-jis"):
            try:
                df = pd.read_csv(io.BytesIO(raw), encoding=enc, dtype=str)
                break
            except Exception:
                continue

        if df is None:
            return jsonify({"error": "CSVの文字コードを判別できませんでした"}), 400
        if df.empty:
            return jsonify({"error": "CSVにデータがありません"}), 400

        # 新規セッションを作成
        keyword_val = "csv_import"
        if "keyword" in df.columns:
            first_kw = str(df["keyword"].iloc[0]).strip()
            if first_kw and first_kw.lower() not in ("nan", ""):
                keyword_val = first_kw

        out_dir, session_id = make_output_dir(keyword_val, step=1)
        dm = DataManager(session_id, out_dir)

        # CSV行をアイテムとして追加
        for _, row in df.iterrows():
            item = {k: v for k, v in row.items() if pd.notna(v)}
            # 数値列を適切な型に変換
            for int_col in ("price", "shipping", "total", "group_size"):
                if int_col in item:
                    try:
                        item[int_col] = int(float(item[int_col]))
                    except (ValueError, TypeError):
                        item[int_col] = 0
            for bool_col in ("needs_review", "scraped_detail"):
                if bool_col in item:
                    item[bool_col] = str(item[bool_col]).lower() in ("true", "1", "yes")
            if isinstance(item.get("images_local"), str):
                # "[]" や JSON文字列をリストに変換
                try:
                    import json as _json
                    item["images_local"] = _json.loads(item["images_local"])
                except Exception:
                    item["images_local"] = []
            dm.add_item(item)

        loaded = dm.total_items
        dm.update_progress(
            keyword=keyword_val,
            status="done",
            total_items=loaded,
        )
        dm.save_all()

        with _lock:
            _data_manager = dm
            _image_processor = ImageProcessor(Path(config.LOCAL_IMAGE_CACHE_DIR) / out_dir.name / "images")
            _gemini_client = GeminiClient()
            _session_output_dir = out_dir

        logger.info(f"CSV読み込み完了: {file.filename} → {loaded}件 セッション={session_id}")
        return jsonify({
            "success": True,
            "session_id": session_id,
            "total_items": loaded,
            "keyword": keyword_val,
        })

    except Exception as e:
        logger.error(f"CSV読み込みエラー: {e}")
        return jsonify({"error": str(e)}), 500


# ─────────────────────────────────────────────
# マスターセラーリスト CSV / HTML エクスポート・インポート
# ─────────────────────────────────────────────

@app.route("/api/master_sellers/export/csv")
def api_master_export_csv():
    """マスターセラーリストを CSV でダウンロード"""
    from urllib.parse import quote

    records = _sellers_master.get_all()
    if not records:
        return jsonify({"error": "マスターリストが空です"}), 400

    df = pd.DataFrame(records)
    buf = io.BytesIO()
    df.to_csv(buf, index=False, encoding="utf-8-sig")
    buf.seek(0)

    filename = f"sellers_master_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    response = Response(buf.getvalue(), mimetype="text/csv; charset=utf-8-sig")
    response.headers["Content-Disposition"] = (
        f"attachment; filename*=UTF-8''{quote(filename)}"
    )
    return response


@app.route("/api/master_sellers/export/html")
def api_master_export_html():
    """マスターセラーリストをスタンドアロン HTML でダウンロード"""
    from urllib.parse import quote

    records = _sellers_master.get_all()
    if not records:
        return jsonify({"error": "マスターリストが空です"}), 400

    stats = _sellers_master.stats()
    exported_at = datetime.now().strftime("%Y年%m月%d日 %H:%M")

    rows_html = ""
    for r in records:
        scraped = r.get("last_scraped_date") or '<span style="color:#dc2626">未</span>'
        cands = r.get("candidates_count")
        cands_str = str(cands) if cands is not None else "—"
        rows_html += f"""
        <tr>
          <td style="font-family:monospace">{r.get('seller_id', '')}</td>
          <td>{r.get('first_seen_date', '—')}</td>
          <td>{scraped}</td>
          <td style="text-align:center">{cands_str}</td>
          <td>{r.get('source_keyword', '—')}</td>
        </tr>"""

    html = f"""<!DOCTYPE html>
<html lang="ja">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>マスターセラーリスト</title>
  <style>
    *, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}
    body {{ font-family: -apple-system, BlinkMacSystemFont, 'Hiragino Sans', sans-serif;
           font-size: 14px; background: #f9fafb; color: #111827; padding: 16px; }}
    .header {{ background: #2563eb; color: #fff; padding: 12px 16px;
               border-radius: 8px; margin-bottom: 12px; }}
    .header h1 {{ font-size: 16px; font-weight: 700; }}
    .header-meta {{ font-size: 11px; opacity: .8; margin-top: 2px; }}
    .search-bar {{ margin-bottom: 10px; }}
    .search-bar input {{ width: 100%; padding: 8px 12px; border: 1px solid #d1d5db;
                         border-radius: 20px; font-size: 14px; outline: none; }}
    table {{ width: 100%; border-collapse: collapse; background: #fff;
             border-radius: 8px; overflow: hidden;
             box-shadow: 0 1px 3px rgba(0,0,0,.08); }}
    th {{ padding: 10px 12px; background: #f3f4f6; font-weight: 600;
          text-align: left; border-bottom: 2px solid #e5e7eb; font-size: 12px; }}
    td {{ padding: 9px 12px; border-bottom: 1px solid #f3f4f6; font-size: 13px; }}
    tr:last-child td {{ border-bottom: none; }}
    tr:hover td {{ background: #f0f9ff; }}
    .count-info {{ color: #6b7280; font-size: 12px; margin: 8px 0 4px; }}
  </style>
</head>
<body>
  <div class="header">
    <h1>📋 マスターセラーリスト</h1>
    <div class="header-meta">
      合計 {stats['total']}件 ／ 未スクレイピング {stats['unscraped']}件 ／ 書き出し: {exported_at}
    </div>
  </div>
  <div class="search-bar">
    <input type="search" id="searchInput" placeholder="セラーIDで絞り込み..." oninput="filterTable()" autocomplete="off">
  </div>
  <div class="count-info" id="countInfo">表示中: {stats['total']}件</div>
  <table id="masterTable">
    <thead>
      <tr>
        <th>セラーID</th>
        <th>初回検出日</th>
        <th>最終スクレイピング</th>
        <th>候補件数</th>
        <th>キーワード</th>
      </tr>
    </thead>
    <tbody id="tableBody">{rows_html}</tbody>
  </table>
  <script>
    const allRows = Array.from(document.querySelectorAll('#tableBody tr'));
    function filterTable() {{
      const q = document.getElementById('searchInput').value.trim().toLowerCase();
      let count = 0;
      allRows.forEach(tr => {{
        const text = tr.textContent.toLowerCase();
        const show = !q || text.includes(q);
        tr.style.display = show ? '' : 'none';
        if (show) count++;
      }});
      document.getElementById('countInfo').textContent = '表示中: ' + count + '件';
    }}
  </script>
</body>
</html>"""

    filename = f"sellers_master_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
    response = Response(html, mimetype="text/html; charset=utf-8")
    response.headers["Content-Disposition"] = (
        f"attachment; filename*=UTF-8''{quote(filename)}"
    )
    return response


@app.route("/api/master_sellers/import/csv", methods=["POST"])
def api_master_import_csv():
    """
    seller_id 列を含む CSV をアップロードしてマスターリストに追記する。
    既存セラーはスキップ（重複無視）。
    """
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "ファイルがありません"}), 400

    try:
        raw = file.read()
        df = None
        for enc in ("utf-8-sig", "utf-8", "shift-jis"):
            try:
                df = pd.read_csv(io.BytesIO(raw), encoding=enc, dtype=str)
                break
            except Exception:
                continue

        if df is None:
            return jsonify({"error": "CSVの文字コードを判別できませんでした"}), 400
        if "seller_id" not in df.columns:
            return jsonify({"error": "seller_id 列が見つかりません"}), 400

        seller_ids = []
        for _, row in df.iterrows():
            sid = str(row.get("seller_id", "")).strip()
            if sid and sid.lower() not in ("nan", ""):
                seller_ids.append(sid)

        if not seller_ids:
            return jsonify({"error": "有効な seller_id がありません"}), 400

        added = _sellers_master.upsert_sellers(seller_ids, source_keyword="csv_import")

        logger.info(f"マスターリストCSVインポート: {len(seller_ids)}件中 {added}件追加")
        return jsonify({
            "success": True,
            "added": added,
            "total_in_file": len(seller_ids),
            "stats": _sellers_master.stats(),
        })

    except Exception as e:
        logger.error(f"マスターCSVインポートエラー: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/master/merge", methods=["POST"])
def api_master_merge():
    """
    別Macで蓄積した sellers_master.json をアップロードしてマージする。
    seller_id で重複排除し、新しいIDだけを追加する。
    既存データ（first_seen_date等）は上書きしない。
    """
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "ファイルがありません"}), 400

    # 一時ファイルに書き出してから merge_from_file を呼ぶ
    import tempfile
    try:
        with tempfile.NamedTemporaryFile(
            suffix=".json", delete=False, mode="wb"
        ) as tmp:
            tmp_path = tmp.name
            file.save(tmp)

        result = _sellers_master.merge_from_file(tmp_path)

        logger.info(
            f"マスターリストマージ: {result['added']}件追加、"
            f"{result['skipped']}件スキップ、合計{result['total']}件"
        )
        return jsonify({
            "success": True,
            "added": result["added"],
            "skipped": result["skipped"],
            "total": result["total"],
            "message": (
                f"{result['added']}件追加、"
                f"{result['skipped']}件スキップ（重複）、"
                f"合計{result['total']}件"
            ),
        })

    except (ValueError, FileNotFoundError) as e:
        logger.warning(f"マスターリストマージエラー: {e}")
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        logger.error(f"マスターリストマージ予期せぬエラー: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        try:
            import os
            os.unlink(tmp_path)
        except Exception:
            pass


# ─────────────────────────────────────────────
# Chrome タブ一覧取得
# ─────────────────────────────────────────────

@app.route("/api/tabs")
def api_tabs():
    """MacのChromeで開いているAucFanタブ一覧を返す"""
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options as ChromeOptions
    from selenium.common.exceptions import WebDriverException

    try:
        options = ChromeOptions()
        options.add_experimental_option(
            "debuggerAddress",
            f"{config.CHROME_DEBUG_HOST}:{config.CHROME_DEBUG_PORT}"
        )
        options.add_argument("--no-sandbox")
        driver = webdriver.Chrome(options=options)

        tabs = []
        current_handle = driver.current_window_handle
        for handle in driver.window_handles:
            driver.switch_to.window(handle)
            url   = driver.current_url
            title = driver.title
            tabs.append({
                "handle": handle,
                "url":    url,
                "title":  title,
                "is_aucfan": "aucfan.com" in url,
            })

        # 元のタブに戻す
        try:
            driver.switch_to.window(current_handle)
        except Exception:
            pass

        driver.quit()

        aucfan_tabs = [t for t in tabs if t["is_aucfan"]]
        return jsonify({"tabs": aucfan_tabs, "all_tabs": tabs})

    except WebDriverException:
        return jsonify({"error": "Chromeに接続できません。bash start.sh でアプリを起動してください。", "tabs": []}), 200
    except Exception as e:
        return jsonify({"error": f"エラー: {e}", "tabs": []}), 200


# ─────────────────────────────────────────────
# セッション管理
# ─────────────────────────────────────────────

@app.route("/api/sessions")
def api_sessions():
    """過去のセッション一覧。?step=1|2|3 でフィルタ可能。"""
    step_param = request.args.get("step")
    step_int = int(step_param) if step_param and step_param.isdigit() else None
    return jsonify({"sessions": _list_sessions(step=step_int)})


@app.route("/api/current_session")
def api_current_session():
    """現在グリッドに表示中のセッション情報を返す"""
    if _session_output_dir is None or _data_manager is None:
        return jsonify({"session": None})
    info = _parse_session_info(_session_output_dir.name)
    return jsonify({
        "session": {
            "name": _session_output_dir.name,
            "label": info["label"],
            "date_str": info["date_str"],
            "step": info["step"],
            "total_items": _data_manager.total_items,
            "status": _data_manager.get_progress().get("status", ""),
        }
    })


@app.route("/api/sessions/<session_name>", methods=["DELETE"])
def api_delete_session(session_name):
    """
    セッションフォルダを丸ごと削除する。
    現在メモリにロードされているセッションを削除する場合は _data_manager もリセット。
    """
    global _data_manager, _image_processor, _gemini_client, _session_output_dir

    # パストラバーサル防止
    if ".." in session_name or "/" in session_name or "\\" in session_name:
        return jsonify({"success": False, "message": "不正なセッション名"}), 400

    base = Path(config.OUTPUT_BASE_DIR)
    session_dir = base / session_name

    if not session_dir.exists():
        return jsonify({"success": False, "message": "セッションが見つかりません"}), 404

    try:
        shutil.rmtree(session_dir)
        logger.info(f"セッション削除: {session_dir}")

        # ローカル画像キャッシュも削除（img_cache/セッション名/）
        local_img_dir = Path(config.LOCAL_IMAGE_CACHE_DIR) / session_name
        if local_img_dir.exists():
            shutil.rmtree(local_img_dir)
            logger.info(f"ローカル画像キャッシュ削除: {local_img_dir}")

        # 削除したセッションが現在ロード中なら状態をリセット
        if _session_output_dir is not None and _session_output_dir.resolve() == session_dir.resolve():
            _data_manager = None
            _image_processor = None
            _gemini_client = None
            _session_output_dir = None
            logger.info("現在のセッションを削除したためメモリをリセットしました")

        return jsonify({"success": True, "name": session_name})
    except Exception as e:
        logger.error(f"セッション削除エラー: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


@app.route("/api/sessions/<session_name>/seller_ids", methods=["POST"])
def api_seller_ids_from_session(session_name):
    """
    指定した過去のSTEP 1セッションの results.csv から seller_id を抽出して
    _seller_state にセットする。
    """
    # パストラバーサル防止
    if ".." in session_name or "/" in session_name or "\\" in session_name:
        return jsonify({"error": "不正なセッション名"}), 400

    base = Path(config.OUTPUT_BASE_DIR)
    session_dir = base / session_name
    csv_path = session_dir / "results.csv"

    if not csv_path.exists():
        # CSV がなければ items.json から試みる
        items_path = session_dir / "items.json"
        if items_path.exists():
            try:
                with open(items_path, encoding="utf-8") as f:
                    items_dict = json.load(f)
                seen = {}
                for item in items_dict.values():
                    sid = str(item.get("seller_id", "")).strip()
                    if not sid or sid.lower() in ("nan", ""):
                        continue
                    if sid not in seen:
                        url = str(item.get("seller_url", "")).strip()
                        seen[sid] = "" if url.lower() in ("nan", "") else url
                sellers = [{"seller_id": s, "seller_url": u, "status": "pending"} for s, u in seen.items()]
                src_kw = _parse_session_info(session_name).get("label", session_name)
                with _seller_lock:
                    _seller_state["sellers"] = sellers
                    _seller_state["current_index"] = -1
                    _seller_state["running"] = False
                    _seller_state["stop_requested"] = False
                    _seller_state["session_dirs"] = []
                    _seller_state["source_keyword"] = src_kw
                return jsonify({
                    "count": len(sellers),
                    "has_seller_url": any(s["seller_url"] for s in sellers),
                    "sellers": sellers,
                    "session_name": session_name,
                })
            except Exception as e:
                return jsonify({"error": f"items.json 読み込み失敗: {e}"}), 500
        return jsonify({"error": f"results.csv が見つかりません: {session_name}"}), 404

    try:
        df = pd.read_csv(csv_path, encoding="utf-8-sig", dtype=str)
        if "seller_id" not in df.columns:
            return jsonify({"error": "seller_id 列が見つかりません"}), 400

        has_seller_url = "seller_url" in df.columns
        seen = {}
        for _, row in df.iterrows():
            sid = str(row.get("seller_id", "")).strip()
            if not sid or sid.lower() in ("nan", ""):
                continue
            if sid not in seen:
                seller_url = ""
                if has_seller_url:
                    u = str(row.get("seller_url", "")).strip()
                    if u and u.lower() != "nan":
                        seller_url = u
                seen[sid] = seller_url

        sellers = [{"seller_id": s, "seller_url": u, "status": "pending"} for s, u in seen.items()]

        src_kw = _parse_session_info(session_name).get("label", session_name)
        with _seller_lock:
            _seller_state["sellers"] = sellers
            _seller_state["current_index"] = -1
            _seller_state["running"] = False
            _seller_state["stop_requested"] = False
            _seller_state["session_dirs"] = []
            _seller_state["source_keyword"] = src_kw  # STEP2セッションに引き継ぐキーワード名

        logger.info(f"セッション({session_name})からセラーID抽出: {len(sellers)}件")
        return jsonify({
            "count": len(sellers),
            "has_seller_url": has_seller_url,
            "sellers": sellers,
            "session_name": session_name,
        })
    except Exception as e:
        logger.error(f"セラーID抽出エラー: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/sessions/<session_name>/load", methods=["POST"])
def api_load_session(session_name):
    """過去セッションをロード"""
    global _data_manager, _image_processor, _gemini_client, _session_output_dir

    base = Path(config.OUTPUT_BASE_DIR)
    session_dir = base / session_name

    if not session_dir.exists():
        return jsonify({"success": False, "message": "セッションが見つかりません"}), 404

    _session_output_dir = session_dir
    _data_manager = DataManager(session_name, session_dir)
    _data_manager.load_previous_session()
    _gemini_client = GeminiClient()

    return jsonify({
        "success": True,
        "session": session_name,
        "total_items": _data_manager.total_items,
    })




def _parse_session_info(name: str) -> dict:
    """フォルダ名から表示用情報を解析する（新旧両命名規則対応）。"""
    import re

    # 新命名規則: S1_20260506_01_バフ / S2_20260506_01 / S3_20260506_01
    m = re.match(r'^S(\d)_(\d{4})(\d{2})(\d{2})_(\d+)(?:_(.+))?$', name)
    if m:
        s, y, mo, d, num, kw = m.groups()
        step = int(s)
        st_map = {1: "keyword", 2: "seller", 3: "master"}
        lbl_map = {1: "STEP 1", 2: "セラー分析", 3: "マスター分析"}
        label = kw if (step == 1 and kw) else lbl_map.get(step, f"STEP {s}")
        return {
            "step": step,
            "session_type": st_map.get(step, "keyword"),
            "label": label,
            "date_str": f"{y}/{mo}/{d}",
            "num": int(num),
        }

    # 旧命名規則: keyword_YYYYMMDD_HHMMSS
    m = re.match(r'^(.+?)_(\d{4})(\d{2})(\d{2})_(\d{2})(\d{2})\d{2}$', name)
    if m:
        kw, y, mo, d, h, mi = m.groups()
        if kw.startswith("seller_analysis"):
            step, st, label = 2, "seller", "セラー分析"
        elif kw.startswith("master_analysis"):
            step, st, label = 3, "master", "マスター分析"
        else:
            step, st, label = 1, "keyword", kw
        return {
            "step": step,
            "session_type": st,
            "label": label,
            "date_str": f"{y}/{mo}/{d} {h}:{mi}",
            "num": 0,
        }

    return {"step": 1, "session_type": "keyword", "label": name, "date_str": "", "num": 0}


def _list_sessions(step: int = None):
    """過去セッションの一覧を返す。step を指定するとそのステップのみ返す。"""
    base = Path(config.OUTPUT_BASE_DIR)
    # Google Drive 未接続の場合はローカルフォールバック
    if not base.exists():
        _gdrive_prefix = os.path.expanduser("~/Library/CloudStorage/")
        if str(base).startswith(_gdrive_prefix):
            base = Path("リサーチ結果")
    if not base.exists():
        return []

    # 現在スクレイピング中のセッション名（is_running フラグ用）
    running_names: set = set()
    if _scraper_thread and _scraper_thread.is_alive() and _session_output_dir:
        running_names.add(_session_output_dir.name)
    with _seller_lock:
        if _seller_state["running"] and _seller_state.get("output_dir"):
            running_names.add(Path(_seller_state["output_dir"]).name)
    with _master_lock:
        if _master_state["running"] and _master_state.get("output_dir"):
            running_names.add(Path(_master_state["output_dir"]).name)

    import re as _re_sess
    # 有効なセッション名パターン（新形式: S1_20260506_01_キーワード / 旧形式: keyword_20260506_123456）
    _SESS_PAT = _re_sess.compile(r'^S[123]_\d{8}_\d+|^.+_\d{8}_\d{6}$')

    sessions = []
    try:
        entries = sorted(base.iterdir(), reverse=True)
    except Exception as e:
        logger.warning(f"セッション一覧の読み込みに失敗しました（Google Drive同期中の可能性）: {e}")
        return []

    for d in entries:
        try:
            if not d.is_dir():
                continue
        except Exception:
            continue

        info = _parse_session_info(d.name)
        if step is not None and info["step"] != step:
            continue

        progress_file = d / "progress.json"
        is_valid_session_name = bool(_SESS_PAT.match(d.name))

        if not is_valid_session_name:
            # 命名規則外のフォルダはデータファイルがある場合のみ表示
            try:
                has_data = (progress_file.exists() or
                            (d / "results.csv").exists() or
                            (d / "items.json").exists())
            except Exception:
                has_data = False
            if not has_data:
                continue

        p = {}
        try:
            if progress_file.exists():
                with open(progress_file, encoding="utf-8") as f:
                    p = json.load(f)
        except Exception:
            pass  # Google Drive 同期中で読み込めない場合はスキップせず空データで続行

        # 完全に空（ファイルなし）なフォルダは削除残骸とみなしてスキップ
        if is_valid_session_name and not p:
            try:
                has_any_file = any(True for _ in d.iterdir())
            except Exception:
                has_any_file = True  # 確認できない場合は表示する
            if not has_any_file:
                continue

        sessions.append({
            "name": d.name,
            "step": info["step"],
            "session_type": info["session_type"],   # 後方互換
            "label": info["label"],
            "date_str": info["date_str"],
            "num": info["num"],
            "keyword": p.get("keyword", info["label"]),
            "source_keyword": p.get("source_keyword", ""),   # 元STEP1キーワード
            "machine_name": p.get("machine_name", ""),       # 実行したMac名
            "status": p.get("status", "unknown"),
            "total_items": p.get("total_items", 0),
            "started_at": p.get("started_at", ""),
            "updated_at": p.get("updated_at", ""),
            "has_csv": (d / "results.csv").exists(),
            "is_running": d.name in running_names,
        })

    return sessions[:100]


# ─────────────────────────────────────────────
# レポート API
# ─────────────────────────────────────────────

@app.route("/api/report")
def api_report():
    """セラー分析・画像グループ分析レポートを返す"""
    dm = get_dm()
    if not dm:
        return jsonify({"error": "データがありません"}), 400

    items = dm.get_all_items()
    if not items:
        return jsonify({"error": "商品データがありません"}), 400

    from collections import defaultdict

    # ── セラーごとの集計 ──
    seller_items = defaultdict(list)   # seller_id -> [item, ...]
    seller_groups = defaultdict(set)   # seller_id -> {group_id, ...}

    for item in items:
        sid = item.get("seller_id", "").strip()
        if not sid:
            continue
        gid = item.get("group_id") or item.get("item_id", "")
        seller_items[sid].append(item)
        seller_groups[sid].add(gid)

    seller_ranking = []
    for sid, sitems in seller_items.items():
        prices = [i.get("price", 0) for i in sitems if i.get("price", 0) > 0]
        seller_ranking.append({
            "seller_id": sid,
            "item_count": len(sitems),
            "group_count": len(seller_groups[sid]),
            "min_price": min(prices) if prices else 0,
            "max_price": max(prices) if prices else 0,
        })
    seller_ranking.sort(key=lambda x: x["item_count"], reverse=True)

    # ── グループごとの集計 ──
    group_map = defaultdict(list)
    for item in items:
        gid = item.get("group_id") or item.get("item_id", "")
        group_map[gid].append(item)

    group_report = []
    for gid, gitems in group_map.items():
        prices = [i.get("price", 0) for i in gitems if i.get("price", 0) > 0]
        sellers = [i.get("seller_id", "").strip() for i in gitems if i.get("seller_id", "").strip()]
        seller_counter = defaultdict(int)
        for s in sellers:
            seller_counter[s] += 1
        unique_sellers = list(seller_counter.keys())
        dup_sellers = [s for s, c in seller_counter.items() if c > 1]

        title = (gitems[0].get("title_full") or gitems[0].get("title_short", ""))[:60]
        thumb = gitems[0].get("thumbnail_local", "")
        thumb_name = Path(thumb).name if thumb else ""

        # グループが大きすぎる場合はpHashの誤グループ化の可能性あり
        too_large = len(gitems) > 50

        group_report.append({
            "group_id": gid,
            "title": title,
            "item_count": len(gitems),
            "seller_count": len(unique_sellers),
            "sellers": unique_sellers[:10],
            "dup_sellers": dup_sellers,
            "min_price": min(prices) if prices else 0,
            "max_price": max(prices) if prices else 0,
            "status": gitems[0].get("status", ""),
            "thumbnail": thumb_name,
            "too_large": too_large,
        })
    group_report.sort(key=lambda x: x["item_count"], reverse=True)

    # ── 同一セラー×同一グループ（自演出品候補） ──
    suspicious = []
    for gr in group_report:
        if gr["dup_sellers"]:
            suspicious.append({
                "group_id": gr["group_id"],
                "title": gr["title"],
                "dup_sellers": gr["dup_sellers"],
                "item_count": gr["item_count"],
            })

    return jsonify({
        "seller_ranking": seller_ranking[:50],
        "group_report": group_report[:100],
        "suspicious": suspicious,
        "total_items": len(items),
        "total_sellers": len(seller_ranking),
        "total_groups": len(group_report),
    })


# ─────────────────────────────────────────────
# セラー分析機能
# ─────────────────────────────────────────────

@app.route("/api/seller_ids_from_current_session", methods=["POST"])
def api_seller_ids_from_current_session():
    """
    現在メモリに読み込まれているキーワードリサーチ結果から
    seller_id のユニークリストを抽出して _seller_state にセットする。
    CSV ファイルを介さずに STEP 1 → STEP 2 を直結する。
    """
    dm = get_dm()
    if dm is None or dm.total_items == 0:
        return jsonify({"error": "現在のセッションにデータがありません。先にSTEP 1のスクレイピングを実行してください"}), 404

    all_items = dm.get_all_items()
    seen = {}
    for item in all_items:
        sid = str(item.get("seller_id", "")).strip()
        if not sid or sid.lower() in ("nan", ""):
            continue
        if sid not in seen:
            seller_url = str(item.get("seller_url", "")).strip()
            if seller_url.lower() in ("nan", ""):
                seller_url = ""
            seen[sid] = seller_url

    if not seen:
        return jsonify({"error": "seller_id が見つかりませんでした（スクレイピング結果を確認してください）"}), 404

    sellers = [
        {"seller_id": sid, "seller_url": url, "status": "pending"}
        for sid, url in seen.items()
    ]

    keyword = dm.get_progress().get("keyword", "（不明）")
    with _seller_lock:
        _seller_state["sellers"] = sellers
        _seller_state["current_index"] = -1
        _seller_state["running"] = False
        _seller_state["stop_requested"] = False
        _seller_state["session_dirs"] = []
        _seller_state["source_keyword"] = keyword  # STEP2セッションに引き継ぐキーワード名

    has_seller_url = any(s["seller_url"] for s in sellers)
    logger.info(f"現在のセッション({keyword})からセラーID抽出: {len(sellers)}件")

    return jsonify({
        "count": len(sellers),
        "has_seller_url": has_seller_url,
        "sellers": sellers,
        "keyword": keyword,
    })


@app.route("/api/latest_csv_import", methods=["POST"])
def api_latest_csv_import():
    """
    Mac上の最新 results.csv を自動で読み込む（iPhone からのアップロード不要）。
    現在アクティブなセッション → なければ最新セッションの順で探す。
    """
    csv_path = None
    session_name = ""

    # 1. 現在アクティブなセッションの CSV を優先
    if _session_output_dir is not None:
        candidate = _session_output_dir / "results.csv"
        if candidate.exists():
            csv_path = candidate
            session_name = _session_output_dir.name

    # 2. なければ最新セッションを探す
    if csv_path is None:
        base = Path(config.OUTPUT_BASE_DIR)
        if base.exists():
            for d in sorted(base.iterdir(), reverse=True):
                if d.is_dir():
                    candidate = d / "results.csv"
                    if candidate.exists():
                        csv_path = candidate
                        session_name = d.name
                        break

    if csv_path is None:
        return jsonify({"error": "results.csv が見つかりません。先にキーワードスクレイピングを実行してください"}), 404

    try:
        df = pd.read_csv(csv_path, encoding="utf-8-sig", dtype=str)

        if "seller_id" not in df.columns:
            return jsonify({"error": f"seller_id 列が見つかりません: {csv_path}"}), 400

        has_seller_url = "seller_url" in df.columns

        seen = {}
        for _, row in df.iterrows():
            sid = str(row.get("seller_id", "")).strip()
            if not sid or sid.lower() == "nan":
                continue
            if sid not in seen:
                seller_url = ""
                if has_seller_url:
                    u = str(row.get("seller_url", "")).strip()
                    if u and u.lower() != "nan":
                        seller_url = u
                seen[sid] = seller_url

        sellers = [
            {"seller_id": sid, "seller_url": url, "status": "pending"}
            for sid, url in seen.items()
        ]

        with _seller_lock:
            _seller_state["sellers"] = sellers
            _seller_state["current_index"] = -1
            _seller_state["running"] = False
            _seller_state["stop_requested"] = False
            _seller_state["session_dirs"] = []

        logger.info(f"最新CSV自動読み込み: {csv_path} → {len(sellers)}件のユニークセラー")
        return jsonify({
            "count": len(sellers),
            "has_seller_url": has_seller_url,
            "sellers": sellers,
            "session_name": session_name,
            "csv_path": str(csv_path),
        })

    except Exception as e:
        logger.error(f"最新CSV読み込みエラー: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/import_csv", methods=["POST"])
def api_import_csv():
    """
    CSVをアップロードしてseller_idを抽出する。
    CSVフォーマット: results.csv と同一（seller_id列必須、seller_url列は任意）
    """
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "ファイルがありません"}), 400

    try:
        raw = file.read()
        # UTF-8 BOM → UTF-8 → Shift-JIS の順で試す
        for enc in ("utf-8-sig", "utf-8", "shift-jis"):
            try:
                df = pd.read_csv(io.BytesIO(raw), encoding=enc, dtype=str)
                break
            except Exception:
                continue
        else:
            return jsonify({"error": "CSVの文字コードを判別できませんでした"}), 400

        if "seller_id" not in df.columns:
            return jsonify({"error": "seller_id 列が見つかりません"}), 400

        has_seller_url = "seller_url" in df.columns

        # ユニークセラーを抽出（最初に出現した seller_url を使う）
        seen = {}
        for _, row in df.iterrows():
            sid = str(row.get("seller_id", "")).strip()
            if not sid or sid.lower() == "nan":
                continue
            if sid not in seen:
                seller_url = ""
                if has_seller_url:
                    u = str(row.get("seller_url", "")).strip()
                    if u and u.lower() != "nan":
                        seller_url = u
                seen[sid] = seller_url

        sellers = [
            {"seller_id": sid, "seller_url": url, "status": "pending"}
            for sid, url in seen.items()
        ]

        with _seller_lock:
            _seller_state["sellers"] = sellers
            _seller_state["current_index"] = -1
            _seller_state["running"] = False
            _seller_state["stop_requested"] = False
            _seller_state["session_dirs"] = []

        logger.info(f"CSVインポート完了: {len(sellers)}件のユニークセラー（seller_url={'あり' if has_seller_url else 'なし'}）")
        return jsonify({
            "count": len(sellers),
            "has_seller_url": has_seller_url,
            "sellers": sellers,
        })

    except Exception as e:
        logger.error(f"CSVインポートエラー: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/seller_scrape/start", methods=["POST"])
def api_seller_scrape_start():
    """セラー分析スクレイピングを開始"""
    with _seller_lock:
        if _seller_state["running"]:
            return jsonify({"error": "既に実行中です"}), 400
        sellers = _seller_state["sellers"]
        if not sellers:
            return jsonify({"error": "セラーリストがありません。先にCSVをインポートしてください"}), 400

        stop_ev = threading.Event()
        _seller_state["running"] = True
        _seller_state["phase"] = "scraping_list"
        _seller_state["stop_event"] = stop_ev
        _seller_state["session_id"] = None
        _seller_state["output_dir"] = None
        _seller_state["dm"] = None

    t = threading.Thread(
        target=_run_seller_analysis,
        args=(stop_ev,),
        daemon=True,
        name="seller-analyzer",
    )
    t.start()
    with _seller_lock:
        _seller_state["thread"] = t

    logger.info(f"セラー分析スクレイピング開始: {len(sellers)}件")
    return jsonify({"status": "started", "total": len(sellers)})


def _run_seller_analysis(stop_ev: threading.Event):
    """
    SellerAnalyzer を使って全セラーを 1 セッションにまとめてスクレイプする
    バックグラウンドスレッド。完了後に _data_manager を差し替えてメイン UI で表示可能にする。
    """
    global _data_manager, _image_processor, _gemini_client, _session_output_dir

    with _seller_lock:
        sellers = list(_seller_state["sellers"])
        source_keyword = _seller_state.get("source_keyword", "")

    # 1 セッションフォルダを作成 (S2_YYYYMMDD_NN)
    out_dir, session_id = make_output_dir("seller_analysis", step=2)
    dm = DataManager(session_id, out_dir)
    img = ImageProcessor(Path(config.LOCAL_IMAGE_CACHE_DIR) / out_dir.name / "images")
    gc = GeminiClient()

    # 元キーワードをセッション情報として記録（セッション履歴で表示するため）
    if source_keyword:
        dm.update_progress(source_keyword=source_keyword)

    with _seller_lock:
        _seller_state["session_id"] = session_id
        _seller_state["output_dir"] = out_dir
        _seller_state["dm"] = dm

    # セラーごとの進捗をコールバックで _seller_state に反映
    def on_progress(index: int, status: str, extra: dict = None):
        with _seller_lock:
            if 0 <= index < len(_seller_state["sellers"]):
                _seller_state["sellers"][index]["status"] = status
                # used_skip 時などに追加情報（used_count 等）を保存
                if extra:
                    _seller_state["sellers"][index].update(extra)
            _seller_state["current_index"] = index if status == "running" else -1
            # phase を DataManager の進捗から同期
            dm_status = dm.get_progress().get("status", "scraping_list")
            _seller_state["phase"] = dm_status

    analyzer = SellerAnalyzer(
        sellers=sellers,
        data_manager=dm,
        image_processor=img,
        gemini_client=gc,
        stop_event=stop_ev,
        on_seller_progress=on_progress,
        login_check_event=_login_check_event,
    )
    analyzer.run()

    final_status = dm.get_progress().get("status", "done")

    # CSV・HTML を自動保存（完了時のみ）
    if final_status in ("done", "stopped"):
        _save_export_files(dm, out_dir)

    # 完了後: メイン UI でそのまま結果を表示できるよう差し替え
    with _lock:
        _data_manager = dm
        _image_processor = img
        _gemini_client = gc
        _session_output_dir = out_dir

    with _seller_lock:
        _seller_state["running"] = False
        _seller_state["current_index"] = -1
        _seller_state["phase"] = final_status

    logger.info("=" * 50)
    logger.info(f"=== STEP 2 スクレイピング完了 === 全{dm.total_items}件処理 ({final_status})")
    logger.info("=" * 50)
    logger.info(">>> 待機中 (アプリは起動中) <<<  次の操作をブラウザから行ってください")


@app.route("/api/seller_scrape/stop", methods=["POST"])
def api_seller_scrape_stop():
    """セラー分析スクレイピングを停止"""
    with _seller_lock:
        ev = _seller_state.get("stop_event")
    if ev:
        ev.set()
    logger.info("セラー分析停止リクエスト")
    return jsonify({"status": "stopping"})


@app.route("/api/login_check", methods=["POST"])
def api_login_check():
    """
    UIの「今すぐ確認して再開」ボタンから呼び出す。
    _login_check_event を set() して、スクレイパーの待機ループを即座に起こす。
    STEP1 / STEP2 / STEP3 すべてで共通の Event を使用する。
    """
    _login_check_event.set()
    logger.info("ログイン即時チェックトリガー受信")
    return jsonify({"status": "checking"})


@app.route("/api/seller_scrape/status")
def api_seller_scrape_status():
    """セラー分析スクレイピングの現在状態を返す"""
    with _seller_lock:
        sellers = list(_seller_state["sellers"])
        current_index = _seller_state["current_index"]
        running = _seller_state["running"]
        phase = _seller_state["phase"]
        session_id = _seller_state["session_id"]
        dm = _seller_state["dm"]

    total = len(sellers)
    done_sellers = sum(1 for s in sellers if s.get("status") == "done")
    errors = sum(1 for s in sellers if s.get("status") == "error")
    used_skipped = sum(1 for s in sellers if s.get("status") == "used_skip")

    # DataManager から詳細進捗を取得
    dm_progress = {}
    total_items = 0
    if dm is not None:
        try:
            dm_progress = dm.get_progress()
            total_items = dm.total_items
        except Exception:
            pass

    # DataManager のステータスを優先する
    # （_seller_state["phase"] は on_progress コールバック呼び出し時にしか更新されないため、
    #   セラーループ終了後の grouping / vision_check / login_required フェーズは
    #   dm.get_progress() から直接読む必要がある）
    dm_status = dm_progress.get("status", "")
    _DM_OVERRIDE_PHASES = {
        "login_required", "grouping", "vision_check",
        "scraping_detail", "done", "stopped", "error",
    }
    effective_phase = dm_status if dm_status in _DM_OVERRIDE_PHASES else phase

    return jsonify({
        "running": running,
        "phase": effective_phase,
        "current_index": current_index,
        "total": total,
        "done": done_sellers,
        "errors": errors,
        "used_skipped": used_skipped,
        "sellers": sellers,
        "session_id": session_id,
        "total_items": total_items,
        "detail_pages_done": dm_progress.get("detail_pages_done", 0),
        "detail_pages_total": dm_progress.get("detail_pages_total", 0),
    })


@app.route("/api/seller_scrape/reset", methods=["POST"])
def api_seller_scrape_reset():
    """セラーリストをリセット"""
    with _seller_lock:
        if _seller_state["running"]:
            return jsonify({"error": "実行中はリセットできません"}), 400
        _seller_state["sellers"] = []
        _seller_state["current_index"] = -1
        _seller_state["phase"] = "idle"
        _seller_state["stop_event"] = None
        _seller_state["dm"] = None
        _seller_state["session_id"] = None
        _seller_state["output_dir"] = None
    return jsonify({"status": "reset"})


# ─────────────────────────────────────────────
# STEP 3: マスターセラーリサーチ
# ─────────────────────────────────────────────

@app.route("/api/master_sellers/add", methods=["POST"])
def api_master_sellers_add():
    """
    セラーIDを手動でマスターリストに追加する。
    リクエスト: {"seller_ids": ["id1", "id2", ...], "source_keyword": "任意メモ"}
    """
    data = request.get_json(force=True, silent=True) or {}
    raw = data.get("seller_ids", [])
    source_keyword = data.get("source_keyword", "手動追加")
    # 文字列1件でも受け付けられるよう正規化
    if isinstance(raw, str):
        raw = [raw]
    seller_ids = [s.strip() for s in raw if str(s).strip()]
    if not seller_ids:
        return jsonify({"error": "seller_id を1件以上指定してください"}), 400
    try:
        added = _sellers_master.upsert_sellers(seller_ids, source_keyword=source_keyword)
        logger.info(f"マスターリスト手動追加: {added}件追加 (入力{len(seller_ids)}件)")
        return jsonify({
            "success": True,
            "added": added,
            "total": len(seller_ids),
            "stats": _sellers_master.stats(),
        })
    except Exception as e:
        logger.error(f"マスターリスト手動追加エラー: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/master_sellers")
def api_master_sellers():
    """マスターセラーリスト取得（ソート・件数制限付き）"""
    sort_order = request.args.get("sort_order", "desc")   # asc / desc
    limit = request.args.get("limit", "0")
    try:
        limit = int(limit)
    except ValueError:
        limit = 0
    records = _sellers_master.get_all(sort_order=sort_order)
    if limit > 0:
        records = records[:limit]
    return jsonify({"sellers": records, "total": len(records)})


@app.route("/api/master_sellers/stats")
def api_master_sellers_stats():
    """マスターリスト統計"""
    return jsonify(_sellers_master.stats())


@app.route("/api/master_sellers/scrape/start", methods=["POST"])
def api_master_scrape_start():
    """STEP 3 スクレイピング開始"""
    with _master_lock:
        if _master_state["running"]:
            return jsonify({"error": "既に実行中です"}), 400

    data = request.get_json(silent=True) or {}
    sort_order = data.get("sort_order", "desc")
    batch_size = data.get("batch_size", 0)   # 0 = 上限なし

    all_unscraped = _sellers_master.get_unscraped(sort_order=sort_order)
    if not all_unscraped:
        return jsonify({"error": "未スクレイピングのセラーがありません"}), 400

    targets = all_unscraped[:batch_size] if batch_size > 0 else all_unscraped

    stop_ev = threading.Event()
    with _master_lock:
        _master_state["running"] = True
        _master_state["phase"] = "scraping_list"
        _master_state["stop_event"] = stop_ev
        _master_state["session_id"] = None
        _master_state["output_dir"] = None
        _master_state["dm"] = None
        _master_state["total"] = len(targets)
        _master_state["done"] = 0
        _master_state["current_seller"] = ""

    t = threading.Thread(
        target=_run_master_analysis,
        args=(targets, stop_ev),
        daemon=True,
        name="master-analyzer",
    )
    t.start()
    with _master_lock:
        _master_state["thread"] = t

    logger.info(f"STEP 3 スクレイピング開始: {len(targets)}件")
    return jsonify({"status": "started", "total": len(targets)})


@app.route("/api/master_sellers/scrape/stop", methods=["POST"])
def api_master_scrape_stop():
    """STEP 3 スクレイピング停止"""
    with _master_lock:
        ev = _master_state.get("stop_event")
    if ev:
        ev.set()
    logger.info("STEP 3 停止リクエスト")
    return jsonify({"status": "stopping"})


@app.route("/api/master_sellers/scrape/status")
def api_master_scrape_status():
    """STEP 3 進捗"""
    with _master_lock:
        running = _master_state["running"]
        phase = _master_state["phase"]
        total = _master_state["total"]
        done = _master_state["done"]
        current_seller = _master_state["current_seller"]
        session_id = _master_state["session_id"]
        dm = _master_state["dm"]

    total_items = 0
    processed_items = 0
    dm_status = ""
    if dm is not None:
        try:
            dm_progress = dm.get_progress()
            total_items = dm.total_items
            processed_items = dm_progress.get("processed_items", total_items)
            dm_status = dm_progress.get("status", "")
        except Exception:
            pass

    # DataManager のステータスを優先する（STEP2と同じロジック）
    _DM_OVERRIDE_PHASES = {
        "login_required", "grouping", "vision_check",
        "scraping_detail", "done", "stopped", "error",
    }
    effective_phase = dm_status if dm_status in _DM_OVERRIDE_PHASES else phase

    return jsonify({
        "running": running,
        "phase": effective_phase,
        "total": total,
        "done": done,
        "current_seller": current_seller,
        "session_id": session_id,
        "total_items": total_items,
        "processed_items": processed_items,
    })


@app.route("/api/master_sellers/all", methods=["DELETE"])
def api_master_delete_all():
    """マスターリスト全件削除"""
    deleted = _sellers_master.clear_all()
    logger.info(f"マスターリスト全削除: {deleted}件")
    return jsonify({"success": True, "deleted": deleted})


@app.route("/api/master_sellers/<seller_id>", methods=["DELETE"])
def api_master_delete_seller(seller_id):
    """個別セラー削除"""
    ok = _sellers_master.delete_seller(seller_id)
    if not ok:
        return jsonify({"success": False, "error": "seller_id が見つかりません"}), 404
    return jsonify({"success": True, "seller_id": seller_id})


def _run_master_analysis(targets: list, stop_ev: threading.Event):
    """STEP 3 バックグラウンドスレッド"""
    global _data_manager, _image_processor, _gemini_client, _session_output_dir

    out_dir, session_id = make_output_dir("master_analysis", step=3)
    dm = DataManager(session_id, out_dir)
    img = ImageProcessor(Path(config.LOCAL_IMAGE_CACHE_DIR) / out_dir.name / "images")
    gc = GeminiClient()

    with _master_lock:
        _master_state["session_id"] = session_id
        _master_state["output_dir"] = out_dir
        _master_state["dm"] = dm

    # SellerAnalyzer に渡す形式に変換
    # seller_url は sellers_master.json に保存済みのURLを優先して使用する。
    # 保存されていない場合（旧データ）は ?aucnm= フォールバックURLを使うが、
    # このURLは商品が表示されないケースがあるため、STEP1を再実行してURLを取得推奨。
    sellers_for_analyzer = [
        {
            "seller_id": r["seller_id"],
            "seller_url": (
                r.get("seller_url", "").strip()
                or f"https://aucfan.com/search1/s-ya/?seller={r['seller_id']}&shopid="
            ),
            "status": "pending",
        }
        for r in targets
    ]

    def on_progress(index: int, status: str):
        with _master_lock:
            sid = targets[index]["seller_id"] if 0 <= index < len(targets) else ""
            if status == "running":
                _master_state["current_seller"] = sid
            if status == "done":
                _master_state["done"] += 1
                # last_scraped_date と candidates_count を書き戻す
                try:
                    # 完了時点でのグループ化済み候補数を取得
                    cnt = sum(
                        1 for i in dm.get_all_items()
                        if i.get("seller_id") == sid
                        and i.get("status") in (
                            config.STATUS_CANDIDATE,
                            config.STATUS_NEXT_CANDIDATE,
                            config.STATUS_OK,
                        )
                    )
                    _sellers_master.update_scraped(sid, candidates_count=cnt)
                except Exception as _e:
                    logger.warning(f"sellers_master 更新失敗 ({sid}): {_e}")
            dm_status = dm.get_progress().get("status", "scraping_list")
            _master_state["phase"] = dm_status

    analyzer = SellerAnalyzer(
        sellers=sellers_for_analyzer,
        data_manager=dm,
        image_processor=img,
        gemini_client=gc,
        stop_event=stop_ev,
        on_seller_progress=on_progress,
        login_check_event=_login_check_event,
    )
    analyzer.run()

    final_status = dm.get_progress().get("status", "done")
    if final_status in ("done", "stopped"):
        _save_export_files(dm, out_dir)

    # 完了後にメイン UI に差し替え
    with _lock:
        _data_manager = dm
        _image_processor = img
        _gemini_client = gc
        _session_output_dir = out_dir

    with _master_lock:
        _master_state["running"] = False
        _master_state["current_seller"] = ""
        _master_state["phase"] = final_status

    logger.info("=" * 50)
    logger.info(f"=== STEP 3 スクレイピング完了 === 全{dm.total_items}件処理 ({final_status})")
    logger.info("=" * 50)


# ─────────────────────────────────────────────
# Gemini レート制限ステータス API
# ─────────────────────────────────────────────

@app.route("/api/gemini_status")
def api_gemini_status():
    """Gemini API エラーフラグの現在状態を返す
    Response: {"rate_limit_hit": bool, "type": str|null, "time": str|null}
    """
    status = get_rate_limit_status()
    return jsonify({
        "rate_limit_hit": status["rate_limit_hit"],
        "type": status["type"],
        "time": status["time"],
    })


@app.route("/api/gemini_status/reset", methods=["POST"])
def api_gemini_status_reset():
    """Gemini API エラーフラグをリセットする"""
    reset_rate_limit_flag()
    return jsonify({"success": True})


# ─────────────────────────────────────────────
# Amazon確認 API
# ─────────────────────────────────────────────

@app.route("/api/amazon/fetch", methods=["POST"])
def api_amazon_fetch():
    """
    現在Chromeで開いているAmazon商品ページからデータを取得する。
    group_id が指定された場合は DataManager に保存する。

    Request body（JSON、任意）:
      {"group_id": "グループID"}

    Response（成功時）:
      {
        "success": true,
        "asin": "B0XXXXXXXX",
        "title": "商品タイトル",
        "price": "¥1,980",
        "image_url": "https://...",
        "bullets": ["特徴1", "特徴2", ...],
        "description": "商品説明",
        "specs": {"重量": "300g", ...},
        "rating": "4.3",
        "review_count": "1,234件の評価",
        "has_aplus": true,
        "url": "https://www.amazon.co.jp/dp/B0XXXXXXXX"
      }
    Response（失敗時）:
      {"success": false, "error": "エラーメッセージ"}
    """
    from amazon_scraper import fetch_amazon_product

    # group_id が送られてきた場合は DataManager に保存する
    body = request.get_json(silent=True) or {}
    group_id = body.get("group_id", "").strip()

    result = fetch_amazon_product()

    if result.get("success") and group_id:
        dm = get_dm()
        if dm:
            dm.save_amazon_data(group_id, result)

    status_code = 200 if result.get("success") else 400
    return jsonify(result), status_code


@app.route("/api/amazon/data/<group_id>", methods=["GET"])
def api_amazon_get(group_id: str):
    """保存済みのAmazonデータをグループIDで取得する。"""
    dm = get_dm()
    if not dm:
        return jsonify({"success": False, "error": "セッションがありません"}), 404
    data = dm.get_amazon_data(group_id)
    if data is None:
        return jsonify({"success": False, "error": "Amazonデータ未取得"}), 404
    return jsonify({"success": True, **data})



# ─────────────────────────────────────────────
# /research — Excel 追記ツール
# ─────────────────────────────────────────────

# Amazon取得中ステータス（フロントエンドがポーリングして進捗表示に使う）
_research_fetch_status = {
    "running": False,
    "step":    "",      # 現在の処理ステップ説明
    "elapsed": 0,       # 開始からの経過秒（フロント側で計算）
}
# 2重実行防止ロック（同時リクエストが来ても1件しか処理しない）
_research_fetch_lock = threading.Lock()

# 1688取得中ステータス（フロントエンドがポーリングして進捗表示に使う）
_research_1688_fetch_status = {
    "running": False,
    "step":    "",
}
# 1688 2重実行防止ロック
_research_1688_fetch_lock = threading.Lock()

_RESEARCH_HTML = """<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>リサーチ追記ツール</title>
<style>
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body { font-family: "BIZ UDGothic", "Hiragino Sans", sans-serif;
         background: #f0f2f5; color: #333; min-height: 100vh; }

  .header { background: #1F4E79; color: white; padding: 14px 20px;
            display: flex; align-items: center; justify-content: space-between; }
  .header-title { font-size: 16px; font-weight: bold; }
  .header-sub   { font-size: 11px; opacity: 0.75; margin-top: 2px; }
  .back-link { color: rgba(255,255,255,0.8); font-size: 12px; text-decoration: none;
               border: 1px solid rgba(255,255,255,0.4); padding: 4px 10px; border-radius: 4px; }
  .back-link:hover { background: rgba(255,255,255,0.15); }

  .main { max-width: 820px; margin: 20px auto; padding: 0 14px; }

  .card { background: white; border-radius: 10px; padding: 18px;
          box-shadow: 0 2px 8px rgba(0,0,0,0.08); margin-bottom: 14px; }
  .card-title { font-size: 13px; font-weight: bold; color: #555;
                border-left: 4px solid #1F4E79; padding-left: 9px;
                margin-bottom: 14px; display: flex; align-items: center; gap: 8px; }

  /* ファイル一覧 */
  .folder-bar { display: flex; gap: 8px; align-items: center; margin-bottom: 10px; }
  .folder-input { flex: 1; padding: 8px 10px; border: 2px solid #ddd;
                  border-radius: 6px; font-size: 12px; font-family: monospace; }
  .folder-input:focus { outline: none; border-color: #1F4E79; }

  .file-list { border: 1px solid #e5e7eb; border-radius: 8px; overflow: hidden; }
  .file-item { display: flex; align-items: center; gap: 10px; padding: 10px 14px;
               border-bottom: 1px solid #f0f0f0; cursor: pointer; transition: background 0.15s; }
  .file-item:last-child { border-bottom: none; }
  .file-item:hover { background: #EBF3FB; }
  .file-item.selected { background: #D6E8F7; border-left: 3px solid #1F4E79; }
  .file-icon { font-size: 20px; flex-shrink: 0; }
  .file-info { flex: 1; min-width: 0; }
  .file-name { font-size: 13px; font-weight: bold; color: #1F4E79;
               white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }
  .file-meta { font-size: 11px; color: #888; margin-top: 2px; }
  .file-badge { font-size: 10px; background: #C55A11; color: white;
                padding: 1px 6px; border-radius: 8px; white-space: nowrap; }
  .file-empty { padding: 20px; text-align: center; color: #aaa; font-size: 13px; }

  /* 読み込み済みExcel表示 */
  .excel-loaded { background: #EBF3FB; border-radius: 8px; padding: 12px 14px; }
  .excel-loaded-name { font-size: 14px; font-weight: bold; color: #1F4E79; }
  .excel-loaded-meta { font-size: 12px; color: #666; margin-top: 3px; }
  .sheet-tags { display: flex; gap: 5px; flex-wrap: wrap; margin-top: 7px; }
  .sheet-tag  { background: #1F4E79; color: white; font-size: 10px;
                padding: 2px 7px; border-radius: 10px; }

  .btn { padding: 8px 16px; border: none; border-radius: 6px; font-size: 13px;
         font-weight: bold; cursor: pointer; transition: all 0.2s; white-space: nowrap; }
  .btn:hover:not(:disabled) { filter: brightness(0.9); }
  .btn:active:not(:disabled) { transform: scale(0.97); }
  .btn-blue   { background: #1F4E79; color: white; }
  .btn-orange { background: #C55A11; color: white; }
  .btn-calc   { background: #6B4C9A; color: white; font-size: 12px; padding: 6px 12px; }
  .btn-sm     { padding: 5px 10px; font-size: 11px; }
  .btn-gray   { background: #ccc; color: #777; cursor: not-allowed; }

  .result { margin-top: 10px; padding: 11px 13px; border-radius: 7px;
            font-size: 12px; line-height: 1.6; }
  .result.success { background: #E9F7EF; border: 1px solid #27AE60; color: #1a6b3a; }
  .result.error   { background: #FDE8E8; border: 1px solid #E74C3C; color: #a11a1a; }
  .result.info    { background: #EBF5FB; border: 1px solid #2980B9; color: #1a4a6b; }

  .badge { display: inline-block; background: #C55A11; color: white;
           font-size: 10px; padding: 1px 7px; border-radius: 9px; margin-left: 6px; }

  .spinner { display: inline-block; width: 14px; height: 14px;
             border: 2px solid rgba(255,255,255,0.4); border-top-color: white;
             border-radius: 50%; animation: spin 0.7s linear infinite;
             margin-right: 5px; vertical-align: middle; }
  @keyframes spin { to { transform: rotate(360deg); } }

  .hint { font-size: 11px; color: #999; margin-top: 8px; line-height: 1.6; }
  .dim { opacity: 0.45; }
  .hidden { display: none; }
</style>
</head>
<body>

<div class="header">
  <div>
    <div class="header-title">📊 リサーチ追記ツール</div>
    <div class="header-sub">Excel に Amazon / 1688 データを追記する</div>
  </div>
  <a href="/" class="back-link">← AucFanアプリ</a>
</div>

<div class="main">

  <!-- ① Excel選択 -->
  <div class="card">
    <div class="card-title">
      ① Excel ファイルを選択
      <button class="btn btn-blue btn-sm" onclick="refreshList()" style="margin-left:auto;">🔄 更新</button>
    </div>

    <!-- フォルダ変更（詳細） -->
    <details style="margin-bottom:10px;">
      <summary style="font-size:11px; color:#888; cursor:pointer;">📁 フォルダを変更する</summary>
      <div style="margin-top:8px; display:flex; gap:8px;">
        <input id="folderPath" class="folder-input" type="text" placeholder="フォルダパス">
        <button class="btn btn-blue btn-sm" onclick="refreshList()">スキャン</button>
      </div>
      <p class="hint">空白のままにすると保存先（Google Drive/AucFanToolData/リサーチ結果）を自動で探します。</p>
    </details>

    <!-- ファイル一覧 -->
    <div id="fileList">
      <div class="file-empty">読み込み中...</div>
    </div>

    <!-- 読み込み済み表示 -->
    <div id="excelLoaded" class="hidden" style="margin-top:12px;"></div>
  </div>

  <!-- ② Amazon取得 -->
  <div class="card">
    <div class="card-title">
      ② Amazon ライバルデータ取得
      <span id="amazonBadge" class="badge hidden">0件</span>
    </div>

    <!-- URL入力エリア -->
    <div style="display:flex; gap:8px; align-items:flex-start; margin-bottom:10px;">
      <div style="flex:1;">
        <input id="amazonUrl" class="folder-input" type="text"
               placeholder="Amazon URL を貼り付け　例: https://amzn.asia/d/07AQHZFg　または https://www.amazon.co.jp/dp/B0XXXX"
               style="width:100%; padding:9px 11px; font-size:13px;">
        <p class="hint" style="margin-top:5px;">
          短縮URL（amzn.asia/d/...）・通常URL どちらでも対応。複数ライバルは1件ずつ貼り付けて「取得→追記」を繰り返してください。
        </p>
      </div>
      <button id="btnAmazonUrl" class="btn btn-orange" onclick="fetchAmazonUrl()" style="margin-top:1px;">
        🔍 取得 → 追記
      </button>
    </div>

    <div id="amazonResult" class="hidden result"></div>
  </div>

  <!-- ③ 1688仕入れデータ取得 -->
  <div class="card">
    <div class="card-title">
      ③ 1688 仕入れデータ取得
      <span id="1688Badge" class="badge hidden">0件</span>
    </div>

    <!-- URL入力エリア -->
    <div style="display:flex; gap:8px; align-items:flex-start; margin-bottom:10px;">
      <div style="flex:1;">
        <input id="url1688" class="folder-input" type="text"
               placeholder="1688 商品URL を貼り付け　例: https://detail.1688.com/offer/XXXXXXXXXX.html"
               style="width:100%; padding:9px 11px; font-size:13px;">
        <p class="hint" style="margin-top:5px;">
          1688の商品ページURLを貼り付けて「取得→追記」を押してください。Sheet4/5 に追記されます。
        </p>
      </div>
      <button id="btn1688Url" class="btn btn-orange" onclick="fetch1688Url()" style="margin-top:1px;">
        🔍 取得 → 追記
      </button>
    </div>

    <div id="result1688" class="hidden result"></div>
  </div>

</div>

<script>
let loadedPath  = "";
let amazonCount = 0;

// ── ファイル一覧 ────────────────────────────────────────────
async function refreshList() {
  const folder = document.getElementById("folderPath").value.trim();
  const div    = document.getElementById("fileList");
  div.innerHTML = '<div class="file-empty">スキャン中...</div>';

  const res  = await fetch("/api/research/excel/list", {
    method: "POST",
    headers: {"Content-Type": "application/json"},
    body: JSON.stringify({folder: folder})
  });
  const data = await res.json();

  if (!data.success || !data.files.length) {
    const msg = data.error || "Excelファイルが見つかりません";
    div.innerHTML = `<div class="file-empty">📭 ${msg}</div>`;
    return;
  }

  // フォルダパスを表示
  document.getElementById("folderPath").placeholder = data.folder;

  div.innerHTML = '<div class="file-list">' +
    data.files.map(f => `
      <div class="file-item" onclick="selectFile('${f.path.replace(/'/g,"\\'")}', this)"
           title="${f.path}">
        <div class="file-icon">📄</div>
        <div class="file-info">
          <div class="file-name">${f.name}</div>
          <div class="file-meta">${f.modified} &nbsp;·&nbsp; ${f.size}</div>
        </div>
        ${f.amazon_count > 0
          ? `<div class="file-badge">Amazon ${f.amazon_count}件</div>`
          : ''}
      </div>`).join("") +
    '</div>';
}

async function selectFile(path, el) {
  // 選択ハイライト
  document.querySelectorAll(".file-item").forEach(e => e.classList.remove("selected"));
  el.classList.add("selected");

  // Excel情報取得
  const res  = await fetch("/api/research/excel/load", {
    method: "POST",
    headers: {"Content-Type": "application/json"},
    body: JSON.stringify({path: path})
  });
  const data = await res.json();
  const div  = document.getElementById("excelLoaded");

  if (data.success) {
    loadedPath  = path;
    amazonCount = data.amazon_count || 0;
    updateBadge();
    const warn = data.is_research ? "" :
      '<p style="color:#C55A11;font-size:11px;margin-top:4px;">⚠ 5シート構成ではない可能性があります</p>';
    div.innerHTML = `
      <div class="excel-loaded">
        <div class="excel-loaded-name">✅ ${data.filename}</div>
        <div class="excel-loaded-meta">
          ${data.title || "（タイトル未取得）"}
          &nbsp;·&nbsp; Amazon <strong>${amazonCount}件</strong>取得済み
        </div>
        <div class="sheet-tags">
          ${(data.sheets||[]).map(s=>`<span class="sheet-tag">${s}</span>`).join("")}
        </div>
        ${warn}
        <div style="margin-top:8px;">
          <a class="btn btn-blue btn-sm" href="/api/research/excel/download?path=${encodeURIComponent(path)}"
             download style="text-decoration:none;">
            ⬇ Excelをダウンロード
          </a>
        </div>
      </div>`;
    div.classList.remove("hidden");
  } else {
    div.innerHTML = `<div class="result error">❌ ${data.error}</div>`;
    div.classList.remove("hidden");
  }
}

function updateBadge() {
  const b = document.getElementById("amazonBadge");
  b.textContent = amazonCount + "件";
  b.classList.remove("hidden");
}

// ── Amazon URL取得 ──────────────────────────────────────────
let _fetchStartTime = null;
let _fetchPollTimer = null;

function _startProgressPolling(div) {
  _fetchStartTime = Date.now();
  _fetchPollTimer = setInterval(async () => {
    try {
      const r = await fetch("/api/research/amazon/status");
      const s = await r.json();
      if (!s.running) return;
      const sec = Math.floor((Date.now() - _fetchStartTime) / 1000);
      div.innerHTML =
        `<span class="spinner"></span>` +
        `<strong>${s.step}</strong><br>` +
        `<small style="color:#555;">経過 ${sec} 秒 ／ Amazonページを閉じないでください</small>`;
    } catch(e) {}
  }, 1000);
}

function _stopProgressPolling() {
  if (_fetchPollTimer) { clearInterval(_fetchPollTimer); _fetchPollTimer = null; }
}

async function fetchAmazonUrl() {
  if (!loadedPath) { alert("先に Excel ファイルを選択してください"); return; }

  const url = document.getElementById("amazonUrl").value.trim();
  if (!url) { alert("Amazon の URL を入力してください"); return; }

  const btn = document.getElementById("btnAmazonUrl");
  const div = document.getElementById("amazonResult");
  btn.innerHTML = '<span class="spinner"></span>取得中...';
  btn.disabled  = true;
  div.className = "result info";
  div.innerHTML =
    `<span class="spinner"></span><strong>① URLを解析中...</strong><br>` +
    `<small style="color:#555;">しばらくお待ちください（10〜30秒）</small>`;
  div.classList.remove("hidden");
  _startProgressPolling(div);

  try {
    const res  = await fetch("/api/research/amazon/fetch-url-append", {
      method: "POST",
      headers: {"Content-Type": "application/json"},
      body: JSON.stringify({path: loadedPath, url: url})
    });
    const data = await res.json();
    _stopProgressPolling();

    if (data.success) {
      amazonCount++;
      updateBadge();
      document.querySelectorAll(".file-item.selected .file-badge").forEach(b => {
        b.textContent = "Amazon " + amazonCount + "件";
      });
      // URLをクリア
      document.getElementById("amazonUrl").value = "";
      div.className = "result success";
      const imgInfo = data.image_count > 0
        ? `🖼 画像 ${data.image_count}枚保存 (amazon/${data.asin}/)`
        : (data.has_image ? "🖼 画像あり" : "");
      div.innerHTML =
        `✅ Sheet2 の ${data.row} 行目に追記 ${imgInfo}<br>` +
        `<strong>ASIN:</strong> ${data.asin} &nbsp; <strong>価格:</strong> ${data.price}<br>` +
        `<span style="color:#555;">${data.title}</span><br>` +
        `<button class="btn btn-calc" onclick="openCalculator('${data.asin}')" style="margin-top:8px;">` +
        `💴 FBA料金シミュレータで開く</button>`;
    } else if (data.duplicate) {
      // 重複ASIN
      div.className = "result info";
      div.innerHTML =
        `⚠️ <strong>ASIN ${data.asin}</strong> は既に ${data.row} 行目に追記済みです。<br>` +
        `<small style="color:#555;">別のライバルのURLを入力してください。</small><br>` +
        `<button class="btn btn-calc" onclick="openCalculator('${data.asin}')" style="margin-top:8px;">` +
        `💴 FBA料金シミュレータで開く</button>`;
    } else {
      div.className = "result error";
      div.textContent = "❌ " + data.error;
    }
  } catch(e) {
    _stopProgressPolling();
    div.className = "result error";
    div.textContent = "❌ 通信エラー: " + e.message;
  } finally {
    _stopProgressPolling();
    btn.innerHTML = "🔍 取得 → 追記";
    btn.disabled  = false;
  }
}

// ── FBA料金シミュレータ ────────────────────────────────────────
async function openCalculator(asin) {
  const btn = event.target;
  btn.textContent = "⏳ 起動中...";
  btn.disabled = true;
  try {
    const res  = await fetch("/api/research/amazon/open-calculator", {
      method: "POST",
      headers: {"Content-Type": "application/json"},
      body: JSON.stringify({asin: asin})
    });
    const data = await res.json();
    if (data.success) {
      btn.textContent = "✅ Chromeで開きました";
    } else {
      btn.textContent = "❌ " + data.error;
      btn.disabled = false;
    }
  } catch(e) {
    btn.textContent = "❌ 通信エラー";
    btn.disabled = false;
  }
}

// ── 1688 URL取得 ──────────────────────────────────────────────
let _1688FetchStartTime = null;
let _1688FetchPollTimer = null;
let fetch1688Count = 0;

function _start1688ProgressPolling(div) {
  _1688FetchStartTime = Date.now();
  _1688FetchPollTimer = setInterval(async () => {
    try {
      const r = await fetch("/api/research/1688/status");
      const s = await r.json();
      if (!s.running) return;
      const sec = Math.floor((Date.now() - _1688FetchStartTime) / 1000);
      div.innerHTML =
        `<span class="spinner"></span>` +
        `<strong>${s.step}</strong><br>` +
        `<small style="color:#555;">経過 ${sec} 秒 ／ 1688ページを閉じないでください</small>`;
    } catch(e) {}
  }, 1000);
}

function _stop1688ProgressPolling() {
  if (_1688FetchPollTimer) { clearInterval(_1688FetchPollTimer); _1688FetchPollTimer = null; }
}

async function fetch1688Url() {
  if (!loadedPath) { alert("先に Excel ファイルを選択してください"); return; }

  const url = document.getElementById("url1688").value.trim();
  if (!url) { alert("1688 の URL を入力してください"); return; }

  const btn = document.getElementById("btn1688Url");
  const div = document.getElementById("result1688");
  btn.innerHTML = '<span class="spinner"></span>取得中...';
  btn.disabled  = true;
  div.className = "result info";
  div.innerHTML =
    `<span class="spinner"></span><strong>① URLを解析中...</strong><br>` +
    `<small style="color:#555;">しばらくお待ちください（10〜30秒）</small>`;
  div.classList.remove("hidden");
  _start1688ProgressPolling(div);

  try {
    const res  = await fetch("/api/research/1688/fetch-url-append", {
      method: "POST",
      headers: {"Content-Type": "application/json"},
      body: JSON.stringify({path: loadedPath, url: url})
    });
    const data = await res.json();
    _stop1688ProgressPolling();

    if (data.success) {
      fetch1688Count++;
      document.getElementById("url1688").value = "";
      const badge = document.getElementById("1688Badge");
      badge.textContent = fetch1688Count + "件";
      badge.classList.remove("hidden");
      div.className = "result success";
      const imgInfo = data.image_count > 0
        ? `🖼 画像 ${data.image_count}枚保存 (1688/)` : "";
      div.innerHTML =
        `✅ Sheet4 の ${data.row} 行目から追記（${data.variant_count}バリアント）${imgInfo}<br>` +
        `<strong>ショップ:</strong> ${data.shop_name || "不明"} &nbsp;` +
        `<strong>最低単価:</strong> ${data.min_price} 元<br>` +
        `<span style="color:#555;">${data.title}</span>`;
    } else {
      div.className = "result error";
      div.textContent = "❌ " + data.error;
    }
  } catch(e) {
    _stop1688ProgressPolling();
    div.className = "result error";
    div.textContent = "❌ 通信エラー: " + e.message;
  } finally {
    _stop1688ProgressPolling();
    btn.innerHTML = "🔍 取得 → 追記";
    btn.disabled  = false;
  }
}

// Enterキーで送信
document.addEventListener("DOMContentLoaded", () => {
  const inp = document.getElementById("amazonUrl");
  if (inp) inp.addEventListener("keydown", e => { if (e.key === "Enter") fetchAmazonUrl(); });
  const inp2 = document.getElementById("url1688");
  if (inp2) inp2.addEventListener("keydown", e => { if (e.key === "Enter") fetch1688Url(); });
});

// 起動時にファイル一覧を読み込む
refreshList();
</script>
</body>
</html>
"""


@app.route("/research")
def research_page():
    """Excel 追記ツール（AucFan アプリ内ページ）"""
    return _RESEARCH_HTML


@app.route("/api/research/excel/list", methods=["POST"])
def api_research_excel_list():
    """指定フォルダ（省略時は OUTPUT_BASE_DIR）の xlsx ファイル一覧を返す。"""
    import time
    body   = request.get_json(silent=True) or {}
    folder = (body.get("folder") or "").strip() or config.OUTPUT_BASE_DIR
    folder_path = Path(folder)

    if not folder_path.exists():
        return jsonify({"success": False, "error": f"フォルダが見つかりません: {folder}", "files": []})

    try:
        import datetime as dt
        files = []
        # フラット（旧形式）+ サブフォルダ（新形式）の両方を拾う
        xlsx_paths = list(folder_path.glob("*.xlsx")) + list(folder_path.glob("*/*.xlsx"))
        xlsx_paths = sorted(xlsx_paths, key=lambda x: x.stat().st_mtime, reverse=True)

        for p in xlsx_paths:
            stat = p.stat()
            # Amazon取得済み件数を簡易チェック
            amazon_count = 0
            try:
                from openpyxl import load_workbook
                wb = load_workbook(str(p), read_only=True, data_only=True)
                if "②Amazonライバル" in wb.sheetnames:
                    ws = wb["②Amazonライバル"]
                    for r in range(4, (ws.max_row or 3) + 1):
                        if ws.cell(r, 1).value:
                            amazon_count += 1
                wb.close()
            except Exception:
                pass

            modified = dt.datetime.fromtimestamp(stat.st_mtime).strftime("%Y/%m/%d %H:%M")
            size_kb  = stat.st_size // 1024
            size_str = f"{size_kb}KB" if size_kb < 1024 else f"{size_kb//1024}MB"

            # サブフォルダ内のファイルは「フォルダ名/ファイル名」で表示
            rel = p.relative_to(folder_path)
            display_name = str(rel) if len(rel.parts) > 1 else p.name

            files.append({
                "name":          display_name,
                "path":          str(p),
                "modified":      modified,
                "size":          size_str,
                "amazon_count":  amazon_count,
            })

        return jsonify({"success": True, "folder": str(folder_path), "files": files})

    except Exception as e:
        logger.error(f"ファイル一覧取得エラー: {e}")
        return jsonify({"success": False, "error": str(e), "files": []})


@app.route("/api/research/excel/download", methods=["GET"])
def api_research_excel_download():
    """Excelファイルをブラウザにダウンロードさせる。"""
    from urllib.parse import unquote
    from flask import send_file
    path = unquote(request.args.get("path", "").strip())
    if not path:
        return "パスが指定されていません", 400
    p = Path(path)
    if not p.exists() or not p.is_file():
        return "ファイルが見つかりません", 404
    # パストラバーサル防止: OUTPUT_BASE_DIR 配下のみ許可
    try:
        p.resolve().relative_to(Path(config.OUTPUT_BASE_DIR).resolve())
    except ValueError:
        return "アクセスできないパスです", 403
    return send_file(
        str(p),
        as_attachment=True,
        download_name=p.name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/api/research/excel/load", methods=["POST"])
def api_research_excel_load():
    """Excel ファイルの情報を返す。"""
    from excel_append import get_excel_info
    body = request.get_json(silent=True) or {}
    path = body.get("path", "").strip()
    if not path:
        return jsonify({"success": False, "error": "パスが指定されていません"})
    return jsonify(get_excel_info(path))


@app.route("/api/research/amazon/append", methods=["POST"])
def api_research_amazon_append():
    """Chrome の現在開いている Amazon ページを取得して Excel に追記する。"""
    from amazon_scraper import fetch_amazon_product
    from excel_append import append_amazon

    body = request.get_json(silent=True) or {}
    excel_path = body.get("path", "").strip()
    if not excel_path:
        return jsonify({"success": False, "error": "Excelパスが指定されていません"})

    amazon_data = fetch_amazon_product()
    if not amazon_data.get("success"):
        return jsonify(amazon_data), 400

    result = append_amazon(excel_path, amazon_data)
    status = 200 if result.get("success") else 500
    return jsonify(result), status


@app.route("/api/research/amazon/status")
def api_research_amazon_status():
    """Amazon取得中ステータスをポーリングで返す"""
    return jsonify(_research_fetch_status)


@app.route("/api/research/amazon/fetch-url-append", methods=["POST"])
def api_research_amazon_fetch_url_append():
    """URLを指定してAmazonデータを取得し、Excelに追記する。短縮URL対応。"""
    global _research_fetch_status
    from amazon_scraper import fetch_amazon_from_url
    from excel_append import append_amazon

    body = request.get_json(silent=True) or {}
    excel_path = body.get("path", "").strip()
    url        = body.get("url", "").strip()

    if not excel_path:
        return jsonify({"success": False, "error": "Excelパスが指定されていません"})
    if not url:
        return jsonify({"success": False, "error": "URLが指定されていません"})

    # 2重実行防止：前の取得がまだ実行中なら即座に拒否
    if not _research_fetch_lock.acquire(blocking=False):
        return jsonify({
            "success": False,
            "error": "取得処理が実行中です。完了をお待ちください。"
        }), 429

    _research_fetch_status = {"running": True, "step": "① URLを解析中..."}
    try:
        _research_fetch_status["step"] = "② ChromeでAmazonページを開いています..."
        amazon_data = fetch_amazon_from_url(url)
        if not amazon_data.get("success"):
            return jsonify(amazon_data), 400

        _research_fetch_status["step"] = "③ Excelに書き込み中..."
        result = append_amazon(excel_path, amazon_data)
        status = 200 if result.get("success") else 500
        return jsonify(result), status
    finally:
        _research_fetch_status = {"running": False, "step": ""}
        _research_fetch_lock.release()


@app.route("/api/research/amazon/open-calculator", methods=["POST"])
def api_open_fba_calculator():
    """
    FBA料金シミュレータ（非ログイン版）をChromeで開き、ASINを自動入力する。
    既存のSelenium接続（port 9222）を使用。
    """
    body = request.get_json(silent=True) or {}
    asin = body.get("asin", "").strip()
    if not asin:
        return jsonify({"success": False, "error": "ASINが指定されていません"})

    try:
        from amazon_scraper import _connect_chrome
        import time
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC

        CALC_URL = "https://sellercentral.amazon.co.jp/revcal?ref=RC2nonlogin"

        driver = _connect_chrome()
        if not driver:
            return jsonify({"success": False, "error": "Chromeに接続できません（port 9222）"})

        # 既存のシミュレータタブを探す（revcal URLのみ再利用。ログインページは除外）
        calc_handle = None
        for handle in driver.window_handles:
            driver.switch_to.window(handle)
            if "revcal" in driver.current_url:
                calc_handle = handle
                break

        if calc_handle:
            # 既存タブに移動してURLをリロード
            driver.switch_to.window(calc_handle)
            driver.get(CALC_URL)
        else:
            # 新規タブを開く
            driver.execute_script("window.open('');")
            calc_handle = driver.window_handles[-1]
            driver.switch_to.window(calc_handle)
            driver.get(CALC_URL)

        # ページ読み込み待機（最大15秒）
        time.sleep(2)

        # ASIN入力フィールドを探す
        # Seller Central の revcal は KAT UI フレームワーク製で、
        # <kat-input> などのカスタム要素が Shadow DOM 内に <input> を持つ。
        # 通常の CSS セレクターは Shadow DOM を貫通できないため、
        # JS で再帰的に Shadow Root を辿って検索する。

        input_el = None

        # まず通常セレクターで試す（Shadow DOM 外にある場合）
        normal_selectors = [
            "input[id*='asin']",
            "input[name*='asin']",
            "input[data-testid*='asin']",
            "input[placeholder*='ASIN']",
            "input[placeholder*='asin']",
            "#asin-search-input",
        ]
        for sel in normal_selectors:
            try:
                els = driver.find_elements(By.CSS_SELECTOR, sel)
                for el in els:
                    if el.is_displayed() and el.is_enabled():
                        input_el = el
                        break
                if input_el:
                    break
            except Exception:
                pass

        if not input_el:
            # Shadow DOM を再帰的に検索する JS
            # nav/header 内の入力欄はスキップ
            try:
                input_el = driver.execute_script("""
                    var NAV_SELECTOR =
                        '#navbar, #navbar-main, nav, header,' +
                        '[id*="nav-search"], [id*="topnav"],' +
                        '[class*="navbar"], [class*="nav-bar"], [class*="nav-search"]';

                    function isInNav(el) {
                        try { return !!el.closest(NAV_SELECTOR); } catch(e) { return false; }
                    }

                    function findInput(root) {
                        // 1. ASIN関連の属性を持つ input を優先検索
                        var priority = root.querySelectorAll(
                            'input[id*="asin"], input[name*="asin"],' +
                            'input[placeholder*="ASIN"], input[data-testid*="asin"]'
                        );
                        for (var i = 0; i < priority.length; i++) {
                            var el = priority[i];
                            if (el.offsetParent !== null && !isInNav(el)) return el;
                        }
                        // 2. 表示中の text input（nav除外）
                        var inputs = root.querySelectorAll('input[type="text"], input:not([type])');
                        for (var j = 0; j < inputs.length; j++) {
                            var el = inputs[j];
                            if (el.offsetParent !== null && !isInNav(el)) return el;
                        }
                        // 3. Shadow DOM 内を再帰検索
                        var all = root.querySelectorAll('*');
                        for (var k = 0; k < all.length; k++) {
                            if (all[k].shadowRoot) {
                                var found = findInput(all[k].shadowRoot);
                                if (found) return found;
                            }
                        }
                        return null;
                    }
                    return findInput(document);
                """)
            except Exception:
                pass

        if not input_el:
            return jsonify({
                "success": False,
                "error": "ASIN入力欄が見つかりませんでした。シミュレータページが表示されているか確認してください。"
            })

        # 入力フィールドをクリアして ASIN を入力
        driver.execute_script("arguments[0].value = '';", input_el)
        driver.execute_script("arguments[0].click();", input_el)
        time.sleep(0.3)
        input_el.clear()
        input_el.send_keys(asin)
        time.sleep(0.3)

        # 送信ボタンを探してクリック（Shadow DOM 内も検索）
        submitted = False
        try:
            submitted = driver.execute_script("""
                function findButton(root) {
                    // type=submit または search/submit を含む id/class のボタン
                    var btns = root.querySelectorAll(
                        'button[type="submit"], input[type="submit"],' +
                        'button[id*="search"], button[id*="submit"],' +
                        '.kat-button--primary, button.a-button-primary,' +
                        'kat-button[variant="primary"]'
                    );
                    for (var i = 0; i < btns.length; i++) {
                        if (btns[i].offsetParent !== null) {
                            btns[i].click();
                            return true;
                        }
                    }
                    // Shadow DOM 内を再帰検索
                    var all = root.querySelectorAll('*');
                    for (var k = 0; k < all.length; k++) {
                        if (all[k].shadowRoot) {
                            if (findButton(all[k].shadowRoot)) return true;
                        }
                    }
                    return false;
                }
                return findButton(document);
            """)
        except Exception:
            pass

        if not submitted:
            # ボタンが見つからない場合は Enter キーで送信
            from selenium.webdriver.common.keys import Keys
            input_el.send_keys(Keys.RETURN)

        logger.info(f"FBA料金シミュレータ: ASIN={asin} 入力完了")
        return jsonify({"success": True, "asin": asin, "url": CALC_URL})

    except Exception as e:
        logger.error(f"FBA料金シミュレータ起動エラー: {e}", exc_info=True)
        return jsonify({"success": False, "error": str(e)})


@app.route("/api/research/1688/status")
def api_research_1688_status():
    """1688取得中ステータスをポーリングで返す"""
    return jsonify(_research_1688_fetch_status)


@app.route("/api/research/1688/fetch-url-append", methods=["POST"])
def api_research_1688_fetch_url_append():
    """URLを指定して1688データを取得し、Excelに追記する。"""
    global _research_1688_fetch_status
    from excel_append import append_1688

    body = request.get_json(silent=True) or {}
    excel_path = body.get("path", "").strip()
    url        = body.get("url", "").strip()

    if not excel_path:
        return jsonify({"success": False, "error": "Excelパスが指定されていません"})
    if not url:
        return jsonify({"success": False, "error": "URLが指定されていません"})

    # 2重実行防止：前の取得がまだ実行中なら即座に拒否
    if not _research_1688_fetch_lock.acquire(blocking=False):
        return jsonify({
            "success": False,
            "error": "取得処理が実行中です。完了をお待ちください。"
        }), 429

    _research_1688_fetch_status = {"running": True, "step": "① URLを解析中..."}
    try:
        from scraper_1688 import fetch_1688_from_url
        _research_1688_fetch_status["step"] = "② Chromeで1688ページを開いています..."
        data_1688 = fetch_1688_from_url(url)
        if not data_1688.get("success"):
            return jsonify(data_1688), 400

        _research_1688_fetch_status["step"] = "③ Excelに書き込み中..."
        result = append_1688(excel_path, data_1688)
        status = 200 if result.get("success") else 500
        return jsonify(result), status
    finally:
        _research_1688_fetch_status = {"running": False, "step": ""}
        _research_1688_fetch_lock.release()


# ─────────────────────────────────────────────
# エラーハンドラー
# ─────────────────────────────────────────────

@app.errorhandler(404)
def not_found(e):
    return jsonify({"error": "Not Found"}), 404


@app.errorhandler(500)
def server_error(e):
    return jsonify({"error": str(e)}), 500


# ─────────────────────────────────────────────
# 起動
# ─────────────────────────────────────────────

def check_env():
    """起動前チェック"""
    issues = []

    if not config.GEMINI_API_KEY:
        issues.append("⚠️  GEMINI_API_KEY が未設定です（.env ファイルを確認）。pHashのみで動作します。")

    for issue in issues:
        logger.warning(issue)

    logger.info("=" * 60)
    logger.info("  AucFan リサーチツール")
    logger.info("=" * 60)
    logger.info(f"  Flask URL  : http://{config.FLASK_HOST}:{config.FLASK_PORT}")
    logger.info(f"  Chrome接続 : {config.CHROME_DEBUG_HOST}:{config.CHROME_DEBUG_PORT}")
    logger.info(f"  出力先     : {config.OUTPUT_BASE_DIR}/")
    logger.info(f"  Gemini     : {'有効' if config.GEMINI_ENABLED and config.GEMINI_API_KEY else '無効'}")
    logger.info("=" * 60)
    logger.info("")
    logger.info("使い方:")
    logger.info("  1. AucFanで検索条件を設定し1ページ目を表示した状態にする")
    logger.info(f"  2. ブラウザで http://localhost:{config.FLASK_PORT} を開く（自動で開きます）")
    logger.info("  3. UIでキーワードを入力して「スクレイピング開始」をクリック")
    logger.info("")


def _auto_load_latest_session():
    """
    起動時に最新のセッションを自動ロードする。
    前回のスクレイピング結果をすぐに表示できるようにする。
    セッションが存在しない場合は何もしない。
    """
    global _data_manager, _gemini_client, _session_output_dir

    base = Path(config.OUTPUT_BASE_DIR)
    if not base.exists():
        return

    # 最新のセッションフォルダを探す（更新日時順）
    latest_dir = None
    for d in sorted(base.iterdir(), reverse=True):
        if d.is_dir() and (d / "progress.json").exists():
            latest_dir = d
            break

    if latest_dir is None:
        return

    try:
        session_name = latest_dir.name
        _session_output_dir = latest_dir
        _data_manager = DataManager(session_name, latest_dir)
        _data_manager.load_previous_session()
        _gemini_client = GeminiClient()
        logger.info(
            f"[起動時自動ロード] セッション: {session_name}"
            f" ({_data_manager.total_items}件)"
        )
    except Exception as e:
        logger.warning(f"[起動時自動ロード] 失敗（スキップ）: {e}")
        # ロード失敗してもFlaskは起動する


if __name__ == "__main__":
    check_env()

    # 最新セッションをバックグラウンドで自動ロード（Google Drive 同期中でもFlaskをすぐ起動するため）
    auto_load_thread = threading.Thread(target=_auto_load_latest_session, daemon=True)
    auto_load_thread.start()

    # ブラウザを自動で開く（少し遅らせる）
    def open_browser():
        time.sleep(1.5)
        webbrowser.open(f"http://localhost:{config.FLASK_PORT}")

    browser_thread = threading.Thread(target=open_browser, daemon=True)
    browser_thread.start()

    # Flask 起動（スレッド化されたSSEのため use_reloader=False）
    app.run(
        host=config.FLASK_HOST,
        port=config.FLASK_PORT,
        debug=False,
        use_reloader=False,
        threaded=True,
    )
