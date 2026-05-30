"""
excel_exporter.py — Aucfan リサーチシート Excel エクスポーター

【役割】
  AucFan の商品データを Sheet1（概要）に書き込み、
  Sheet2〜5 を空で用意した 5 シート構成の Excel を生成する。

  Sheet1: ①概要      — AucFanデータ + 手入力2セル（販売価格・FBA手数料）+ 自動計算
  Sheet2: ②Amazonライバル — Amazon競合データ追記用（空）
  Sheet3: ③Amazonテキスト — スペック文オマージュ用（空）
  Sheet4: ④1688仕入れ   — 利益計算メイン（空）
  Sheet5: ⑤1688テキスト  — 仕入れ詳細（空）

  Sheet2〜5 への追記は /research ページ（routes/research.py）で行う。
"""
from __future__ import annotations

import io
import logging
import re
import shutil
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ── 定数 ──────────────────────────────────────────────────────────────
THUMB_MAX_W = 120
THUMB_MAX_H = 100

# シート名
SHEET_NAMES = [
    "①概要",
    "②Amazonライバル",
    "③Amazonテキスト",
    "④1688仕入れ",
    "⑤1688テキスト",
]

# カラー定義
COLOR_HEADER_BLUE   = "1F4E79"   # 濃い青（Sheet1ヘッダー）
COLOR_HEADER_ORANGE = "C55A11"   # オレンジ（Amazon系）
COLOR_HEADER_GREEN  = "375623"   # 緑（1688系）
COLOR_SECTION_LIGHT = "D9E1F2"   # 薄青（セクション背景）
COLOR_INPUT_YELLOW  = "FFFF99"   # 黄色（手入力セル）
# SP-API自動転記セル色は config.EXCEL_COLOR_SP_API で管理（デフォルト: "CCE5FF" 水色）
COLOR_WHITE         = "FFFFFF"
COLOR_GRAY          = "F2F2F2"


# ── ファイル名サニタイズ ────────────────────────────────────────────────
def sanitize_filename(title: str, max_len: int = 60) -> str:
    title = re.sub(r'[\\/:*?"<>|]', '', title)
    title = re.sub(r'\s+', ' ', title).strip()
    return title[:max_len] if title else "リサーチ"


# ── 画像リサイズ ────────────────────────────────────────────────────────
def _make_thumb(src_path: Path, tmp_dir: str) -> Optional[str]:
    try:
        from PIL import Image as PILImage
        img = PILImage.open(src_path).convert("RGBA")
        img.thumbnail((THUMB_MAX_W, THUMB_MAX_H), PILImage.LANCZOS)
        bg = PILImage.new("RGB", img.size, (255, 255, 255))
        bg.paste(img, mask=img.split()[3])
        out_path = str(Path(tmp_dir) / src_path.name)
        bg.save(out_path, "JPEG", quality=85)
        return out_path
    except Exception as e:
        logger.debug(f"画像処理スキップ ({src_path.name}): {e}")
        return None


# ── スタイルヘルパー ──────────────────────────────────────────────────
def _header_font(color: str = COLOR_WHITE, bold: bool = True) -> Font:
    return Font(name="BIZ UDGothic", bold=bold, color=color)

def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)

def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def _thin_border() -> Border:
    thin = Side(style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


# ── Sheet1: ①概要 ───────────────────────────────────────────────────
def _build_sheet1(ws, group: dict):
    """Sheet1（①概要）を構築する。"""
    now = datetime.now()

    # 列幅設定
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 6
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 20

    # ── タイトル行 ─────────────────────────────────────────────────
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = f"リサーチシート　{group['title']}"
    title_cell.font = Font(name="BIZ UDGothic", bold=True, size=13, color=COLOR_WHITE)
    title_cell.fill = _fill(COLOR_HEADER_BLUE)
    title_cell.alignment = _center()

    # ── 作成日 ─────────────────────────────────────────────────────
    ws.row_dimensions[2].height = 18
    ws["D2"].value = "作成日"
    ws["D2"].font = Font(name="BIZ UDGothic", bold=True)
    ws["D2"].fill = _fill(COLOR_GRAY)
    ws["E2"].value = now.strftime("%Y/%m/%d")
    ws["E2"].font = Font(name="BIZ UDGothic")

    # ── セクション: AucFan参考データ ──────────────────────────────
    ws.row_dimensions[4].height = 20
    ws.merge_cells("A4:E4")
    sec = ws["A4"]
    sec.value = "▼ AucFan 参考データ"
    sec.font = _header_font(COLOR_WHITE)
    sec.fill = _fill(COLOR_HEADER_BLUE)
    sec.alignment = _left()

    rows_aucfan = [
        (5,  "商品キーワード",  group["title"],          False),
        (6,  "AucFan最安値",   group["min_total"] or "", False),
        (7,  "件数（4件以上）", group["group_size"],     False),
        (8,  "AucFan URL",     group.get("url", ""),     False),
    ]
    for r, label, val, is_input in rows_aucfan:
        ws.row_dimensions[r].height = 18
        ws[f"A{r}"].value = label
        ws[f"A{r}"].font = Font(name="BIZ UDGothic", bold=True)
        ws[f"A{r}"].fill = _fill(COLOR_GRAY)
        ws[f"A{r}"].border = _thin_border()
        ws.merge_cells(f"B{r}:E{r}")
        ws[f"B{r}"].value = val
        ws[f"B{r}"].font = Font(name="BIZ UDGothic")
        ws[f"B{r}"].border = _thin_border()
        if is_input:
            ws[f"B{r}"].fill = _fill(COLOR_INPUT_YELLOW)
        # URL行はハイパーリンク設定
        if label == "AucFan URL" and val:
            ws[f"B{r}"].hyperlink = str(val)
            ws[f"B{r}"].font = Font(name="BIZ UDGothic", color="0563C1", underline="single")

    # ── AucFanサムネイル欄 ─────────────────────────────────────────
    ws.row_dimensions[9].height = 90
    ws["A9"].value = "AucFan画像"
    ws["A9"].font = Font(name="BIZ UDGothic", bold=True)
    ws["A9"].fill = _fill(COLOR_GRAY)
    ws["A9"].border = _thin_border()
    ws.merge_cells("B9:E9")
    ws["B9"].value = "（画像は自動埋め込み）"
    ws["B9"].font = Font(name="BIZ UDGothic", color="999999", italic=True)
    ws["B9"].alignment = _center()

    # ── セクション: 販売価格・FBA（手入力） ───────────────────────
    ws.row_dimensions[11].height = 20
    ws.merge_cells("A11:E11")
    sec2 = ws["A11"]
    sec2.value = "▼ 販売価格・FBA手数料（手入力=黄 / SP-API自動=水色）"
    sec2.font = _header_font(COLOR_WHITE)
    sec2.fill = _fill(COLOR_HEADER_ORANGE)
    sec2.alignment = _left()

    input_rows = [
        (12, "販売予定価格（円）", ""),
        (13, "FBA手数料（円）",   ""),
    ]
    for r, label, val in input_rows:
        ws.row_dimensions[r].height = 22
        ws[f"A{r}"].value = label
        ws[f"A{r}"].font = Font(name="BIZ UDGothic", bold=True)
        ws[f"A{r}"].fill = _fill(COLOR_GRAY)
        ws[f"A{r}"].border = _thin_border()
        ws.merge_cells(f"B{r}:E{r}")
        ws[f"B{r}"].value = val
        ws[f"B{r}"].fill = _fill(COLOR_INPUT_YELLOW)
        ws[f"B{r}"].font = Font(name="BIZ UDGothic")
        ws[f"B{r}"].border = _thin_border()

    # ── セクション: 利益計算（自動） ──────────────────────────────
    ws.row_dimensions[15].height = 20
    ws.merge_cells("A15:E15")
    sec3 = ws["A15"]
    sec3.value = "▼ 利益計算（自動）"
    sec3.font = _header_font(COLOR_WHITE)
    sec3.fill = _fill(COLOR_HEADER_GREEN)
    sec3.alignment = _left()

    # B12=販売予定価格, B13=FBA手数料
    calc_rows = [
        (16, "原価（円）",   "（④1688仕入れシートに入力後、自動参照）"),
        (17, "利益（円）",   '=IF(B12="","",IF(B16="","",(B12-B13-B16)))'),
        (18, "利益率",       '=IF(OR(B12="",B12=0),"",TEXT(B17/B12,"0.0%"))'),
        (19, "判定",         '=IF(OR(B17="",B18=""),"",IF(AND(VALUE(SUBSTITUTE(B18,"%",""))/100>=0.25,B17>=450),"◎ GO","× 再検討"))'),
    ]
    for r, label, formula in calc_rows:
        ws.row_dimensions[r].height = 22
        ws[f"A{r}"].value = label
        ws[f"A{r}"].font = Font(name="BIZ UDGothic", bold=True)
        ws[f"A{r}"].fill = _fill(COLOR_GRAY)
        ws[f"A{r}"].border = _thin_border()
        ws.merge_cells(f"B{r}:E{r}")
        ws[f"B{r}"].value = formula
        ws[f"B{r}"].font = Font(name="BIZ UDGothic", bold=(r == 18))
        ws[f"B{r}"].border = _thin_border()
        if r == 18:
            ws[f"B{r}"].alignment = _center()

    # メモ欄
    ws.row_dimensions[20].height = 20
    ws.merge_cells("A20:E20")
    ws["A20"].value = "▼ メモ"
    ws["A20"].font = _header_font(COLOR_WHITE)
    ws["A20"].fill = _fill(COLOR_HEADER_BLUE)
    ws["A20"].alignment = _left()

    ws.row_dimensions[21].height = 60
    ws.merge_cells("A21:E21")
    ws["A21"].font = Font(name="BIZ UDGothic")
    ws["A21"].border = _thin_border()
    ws["A21"].alignment = Alignment(vertical="top", wrap_text=True)


# ── Sheet2: ②Amazonライバル ──────────────────────────────────────────
def _build_sheet2(ws):
    """Sheet2（②Amazonライバル）のヘッダーを作成する。データは /research ページで追記。"""
    # 列幅
    col_widths = [12, 45, 10, 8, 12, 8, 40, 25, 18]
    headers    = ["ASIN", "タイトル", "価格", "評価", "レビュー数", "A+", "URL（実）", "入力URL", "画像"]
    for i, (w, h) in enumerate(zip(col_widths, headers), 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # タイトル行
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A1:H1")
    ws["A1"].value = "② Amazon ライバル"
    ws["A1"].font = _header_font(COLOR_WHITE, bold=True)
    ws["A1"].fill = _fill(COLOR_HEADER_ORANGE)
    ws["A1"].alignment = _center()

    # ヘッダー行
    ws.row_dimensions[2].height = 22
    for col, label in enumerate(headers, 1):
        c = ws.cell(2, col)
        c.value = label
        c.font = Font(name="BIZ UDGothic", bold=True, color=COLOR_WHITE)
        c.fill = _fill(COLOR_HEADER_ORANGE)
        c.alignment = _center()
        c.border = _thin_border()

    # 補足メモ
    ws.row_dimensions[3].height = 16
    ws.merge_cells("A3:H3")
    ws["A3"].value = "← /research ページの「取得→追記」ボタンで追記されます"
    ws["A3"].font = Font(name="BIZ UDGothic", italic=True, color="999999")


# ── Sheet3: ③Amazonテキスト ─────────────────────────────────────────
def _build_sheet3(ws):
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 80

    ws.row_dimensions[1].height = 28
    ws.merge_cells("A1:B1")
    ws["A1"].value = "③ Amazon テキスト（スペック文オマージュ用）"
    ws["A1"].font = _header_font(COLOR_WHITE, bold=True)
    ws["A1"].fill = _fill(COLOR_HEADER_ORANGE)
    ws["A1"].alignment = _center()

    ws.row_dimensions[2].height = 16
    ws.merge_cells("A2:B2")
    ws["A2"].value = "← /research ページの「取得→追記」ボタンで追記されます"
    ws["A2"].font = Font(name="BIZ UDGothic", italic=True, color="999999")


# ── Sheet4: ④1688仕入れ ─────────────────────────────────────────────
def _build_sheet4(ws):
    # A=仕入れ選択 B=ショップ名 C=ショップURL D=信頼度 E=入驻年数
    # F=商品名(中) G=商品名(日) H=バリアント(中) I=バリアント(日) J=在庫数
    # K=単価(CNY) L=係数(CNY→JPY) M=MOQ
    # N=仕入総額(CNY)=K×M  O=仕入総額(JPY)=K×M×L  P=原価/個(JPY)=K×L
    # Q=利益(JPY) R=利益率 S=判定
    col_widths = [8, 20, 36, 12, 8, 26, 26, 20, 20, 8, 10, 6, 6, 12, 12, 12, 10, 8, 10]
    headers    = [
        "仕入れ選択", "ショップ名", "ショップURL", "信頼度", "年数",
        "商品名（中）", "商品名（日）", "バリアント（中）", "バリアント（日）", "在庫数",
        "単価\n(CNY)", "係数\n(×JPY)", "MOQ",
        "仕入総額\n(CNY)", "仕入総額\n(JPY)", "原価/個\n(JPY)",
        "利益\n(JPY)", "利益率", "判定",
    ]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 28
    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    ws["A1"].value = "④ 1688 仕入れ（利益計算メイン）"
    ws["A1"].font = _header_font(COLOR_WHITE, bold=True)
    ws["A1"].fill = _fill(COLOR_HEADER_GREEN)
    ws["A1"].alignment = _center()

    ws.row_dimensions[2].height = 34
    for col, label in enumerate(headers, 1):
        c = ws.cell(2, col)
        c.value = label
        c.font = Font(name="BIZ UDGothic", bold=True, color=COLOR_WHITE)
        c.fill = _fill(COLOR_HEADER_GREEN)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _thin_border()

    ws.row_dimensions[3].height = 16
    ws.merge_cells(f"A3:{get_column_letter(len(headers))}3")
    ws["A3"].value = "← /research の「1688取得→追記」ボタンで追記されます"
    ws["A3"].font = Font(name="BIZ UDGothic", italic=True, color="999999")


# ── Sheet5: ⑤1688テキスト ───────────────────────────────────────────
def _build_sheet5(ws):
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 80

    ws.row_dimensions[1].height = 28
    ws.merge_cells("A1:B1")
    ws["A1"].value = "⑤ 1688 テキスト（仕入れ詳細）"
    ws["A1"].font = _header_font(COLOR_WHITE, bold=True)
    ws["A1"].fill = _fill(COLOR_HEADER_GREEN)
    ws["A1"].alignment = _center()

    ws.row_dimensions[2].height = 16
    ws.merge_cells("A2:B2")
    ws["A2"].value = "← /research ページの「1688取得→追記」ボタンで追記されます"
    ws["A2"].font = Font(name="BIZ UDGothic", italic=True, color="999999")


# ── グループ情報取得 ────────────────────────────────────────────────
def _get_group(dm, group_id: str, session_name: str = "") -> Optional[dict]:
    all_items = dm.get_all_items()
    g_items = [i for i in all_items
               if (i.get("group_id") or i.get("item_id")) == group_id]
    if not g_items:
        return None

    rep = next((i for i in g_items if i.get("item_id") == group_id), g_items[0])
    totals = [float(i.get("total") or 0) for i in g_items if float(i.get("total") or 0) > 0]

    sellers: list[str] = []
    seen: set[str] = set()
    for i in g_items:
        sid = (i.get("seller_id") or "").strip()
        if sid and sid not in seen:
            seen.add(sid)
            sellers.append(sid)
        if len(sellers) >= 3:
            break

    thumb_path: Optional[Path] = None
    th = rep.get("thumbnail_local", "") or ""
    if th:
        p = Path(th)
        if p.exists():
            thumb_path = p
        elif session_name:
            try:
                import config
                fallback = (
                    Path(config.LOCAL_IMAGE_CACHE_DIR)
                    / session_name / "images" / Path(th).name
                )
                if fallback.exists():
                    thumb_path = fallback
            except Exception:
                pass

    return {
        "title":      rep.get("title_short") or rep.get("title_full") or "",
        "group_size": int(rep.get("group_size") or len(g_items)),
        "min_total":  int(min(totals)) if totals else 0,
        "sellers":    sellers,
        "url":        rep.get("url", "") or "",
        "thumb_path": thumb_path,
    }


# ── 公開 API ───────────────────────────────────────────────────────────
def generate_excel_single(dm, group_id: str,
                          embed_images: bool = True) -> Optional[tuple[bytes, str]]:
    return generate_excel_single_with_session(dm, group_id, group_id, embed_images)


def generate_excel_single_with_session(dm, group_id: str, session_name: str,
                                       embed_images: bool = True) -> Optional[tuple[bytes, str]]:
    """
    5シート構成のExcelを生成し (bytes, filename) を返す。
    Sheet1にAucFanデータを書き込み、Sheet2〜5はヘッダーのみ（空）で作成。
    """
    try:
        group = _get_group(dm, group_id, session_name)
        if not group:
            logger.warning(f"グループが見つかりません: {group_id}")
            return None

        # ── ワークブック作成 ─────────────────────────────────────
        wb = Workbook()

        # デフォルトシートをSheet1として使用
        ws1 = wb.active
        ws1.title = SHEET_NAMES[0]
        _build_sheet1(ws1, group)

        # Sheet2〜5を追加
        ws2 = wb.create_sheet(SHEET_NAMES[1])
        _build_sheet2(ws2)

        ws3 = wb.create_sheet(SHEET_NAMES[2])
        _build_sheet3(ws3)

        ws4 = wb.create_sheet(SHEET_NAMES[3])
        _build_sheet4(ws4)

        ws5 = wb.create_sheet(SHEET_NAMES[4])
        _build_sheet5(ws5)

        # Sheet1をアクティブに
        wb.active = ws1

        # ── AucFan画像埋め込み ───────────────────────────────────
        tmp_dir = tempfile.mkdtemp()
        try:
            if group["thumb_path"] and embed_images:
                thumb_file = _make_thumb(group["thumb_path"], tmp_dir)
                if thumb_file:
                    try:
                        from openpyxl.drawing.image import Image as XLImage
                        xl_img = XLImage(thumb_file)
                        xl_img.anchor = "B9"
                        ws1.add_image(xl_img)
                        ws1["B9"].value = None
                    except Exception as e:
                        logger.warning(f"画像埋め込みエラー: {e}")

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
        finally:
            shutil.rmtree(tmp_dir, ignore_errors=True)

        filename = f"{sanitize_filename(group['title'])}_リサーチ.xlsx"
        logger.info(f"Excel生成完了（5シート）: {filename}")
        return buf.read(), filename

    except Exception as e:
        logger.error(f"Excel生成エラー: {e}", exc_info=True)
        return None
