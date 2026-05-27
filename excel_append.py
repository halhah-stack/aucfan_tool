"""
excel_append.py — Excel リサーチシート追記モジュール

app.py の /research エンドポイントから呼び出される。
aucfan ツールとは独立して使えるよう、依存を最小限に保つ。

【提供する関数】
  get_excel_info(path)              → Excelの現在状態を返す
  append_amazon(excel_path, data)   → Sheet2/3 にAmazonデータを追記
  download_image(url, save_dir, name) → 画像をローカル保存
"""
from __future__ import annotations

import logging
import re
import shutil
import tempfile
import urllib.request
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)

# シート名（excel_exporter.py と揃える）
SHEET_OVERVIEW     = "①概要"
SHEET_AMAZON_LIST  = "②Amazonライバル"
SHEET_AMAZON_TEXT  = "③Amazonテキスト"
SHEET_1688_LIST    = "④1688仕入れ"
SHEET_1688_TEXT    = "⑤1688テキスト"

IMAGE_FOLDER_NAME  = "amazon"   # Excelと同フォルダ内の amazon/ サブフォルダ


# ── 画像ダウンロード ────────────────────────────────────────────────────
def download_image(url: str, save_dir: Path, name: str) -> Optional[Path]:
    """画像URLをローカルに保存して Path を返す。失敗時は None。"""
    if not url:
        return None
    try:
        save_dir.mkdir(parents=True, exist_ok=True)
        ext = ".png" if ".png" in url.lower() else ".jpg"
        dest = save_dir / f"{name}{ext}"
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=10) as resp:
            dest.write_bytes(resp.read())
        logger.info(f"画像保存: {dest}")
        return dest
    except Exception as e:
        logger.warning(f"画像ダウンロード失敗 ({url}): {e}")
        return None


def ensure_product_folders(excel_path: str) -> dict:
    """
    Excelファイルと同じフォルダに amazon/ と 1688/ サブフォルダを作成する。
    （Excelが商品フォルダ内に置かれていない旧形式でも対応）
    """
    try:
        base = Path(excel_path).parent
        (base / "amazon").mkdir(exist_ok=True)
        (base / "1688").mkdir(exist_ok=True)
        return {"success": True, "base": str(base)}
    except Exception as e:
        logger.warning(f"フォルダ作成エラー: {e}")
        return {"success": False, "error": str(e)}


def download_all_images(image_urls: list, save_dir: Path, asin: str) -> list:
    """
    複数の画像URLを {save_dir}/{asin}/ サブフォルダに 01.jpg, 02.jpg… として保存。
    保存に成功した Path のリストを返す。
    """
    if not image_urls:
        return []
    asin_dir = save_dir / asin
    asin_dir.mkdir(parents=True, exist_ok=True)
    saved = []
    for i, url in enumerate(image_urls, 1):
        if not url:
            continue
        try:
            ext = ".png" if ".png" in url.lower() else ".jpg"
            dest = asin_dir / f"{i:02d}{ext}"
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=15) as resp:
                dest.write_bytes(resp.read())
            logger.info(f"画像保存 [{i}]: {dest}")
            saved.append(dest)
        except Exception as e:
            logger.warning(f"画像ダウンロード失敗 [{i}] ({url[:60]}): {e}")
    return saved


# ── Excel情報取得 ───────────────────────────────────────────────────────
def get_excel_info(excel_path: str) -> dict:
    """Excel ファイルの現在の状態を返す。"""
    try:
        from openpyxl import load_workbook
        path = Path(excel_path)
        if not path.exists():
            return {"success": False, "error": f"ファイルが見つかりません: {excel_path}"}

        wb = load_workbook(str(path), read_only=True, data_only=True)
        sheets = list(wb.sheetnames)

        # Sheet1からタイトル取得
        title = ""
        if SHEET_OVERVIEW in sheets:
            ws = wb[SHEET_OVERVIEW]
            val = ws["A1"].value or ""
            title = str(val).replace("リサーチシート　", "").strip()

        # Sheet2の追記件数
        amazon_count = 0
        if SHEET_AMAZON_LIST in sheets:
            ws2 = wb[SHEET_AMAZON_LIST]
            for r in range(4, (ws2.max_row or 3) + 1):
                if ws2.cell(r, 1).value:
                    amazon_count += 1

        wb.close()
        return {
            "success":      True,
            "filename":     path.name,
            "title":        title,
            "sheets":       sheets,
            "amazon_count": amazon_count,
            "is_research":  SHEET_AMAZON_LIST in sheets,
        }
    except Exception as e:
        logger.error(f"Excel情報取得エラー: {e}")
        return {"success": False, "error": str(e)}


# ── Amazon追記 ─────────────────────────────────────────────────────────
def append_amazon(excel_path: str, data: dict) -> dict:
    """
    Sheet2（②Amazonライバル）と Sheet3（③Amazonテキスト）に
    Amazon データを追記して上書き保存する。
    画像は Excelと同じフォルダの _images/ に保存して埋め込む。
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        path = Path(excel_path)
        if not path.exists():
            return {"success": False, "error": f"ファイルが見つかりません: {excel_path}"}

        wb = load_workbook(str(path))

        if SHEET_AMAZON_LIST not in wb.sheetnames:
            return {
                "success": False,
                "error": f"シート「{SHEET_AMAZON_LIST}」がありません。"
                         "aucfanアプリのExcelボタンで作成したファイルを指定してください。"
            }

        thin = Side(style="thin", color="BFBFBF")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # ── Sheet2: ②Amazonライバル ──────────────────────────────
        ws2 = wb[SHEET_AMAZON_LIST]

        # 重複チェック: 同じASINが既に存在する場合はスキップ
        asin_to_add = data.get("asin", "")
        if asin_to_add:
            for r in range(4, (ws2.max_row or 3) + 1):
                if ws2.cell(r, 1).value == asin_to_add:
                    logger.warning(f"重複ASIN スキップ: {asin_to_add} (行{r})")
                    return {
                        "success": False,
                        "duplicate": True,
                        "asin": asin_to_add,
                        "row": r,
                        "error": f"ASIN {asin_to_add} は既に{r}行目に追記済みです。",
                    }

        # 4行目以降で最初の空行を探す
        next_row = 4
        for r in range(4, (ws2.max_row or 3) + 2):
            if ws2.cell(r, 1).value is None:
                next_row = r
                break

        ws2.row_dimensions[next_row].height = 90

        input_url = data.get("input_url", "") or ""
        resolved_url = data.get("url", "") or ""

        row_values = [
            data.get("asin", ""),
            data.get("title", ""),
            data.get("price", ""),
            data.get("rating", ""),
            data.get("review_count", ""),
            "あり" if data.get("has_aplus") else "なし",
            resolved_url,
            input_url,   # 入力URL（短縮URLそのまま）
            "",          # 画像列（後で埋め込み）
        ]
        for col, val in enumerate(row_values, 1):
            c = ws2.cell(next_row, col)
            c.value = val
            c.font = Font(name="BIZ UDGothic")
            c.border = border
            c.alignment = Alignment(
                vertical="center",
                wrap_text=(col == 2),
                horizontal="left" if col in (2, 7, 8) else "center"
            )

        # ── フォルダ確認（amazon/ 1688/ が存在しない旧形式にも対応） ──
        ensure_product_folders(str(path))

        # ── 画像ダウンロード & 埋め込み ──────────────────────────
        image_saved = False
        image_path = None
        img_dir = path.parent / IMAGE_FOLDER_NAME   # amazon/
        safe_asin = re.sub(r'[^\w]', '_', data.get("asin", "img"))

        # メイン画像（Excelに埋め込む用）
        if data.get("image_url"):
            image_path = download_image(data["image_url"], img_dir, safe_asin)

        # 全画像（カタログ参考用） → _images/{ASIN}/ サブフォルダに保存
        all_image_paths = []
        if data.get("image_urls"):
            all_image_paths = download_all_images(
                data["image_urls"], img_dir, safe_asin
            )
            logger.info(f"全画像保存完了: {len(all_image_paths)}枚 → {img_dir / safe_asin}")

        tmp_dir = tempfile.mkdtemp()
        try:
            if image_path and image_path.exists():
                try:
                    from PIL import Image as PILImage
                    img = PILImage.open(image_path).convert("RGBA")
                    img.thumbnail((100, 85), PILImage.LANCZOS)
                    bg = PILImage.new("RGB", img.size, (255, 255, 255))
                    bg.paste(img, mask=img.split()[3])
                    thumb = Path(tmp_dir) / image_path.name
                    bg.save(str(thumb), "JPEG", quality=85)

                    from openpyxl.drawing.image import Image as XLImage
                    xl_img = XLImage(str(thumb))
                    xl_img.anchor = f"I{next_row}"
                    ws2.add_image(xl_img)
                    ws2.cell(next_row, 9).value = None
                    image_saved = True
                except Exception as e:
                    logger.warning(f"画像埋め込みエラー: {e}")
                    ws2.cell(next_row, 9).value = data.get("image_url", "")

            # ── Sheet3: ③Amazonテキスト ──────────────────────────
            if SHEET_AMAZON_TEXT in wb.sheetnames:
                ws3 = wb[SHEET_AMAZON_TEXT]
                cur_max = ws3.max_row or 2
                start = cur_max + 2 if cur_max > 2 else 4

                def _write(r, label, value):
                    height = max(18, min(150, len(str(value)) // 4 + 18))
                    ws3.row_dimensions[r].height = height
                    c_label = ws3.cell(r, 1)
                    c_label.value = label
                    c_label.font = Font(name="BIZ UDGothic", bold=True)
                    c_label.fill = PatternFill("solid", fgColor="E2EFDA")
                    c_label.border = border
                    c_val = ws3.cell(r, 2)
                    c_val.value = value
                    c_val.font = Font(name="BIZ UDGothic")
                    c_val.alignment = Alignment(vertical="top", wrap_text=True)
                    c_val.border = border

                # セパレーター
                ws3.row_dimensions[start - 1].height = 6
                ws3.merge_cells(f"A{start}:B{start}")
                sep = ws3.cell(start, 1)
                sep.value = f"── {data.get('asin', '')}  {data.get('title', '')[:50]} ──"
                sep.font = Font(name="BIZ UDGothic", bold=True, color="FFFFFF")
                sep.fill = PatternFill("solid", fgColor="375623")
                sep.alignment = Alignment(horizontal="left", vertical="center")
                ws3.row_dimensions[start].height = 22

                _write(start + 1, "タイトル",   data.get("title", ""))
                _write(start + 2, "価格",        data.get("price", ""))
                _write(start + 3, "商品の特徴",  "\n".join(data.get("bullets", [])))
                _write(start + 4, "商品説明",    data.get("description", ""))
                specs_text = "\n".join(
                    f"{k}: {v}" for k, v in (data.get("specs") or {}).items()
                )
                _write(start + 5, "仕様・詳細",  specs_text)

            # 保存（tmpフォルダ削除前に必ず実行）
            wb.save(str(path))

        finally:
            shutil.rmtree(tmp_dir, ignore_errors=True)

        return {
            "success":     True,
            "row":         next_row,
            "asin":        data.get("asin", ""),
            "title":       data.get("title", ""),
            "price":       data.get("price", ""),
            "has_image":   image_saved,
            "image_count": len(all_image_paths),  # 保存した全画像枚数
        }

    except Exception as e:
        logger.error(f"Excel追記エラー: {e}", exc_info=True)
        return {"success": False, "error": str(e)}


# ── 1688追記 ────────────────────────────────────────────────────────────
def append_1688(excel_path: str, data: dict) -> dict:
    """
    Sheet4（④1688仕入れ）と Sheet5（⑤1688テキスト）に
    1688データを追記して上書き保存する。

    Sheet4 列構成 (A〜R / 18列):
      A=仕入れ選択（手入力）  B=ショップ名  C=ショップURL  D=信頼度  E=入驻年数
      F=商品名（中）  G=商品名（日・手入力）  H=バリアント（中）  I=バリアント（日・手入力）
      J=在庫数  K=単価(CNY)  L=MOQ  M=仕入総額(CNY)=K×L  N=仕入総額(JPY)=K×L×rate
      O=原価/個(JPY)=K×rate  P=利益(JPY)  Q=利益率  R=判定

    data キー:
      title, shop_name, shop_url, shop_rating, shop_repeat_rate, shop_years,
      min_price, moq, moq_unit, variants[{name, price, stock}],
      image_urls, attributes, url

    バリアントが複数ある場合は1バリアント1行で追記する。
    """
    SHEET_1688_LIST = "④1688仕入れ"
    SHEET_1688_TEXT = "⑤1688テキスト"

    try:
        import config
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        rate = getattr(config, "CNY_TO_JPY_RATE", 35)  # 元→円 換算係数

        path = Path(excel_path)
        if not path.exists():
            return {"success": False, "error": f"ファイルが見つかりません: {excel_path}"}

        wb = load_workbook(str(path))
        if SHEET_1688_LIST not in wb.sheetnames:
            return {"success": False,
                    "error": f"シート「{SHEET_1688_LIST}」がありません。"}

        # ── 列数チェック: 旧形式（18列未満）は拒否してユーザーに知らせる ──
        # Sheet4の構成: Row1=タイトル Row2=列ヘッダー Row3=説明文
        # → Row2を見て列数を確認する
        _ws4_check = wb[SHEET_1688_LIST]
        _header_row = [_ws4_check.cell(2, c).value for c in range(1, 20)]
        _col_count = sum(1 for v in _header_row if v is not None)
        if _col_count < 18:
            return {
                "success": False,
                "error": (
                    f"このExcelファイルは旧フォーマット（{_col_count}列）です。\n"
                    "aucfanアプリの「Excelを作成」ボタンで新しいファイルを作成してください。"
                )
            }

        thin   = Side(style="thin", color="BFBFBF")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # ── Sheet4: ④1688仕入れ ─────────────────────────────────────
        ws4 = wb[SHEET_1688_LIST]

        # 4行目以降の最初の空行を探す
        # ショップ名(B)ではなく単価(K列=11)で判定（ショップ名が取れない場合でも正しく動く）
        next_row = 4
        for r in range(4, (ws4.max_row or 3) + 2):
            if ws4.cell(r, 11).value is None:   # K列(単価CNY)で判定
                next_row = r
                break

        variants = data.get("variants") or [
            {"name": "デフォルト", "price": data.get("min_price", 0), "stock": 0}
        ]
        shop_url   = data.get("shop_url", "")
        shop_name  = data.get("shop_name", "")
        shop_years = data.get("shop_years", "")
        title      = data.get("title", "")
        title_ja   = data.get("title_ja", "")

        # 信頼度テキスト（評価/回頭率）
        trust = "/".join(filter(None, [
            data.get("shop_rating", ""),
            data.get("shop_repeat_rate", "")
        ]))

        written_rows = []
        for v in variants:
            price = float(v.get("price") or 0)
            moq   = int(data.get("moq") or 1)
            stock = int(v.get("stock") or 0)

            # A=仕入れ選択（空欄・手入力）
            ws4.cell(next_row,  1).value = ""
            # B=ショップ名
            ws4.cell(next_row,  2).value = shop_name
            # C=ショップURL
            ws4.cell(next_row,  3).value = shop_url
            # D=信頼度
            ws4.cell(next_row,  4).value = trust
            # E=入驻年数
            ws4.cell(next_row,  5).value = shop_years
            # F=商品名（中）
            ws4.cell(next_row,  6).value = title
            # G=商品名（日）翻訳済み（編集可）
            ws4.cell(next_row,  7).value = title_ja
            # H=バリアント（中）
            ws4.cell(next_row,  8).value = v.get("name", "")
            # I=バリアント（日）翻訳済み（編集可）
            ws4.cell(next_row,  9).value = v.get("name_ja", "")
            # J=在庫数
            ws4.cell(next_row, 10).value = stock if stock > 0 else ""
            # K=単価(CNY)
            ws4.cell(next_row, 11).value = price
            # L=係数(CNY→JPY) ← 常に書き込み・編集可
            ws4.cell(next_row, 12).value = rate
            # M=MOQ ← 常に書き込み（デフォルト1）
            ws4.cell(next_row, 13).value = moq

            if price > 0:
                r = next_row
                # N=仕入総額(CNY) = 単価×MOQ
                ws4.cell(r, 14).value = f"=K{r}*M{r}"
                # O=仕入総額(JPY) = 単価×MOQ×係数
                ws4.cell(r, 15).value = f"=K{r}*M{r}*L{r}"
                # P=原価/個(JPY) = 単価×係数
                ws4.cell(r, 16).value = f"=K{r}*L{r}"
                # Q=利益(JPY) = 販売価格 - FBA手数料 - 原価/個
                ws4.cell(r, 17).value = (
                    f'=IFERROR(Sheet1!B12-Sheet1!B13-P{r},"")'
                )
                # R=利益率
                ws4.cell(r, 18).value = (
                    f'=IFERROR(TEXT(Q{r}/Sheet1!B12,"0.0%"),"")'
                )
                # S=判定
                ws4.cell(r, 19).value = (
                    f'=IFERROR(IF(AND(VALUE(SUBSTITUTE(R{r},"%",""))/100>=0.25,'
                    f'Q{r}>=450),"◎ GO","× 再検討"),"")'
                )
            else:
                for col in range(14, 20):
                    ws4.cell(next_row, col).value = ""

            # ── 書式設定 ──
            # 左揃え列: B,C,D,E,F,G,H,I
            LEFT_COLS  = {2, 3, 4, 5, 6, 7, 8, 9}
            # URL列（ハイパーリンク書式）: C
            URL_COL    = 3
            # 折り返し列: F,G,H,I
            WRAP_COLS  = {6, 7, 8, 9}
            # 入力促進（薄い黄色背景）: A,G,I
            INPUT_COLS = {1, 7, 9}

            for col in range(1, 20):
                c = ws4.cell(next_row, col)
                is_url = (col == URL_COL)
                c.font = Font(
                    name="BIZ UDGothic",
                    color="0563C1" if is_url else "000000",
                    underline="single" if is_url else "none"
                )
                c.border = border
                c.alignment = Alignment(
                    vertical="center",
                    horizontal="left" if col in LEFT_COLS else "center",
                    wrap_text=(col in WRAP_COLS)
                )
                if col in INPUT_COLS:
                    c.fill = PatternFill("solid", fgColor="FFFDE7")  # 薄い黄

            # URLにハイパーリンク（C列）
            if shop_url:
                ws4.cell(next_row, 3).hyperlink = shop_url

            written_rows.append(next_row)
            ws4.row_dimensions[next_row].height = 34
            next_row += 1

        # ── Sheet5: ⑤1688テキスト ──────────────────────────────────
        if SHEET_1688_TEXT in wb.sheetnames:
            ws5 = wb[SHEET_1688_TEXT]
            cur_max = ws5.max_row or 2
            start5  = cur_max + 2 if cur_max > 2 else 4

            def _write5(r, label, value):
                height = max(18, min(120, len(str(value)) // 4 + 18))
                ws5.row_dimensions[r].height = height
                c_l = ws5.cell(r, 1)
                c_l.value = label
                c_l.font  = Font(name="BIZ UDGothic", bold=True)
                c_l.fill  = PatternFill("solid", fgColor="E2EFDA")
                c_l.border = border
                c_v = ws5.cell(r, 2)
                c_v.value = value
                c_v.font  = Font(name="BIZ UDGothic")
                c_v.alignment = Alignment(vertical="top", wrap_text=True)
                c_v.border = border

            # セパレーター
            ws5.merge_cells(f"A{start5}:B{start5}")
            sep = ws5.cell(start5, 1)
            sep.value = f"── {title[:60]} ──"
            sep.font  = Font(name="BIZ UDGothic", bold=True, color="FFFFFF")
            sep.fill  = PatternFill("solid", fgColor="375623")
            sep.alignment = Alignment(horizontal="left", vertical="center")
            ws5.row_dimensions[start5].height = 22

            _write5(start5 + 1, "商品名",      title)
            _write5(start5 + 2, "ショップ名",  shop_name)
            _write5(start5 + 3, "ショップURL", shop_url)
            _write5(start5 + 4, "入驻年数",    shop_years)
            _write5(start5 + 5, "信頼度",      trust)
            _write5(start5 + 6, "最低単価",    f"¥{data.get('min_price', 0)} 元")
            _write5(start5 + 7, "MOQ",
                    f"{data.get('moq', 1)}{data.get('moq_unit', '个')} 起批")

            # バリアント一覧
            variants_text = "\n".join(
                f"{v['name']}  ¥{v['price']}  在庫{v.get('stock', 0)}"
                for v in variants
            )
            _write5(start5 + 8, "バリアント", variants_text)

            # 商品属性
            attrs = data.get("attributes") or {}
            attrs_text = "\n".join(f"{k}: {v}" for k, v in attrs.items())
            _write5(start5 + 9, "商品属性",   attrs_text)

        # ── 画像ダウンロード & 保存 ──────────────────────────────────
        ensure_product_folders(str(path))
        img_dir = path.parent / "1688"
        image_urls = data.get("image_urls") or []
        saved_count = 0
        if image_urls:
            shop_key = re.sub(r'[^\w]', '_', shop_name[:20]) if shop_name else "shop"
            saved = download_all_images(image_urls, img_dir, shop_key)
            saved_count = len(saved)

        wb.save(str(path))

        return {
            "success":       True,
            "rows":          written_rows,
            "row":           written_rows[0] if written_rows else 0,
            "title":         title,
            "shop_name":     shop_name,
            "min_price":     data.get("min_price", 0),
            "variant_count": len(variants),
            "image_count":   saved_count,
        }

    except Exception as e:
        logger.error(f"1688 Excel追記エラー: {e}", exc_info=True)
        return {"success": False, "error": str(e)}
